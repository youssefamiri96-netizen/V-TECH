"""Gestione claim (reclami vettore) per V-Tech Trasporti.

Modulo AGGIUNTIVO e indipendente:
- crea tabelle proprie nello stesso database del gestionale (claims,
  claim_attachments, claim_events) e non modifica nulla di esistente;
- salva gli allegati (mail, foto, documenti) in DATA_DIR/claim_attachments,
  cioe' sul volume persistente Railway, come il database;
- genera l'export Excel da condividere con il cliente.
"""

from __future__ import annotations

import base64
import re
import sqlite3
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from vtech_app import DATA_DIR, DB_PATH, DOWNLOADS_DIR, clean_text, to_float

ATTACHMENTS_DIR = DATA_DIR / "claim_attachments"
MAX_ATTACHMENT_BYTES = 25 * 1024 * 1024

CLAIM_STATUSES = [
    "Aperto",
    "Inviato al vettore",
    "In valutazione",
    "Documenti richiesti",
    "Accettato",
    "Respinto",
    "Chiuso",
]
OPEN_STATUSES = {"Aperto", "Inviato al vettore", "In valutazione", "Documenti richiesti", "Accettato"}

CLAIM_REASON_GROUPS: dict[str, list[str]] = {
    "Trasporto": [
        "Merce danneggiata",
        "Merce mancante",
        "Consegna in ritardo",
        "Mancata consegna",
        "Errore di consegna",
        "Imballo non conforme",
        "Altro",
    ],
    "Magazzino": [
        "Errore di picking",
        "Colli mancanti al carico",
        "Quantita spedita errata",
        "Prodotto errato spedito",
        "Merce danneggiata in magazzino",
        "Bancale rotto o non conforme",
        "Etichetta errata o mancante",
        "Documento di trasporto errato",
        "Danno causato dal vettore in carico",
        "Vettore non presentato al carico",
        "Ritardo del vettore al carico",
        "Merce non ritirata",
        "Reso da cliente",
        "Altro magazzino",
    ],
}

CLAIM_REASONS = [reason for group in CLAIM_REASON_GROUPS.values() for reason in group]

CLAIM_ORIGINS = list(CLAIM_REASON_GROUPS.keys())
DEFAULT_ORIGIN = "Trasporto"


def claim_origin_for_reason(reason: Any) -> str:
    """Ricava il tipo di claim (Trasporto/Magazzino) dal motivo scelto."""
    raw = clean_text(reason)
    for origin, reasons in CLAIM_REASON_GROUPS.items():
        if raw in reasons:
            return origin
    return DEFAULT_ORIGIN

ATTACHMENT_KINDS = ["Mail", "Foto", "Documento", "Altro"]

# L'export contiene solo il foglio principale. Mettere True per riavere
# anche i fogli "Allegati" e "Storico stati".
EXPORT_DETAIL_SHEETS = False

# Due colonne note nell'export:
#   notes          -> "Note KN"    (le nostre)
#   customer_notes -> "Note Vtech" (quelle del cliente)

# Traduzioni per l'export in inglese da mandare al cliente.
STATUS_EN = {
    "Aperto": "Open",
    "Inviato al vettore": "Sent to carrier",
    "In valutazione": "Under review",
    "Documenti richiesti": "Documents requested",
    "Accettato": "Accepted",
    "Respinto": "Rejected",
    "Chiuso": "Closed",
}
REASON_EN = {
    "Merce danneggiata": "Damaged goods",
    "Merce mancante": "Missing goods",
    "Consegna in ritardo": "Late delivery",
    "Mancata consegna": "Failed delivery",
    "Errore di consegna": "Wrong delivery",
    "Imballo non conforme": "Non-compliant packaging",
    "Altro": "Other",
    "Errore di picking": "Picking error",
    "Colli mancanti al carico": "Missing parcels at loading",
    "Quantita spedita errata": "Wrong quantity shipped",
    "Prodotto errato spedito": "Wrong product shipped",
    "Merce danneggiata in magazzino": "Goods damaged in warehouse",
    "Bancale rotto o non conforme": "Broken or non-compliant pallet",
    "Etichetta errata o mancante": "Wrong or missing label",
    "Documento di trasporto errato": "Incorrect delivery note",
    "Danno causato dal vettore in carico": "Damage caused by carrier at loading",
    "Vettore non presentato al carico": "Carrier no-show at loading",
    "Ritardo del vettore al carico": "Carrier late at loading",
    "Merce non ritirata": "Goods not collected",
    "Reso da cliente": "Customer return",
    "Altro magazzino": "Other warehouse issue",
}
ORIGIN_EN = {"Trasporto": "Transport", "Magazzino": "Warehouse"}
KIND_EN = {"Mail": "Email", "Foto": "Photo", "Documento": "Document", "Altro": "Other"}

ALLOWED_ATTACHMENT_EXTENSIONS = {
    ".eml", ".msg", ".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic",
    ".bmp", ".tif", ".tiff", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt", ".zip",
}

HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="D5DDE5")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------
# Utilita'
# --------------------------------------------------------------------------
def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _safe_filename(value: Any, fallback: str = "allegato") -> str:
    raw = clean_text(value) or fallback
    raw = raw.replace("\\", "/").split("/")[-1]
    normalized = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^A-Za-z0-9._ -]+", "_", normalized).strip(" ._")
    return (cleaned or fallback)[:120]


def _decode_base64(value: Any) -> bytes:
    raw = clean_text(value)
    if not raw:
        raise ValueError("File allegato vuoto.")
    if "," in raw[:120] and raw[:5].lower() == "data:":
        raw = raw.split(",", 1)[1]
    try:
        content = base64.b64decode(raw, validate=False)
    except Exception as exc:  # pragma: no cover - difensivo
        raise ValueError("Allegato non leggibile.") from exc
    if not content:
        raise ValueError("File allegato vuoto.")
    if len(content) > MAX_ATTACHMENT_BYTES:
        raise ValueError(
            f"Allegato troppo grande ({len(content) / 1048576:.1f} MB). Massimo {MAX_ATTACHMENT_BYTES // 1048576} MB."
        )
    return content


def _normalized_status(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return CLAIM_STATUSES[0]
    for status in CLAIM_STATUSES:
        if status.lower() == text.lower():
            return status
    return text


def _row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


# --------------------------------------------------------------------------
# Database
# --------------------------------------------------------------------------
def init_claims_db(db_path: Path = DB_PATH) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS claims (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                claim_ref TEXT UNIQUE,
                shipment TEXT,
                orders_text TEXT,
                customer TEXT,
                province TEXT,
                carrier TEXT,
                shipment_date TEXT,
                reason TEXT,
                description TEXT,
                status TEXT,
                amount_claimed REAL,
                amount_settled REAL,
                carrier_ref TEXT,
                opened_at TEXT,
                updated_at TEXT,
                closed_at TEXT,
                created_by TEXT,
                notes TEXT
            )
            """
        )
        # Migrazione: colonna origin (tipo claim) aggiunta dopo la prima versione.
        existing = {row[1] for row in conn.execute("PRAGMA table_info(claims)").fetchall()}
        if "customer_notes" not in existing:
            conn.execute("ALTER TABLE claims ADD COLUMN customer_notes TEXT")
        if "origin" not in existing:
            conn.execute("ALTER TABLE claims ADD COLUMN origin TEXT")
            conn.execute("UPDATE claims SET origin = ? WHERE origin IS NULL OR origin = ''", (DEFAULT_ORIGIN,))
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS claim_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                claim_id INTEGER NOT NULL,
                kind TEXT,
                filename TEXT,
                stored_name TEXT,
                size_bytes INTEGER,
                uploaded_at TEXT,
                uploaded_by TEXT
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS claim_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                claim_id INTEGER NOT NULL,
                created_at TEXT,
                author TEXT,
                status TEXT,
                note TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_claim_att ON claim_attachments(claim_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_claim_ev ON claim_events(claim_id)")
    ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)


def _next_claim_ref(conn: sqlite3.Connection) -> str:
    year = datetime.now().year
    prefix = f"CLM-{year}-"
    cursor = conn.execute(
        "SELECT claim_ref FROM claims WHERE claim_ref LIKE ? ORDER BY claim_ref DESC LIMIT 1",
        (prefix + "%",),
    )
    last = cursor.fetchone()
    counter = 1
    if last and last[0]:
        try:
            counter = int(str(last[0]).rsplit("-", 1)[-1]) + 1
        except ValueError:
            counter = 1
    return f"{prefix}{counter:04d}"


def create_claim(payload: dict[str, Any], author: str = "", db_path: Path = DB_PATH) -> dict[str, Any]:
    shipment = clean_text(payload.get("shipment"))
    if not shipment:
        raise ValueError("Seleziona la spedizione a cui aprire il claim.")
    reason = clean_text(payload.get("reason"))
    if not reason:
        raise ValueError("Indica il motivo del claim.")

    init_claims_db(db_path)
    status = _normalized_status(payload.get("status"))
    stamp = _now()
    with sqlite3.connect(db_path) as conn:
        claim_ref = _next_claim_ref(conn)
        cursor = conn.execute(
            """
            INSERT INTO claims (
                claim_ref, shipment, orders_text, customer, province, carrier,
                shipment_date, reason, description, status, amount_claimed,
                amount_settled, carrier_ref, opened_at, updated_at, closed_at,
                created_by, notes, origin, customer_notes
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                claim_ref,
                shipment,
                clean_text(payload.get("ordersText")),
                clean_text(payload.get("customer")),
                clean_text(payload.get("province")),
                clean_text(payload.get("carrier")),
                clean_text(payload.get("shipmentDate")),
                reason,
                clean_text(payload.get("description")),
                status,
                to_float(payload.get("amountClaimed")),
                to_float(payload.get("amountSettled")),
                clean_text(payload.get("carrierRef")),
                stamp,
                stamp,
                stamp if status not in OPEN_STATUSES else "",
                clean_text(author),
                clean_text(payload.get("notes")),
                clean_text(payload.get("origin")) or claim_origin_for_reason(reason),
                clean_text(payload.get("customerNotes")),
            ),
        )
        claim_id = int(cursor.lastrowid)
        conn.execute(
            "INSERT INTO claim_events (claim_id, created_at, author, status, note) VALUES (?,?,?,?,?)",
            (claim_id, stamp, clean_text(author), status, "Claim aperto"),
        )
    return get_claim(claim_id, db_path)


UPDATABLE_FIELDS = {
    "reason": "reason",
    "description": "description",
    "status": "status",
    "amountClaimed": "amount_claimed",
    "amountSettled": "amount_settled",
    "carrierRef": "carrier_ref",
    "notes": "notes",
    "customerNotes": "customer_notes",
    "carrier": "carrier",
}


def update_claim(
    claim_id: int,
    payload: dict[str, Any],
    author: str = "",
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    init_claims_db(db_path)
    current = get_claim(claim_id, db_path)
    if current is None:
        raise ValueError("Claim non trovato.")

    assignments: list[str] = []
    values: list[Any] = []
    for key, column in UPDATABLE_FIELDS.items():
        if key not in payload:
            continue
        raw = payload.get(key)
        if column in {"amount_claimed", "amount_settled"}:
            values.append(to_float(raw))
        elif column == "status":
            values.append(_normalized_status(raw))
        else:
            values.append(clean_text(raw))
        assignments.append(f"{column} = ?")

    new_status = _normalized_status(payload.get("status")) if "status" in payload else current["status"]
    status_changed = "status" in payload and new_status != current["status"]
    stamp = _now()
    assignments.append("updated_at = ?")
    values.append(stamp)
    if status_changed:
        assignments.append("closed_at = ?")
        values.append(stamp if new_status not in OPEN_STATUSES else "")

    if "reason" in payload and clean_text(payload.get("reason")):
        assignments.append("origin = ?")
        values.append(claim_origin_for_reason(payload.get("reason")))

    note = clean_text(payload.get("eventNote"))
    with sqlite3.connect(db_path) as conn:
        if assignments:
            values.append(int(claim_id))
            conn.execute(f"UPDATE claims SET {', '.join(assignments)} WHERE id = ?", values)
        if status_changed or note:
            conn.execute(
                "INSERT INTO claim_events (claim_id, created_at, author, status, note) VALUES (?,?,?,?,?)",
                (
                    int(claim_id),
                    stamp,
                    clean_text(author),
                    new_status,
                    note or f"Stato aggiornato a {new_status}",
                ),
            )
    return get_claim(claim_id, db_path)


def delete_claim(claim_id: int, db_path: Path = DB_PATH) -> None:
    init_claims_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        stored = [
            row["stored_name"]
            for row in conn.execute(
                "SELECT stored_name FROM claim_attachments WHERE claim_id = ?", (int(claim_id),)
            ).fetchall()
        ]
        conn.execute("DELETE FROM claim_attachments WHERE claim_id = ?", (int(claim_id),))
        conn.execute("DELETE FROM claim_events WHERE claim_id = ?", (int(claim_id),))
        conn.execute("DELETE FROM claims WHERE id = ?", (int(claim_id),))
    for name in stored:
        if not name:
            continue
        target = ATTACHMENTS_DIR / name
        try:
            if target.exists():
                target.unlink()
        except OSError:
            pass


def get_claim(claim_id: int, db_path: Path = DB_PATH) -> dict[str, Any] | None:
    init_claims_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM claims WHERE id = ?", (int(claim_id),)).fetchone()
        if row is None:
            return None
        claim = _row_to_dict(row)
        claim["attachments"] = [
            _row_to_dict(item)
            for item in conn.execute(
                "SELECT * FROM claim_attachments WHERE claim_id = ? ORDER BY id DESC",
                (int(claim_id),),
            ).fetchall()
        ]
        claim["events"] = [
            _row_to_dict(item)
            for item in conn.execute(
                "SELECT * FROM claim_events WHERE claim_id = ? ORDER BY id DESC",
                (int(claim_id),),
            ).fetchall()
        ]
    return claim


def list_claims(db_path: Path = DB_PATH) -> list[dict[str, Any]]:
    init_claims_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        claims = [_row_to_dict(row) for row in conn.execute(
            "SELECT * FROM claims ORDER BY id DESC"
        ).fetchall()]
        attachments: dict[int, list[dict[str, Any]]] = {}
        for row in conn.execute("SELECT * FROM claim_attachments ORDER BY id DESC").fetchall():
            attachments.setdefault(int(row["claim_id"]), []).append(_row_to_dict(row))
        events: dict[int, list[dict[str, Any]]] = {}
        for row in conn.execute("SELECT * FROM claim_events ORDER BY id DESC").fetchall():
            events.setdefault(int(row["claim_id"]), []).append(_row_to_dict(row))
    for claim in claims:
        claim["attachments"] = attachments.get(int(claim["id"]), [])
        claim["events"] = events.get(int(claim["id"]), [])
    return claims


# --------------------------------------------------------------------------
# Allegati
# --------------------------------------------------------------------------
def add_claim_attachment(
    claim_id: int,
    filename: Any,
    content_base64: Any,
    kind: Any = "",
    author: str = "",
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    init_claims_db(db_path)
    claim = get_claim(claim_id, db_path)
    if claim is None:
        raise ValueError("Claim non trovato.")

    safe_name = _safe_filename(filename)
    extension = Path(safe_name).suffix.lower()
    if extension not in ALLOWED_ATTACHMENT_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_ATTACHMENT_EXTENSIONS))
        raise ValueError(f"Formato non ammesso ({extension or 'senza estensione'}). Ammessi: {allowed}")

    content = _decode_base64(content_base64)
    ATTACHMENTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp_compact = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    stored_name = f"claim{int(claim_id)}_{stamp_compact}_{safe_name}"
    (ATTACHMENTS_DIR / stored_name).write_bytes(content)

    kind_text = clean_text(kind) or ("Foto" if extension in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".bmp"} else
                                     "Mail" if extension in {".eml", ".msg"} else "Documento")
    stamp = _now()
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            INSERT INTO claim_attachments (claim_id, kind, filename, stored_name, size_bytes, uploaded_at, uploaded_by)
            VALUES (?,?,?,?,?,?,?)
            """,
            (int(claim_id), kind_text, safe_name, stored_name, len(content), stamp, clean_text(author)),
        )
        conn.execute("UPDATE claims SET updated_at = ? WHERE id = ?", (stamp, int(claim_id)))
    return get_claim(claim_id, db_path)


def delete_claim_attachment(attachment_id: int, db_path: Path = DB_PATH) -> dict[str, Any] | None:
    init_claims_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM claim_attachments WHERE id = ?", (int(attachment_id),)
        ).fetchone()
        if row is None:
            raise ValueError("Allegato non trovato.")
        claim_id = int(row["claim_id"])
        stored_name = row["stored_name"]
        conn.execute("DELETE FROM claim_attachments WHERE id = ?", (int(attachment_id),))
        conn.execute("UPDATE claims SET updated_at = ? WHERE id = ?", (_now(), claim_id))
    if stored_name:
        target = ATTACHMENTS_DIR / stored_name
        try:
            if target.exists():
                target.unlink()
        except OSError:
            pass
    return get_claim(claim_id, db_path)


def attachment_file(attachment_id: int, db_path: Path = DB_PATH) -> tuple[Path, str]:
    init_claims_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM claim_attachments WHERE id = ?", (int(attachment_id),)
        ).fetchone()
    if row is None:
        raise ValueError("Allegato non trovato.")
    target = ATTACHMENTS_DIR / str(row["stored_name"])
    if not target.exists():
        raise ValueError("File allegato non piu' disponibile sul server.")
    return target, str(row["filename"] or target.name)


# --------------------------------------------------------------------------
# Export Excel
# --------------------------------------------------------------------------
def _style_header(sheet, headers: list[str], row_index: int = 1) -> None:
    for column_index, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER


def _autosize(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _label(value: Any, mapping: dict[str, str], english: bool) -> str:
    """Restituisce l'etichetta tradotta se serve l'inglese, altrimenti l'originale."""
    raw = clean_text(value)
    if not english or not raw:
        return raw
    return mapping.get(raw, raw)


def _days_open(claim: dict[str, Any]) -> Any:
    opened = clean_text(claim.get("opened_at"))
    if not opened:
        return ""
    try:
        start = datetime.strptime(opened[:10], "%Y-%m-%d")
    except ValueError:
        return ""
    closed = clean_text(claim.get("closed_at"))
    end = datetime.now()
    if closed:
        try:
            end = datetime.strptime(closed[:10], "%Y-%m-%d")
        except ValueError:
            end = datetime.now()
    return max(0, (end - start).days)


def export_claims_report(
    db_path: Path = DB_PATH,
    downloads_dir: Path = DOWNLOADS_DIR,
    only_open: bool = False,
    language: str = "it",
) -> Path:
    """Genera l'Excel dei claim.

    language="it" -> versione interna (italiano)
    language="en" -> versione da mandare al cliente (inglese)
    """
    english = clean_text(language).lower() in {"en", "eng", "inglese", "english"}
    claims = list_claims(db_path)
    if only_open:
        claims = [claim for claim in claims if clean_text(claim.get("status")) in OPEN_STATUSES]
    if not claims:
        raise ValueError("Nessun claim da esportare.")

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Claims" if english else "Claim"
    sheet["A1"] = "V-Tech Trasporti - Claims register" if english else "V-Tech Trasporti - Registro claim"
    sheet["A1"].font = TITLE_FONT
    stamp_label = datetime.now().strftime("%d/%m/%Y %H:%M")
    sheet["A2"] = (
        f"Updated on {stamp_label} - {len(claims)} claims"
        if english
        else f"Aggiornato al {stamp_label} - {len(claims)} claim"
    )
    sheet["A2"].font = Font(italic=True, color="5B6B7B")

    # Nella versione per il cliente le "Note interne" NON vengono esportate.
    headers = [
        "Claim", "Type", "Shipment", "Orders", "Customer", "Province", "Carrier",
        "Shipment date", "Reason", "Description", "Status", "Carrier ref.",
        "Amount claimed", "Amount settled", "Opened on",
        "Last update", "Closed on", "Days", "Attachments", "Note KN", "Note Vtech",
    ] if english else [
        "Claim", "Tipo", "Spedizione", "Ordini", "Cliente", "Provincia", "Vettore",
        "Data spedizione", "Motivo", "Descrizione", "Stato", "Rif. vettore",
        "Importo richiesto", "Importo riconosciuto", "Aperto il",
        "Ultimo aggiornamento", "Chiuso il", "Giorni", "Allegati", "Note KN", "Note Vtech",
    ]
    _style_header(sheet, headers, row_index=4)
    for offset, claim in enumerate(claims, start=5):
        attachments = claim.get("attachments") or []
        values = [
            claim.get("claim_ref") or "",
            _label(claim.get("origin") or DEFAULT_ORIGIN, ORIGIN_EN, english),
            claim.get("shipment") or "",
            claim.get("orders_text") or "",
            claim.get("customer") or "",
            claim.get("province") or "",
            claim.get("carrier") or "",
            claim.get("shipment_date") or "",
            _label(claim.get("reason"), REASON_EN, english),
            claim.get("description") or "",
            _label(claim.get("status"), STATUS_EN, english),
            claim.get("carrier_ref") or "",
            claim.get("amount_claimed"),
            claim.get("amount_settled"),
            claim.get("opened_at") or "",
            claim.get("updated_at") or "",
            claim.get("closed_at") or "",
            _days_open(claim),
            len(attachments),
            claim.get("notes") or "",
            claim.get("customer_notes") or "",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=column_index, value=value)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=column_index in {4, 10, 20, 21})
            if column_index in {13, 14}:
                cell.number_format = '#,##0.00 "EUR"'
    _autosize(sheet, [14, 12, 14, 18, 26, 10, 12, 14, 20, 42, 18, 16, 16, 18, 18, 18, 18, 8, 9, 34, 34])
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, 4 + len(claims))}"

    if EXPORT_DETAIL_SHEETS:
        attachments_sheet = workbook.create_sheet("Attachments" if english else "Allegati")
        attachment_headers = (
            ["Claim", "Shipment", "Type", "File name", "Size KB", "Uploaded on", "Uploaded by"]
            if english
            else ["Claim", "Spedizione", "Tipo", "Nome file", "Dimensione KB", "Caricato il", "Caricato da"]
        )
        _style_header(attachments_sheet, attachment_headers)
        row_index = 2
        for claim in claims:
            for attachment in claim.get("attachments") or []:
                values = [
                    claim.get("claim_ref") or "",
                    claim.get("shipment") or "",
                    _label(attachment.get("kind"), KIND_EN, english),
                    attachment.get("filename") or "",
                    round((to_float(attachment.get("size_bytes")) or 0) / 1024, 1),
                    attachment.get("uploaded_at") or "",
                    attachment.get("uploaded_by") or "",
                ]
                for column_index, value in enumerate(values, start=1):
                    cell = attachments_sheet.cell(row=row_index, column=column_index, value=value)
                    cell.border = CELL_BORDER
                row_index += 1
        if row_index == 2:
            attachments_sheet.cell(
                row=2, column=1, value="No attachments uploaded." if english else "Nessun allegato caricato."
            )
        _autosize(attachments_sheet, [14, 14, 12, 46, 14, 20, 18])
        attachments_sheet.freeze_panes = "A2"

        history_sheet = workbook.create_sheet("Status history" if english else "Storico stati")
        history_headers = (
            ["Claim", "Shipment", "Date", "Status", "Note", "Author"]
            if english
            else ["Claim", "Spedizione", "Data", "Stato", "Nota", "Autore"]
        )
        _style_header(history_sheet, history_headers)
        row_index = 2
        for claim in claims:
            for event in sorted(claim.get("events") or [], key=lambda item: int(item.get("id") or 0)):
                values = [
                    claim.get("claim_ref") or "",
                    claim.get("shipment") or "",
                    event.get("created_at") or "",
                    _label(event.get("status"), STATUS_EN, english),
                    event.get("note") or "",
                    event.get("author") or "",
                ]
                for column_index, value in enumerate(values, start=1):
                    cell = history_sheet.cell(row=row_index, column=column_index, value=value)
                    cell.border = CELL_BORDER
                    cell.alignment = Alignment(vertical="top", wrap_text=column_index == 5)
                row_index += 1
        if row_index == 2:
            history_sheet.cell(
                row=2, column=1, value="No updates recorded." if english else "Nessun aggiornamento registrato."
            )
        _autosize(history_sheet, [14, 14, 20, 20, 50, 18])
        history_sheet.freeze_panes = "A2"


    downloads_dir.mkdir(parents=True, exist_ok=True)
    prefix = "VTech_claims_EN" if english else "VTech_claim"
    output_path = downloads_dir / f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(output_path)
    return output_path


ECONOMIC_FIELD_HINTS = (
    "costo", "margine", "margin", "extra", "tariffa", "attivo", "passivo",
    "prezzo", "importo", "nolo", "fuel",
)
ECONOMIC_FIELD_EXACT = {"gp", "gp%", "gp %"}


def _is_economic_field(key: Any) -> bool:
    text = clean_text(key).lower()
    if not text:
        return False
    if text in ECONOMIC_FIELD_EXACT:
        return True
    return any(hint in text for hint in ECONOMIC_FIELD_HINTS)


def strip_economics_for_claims(payload: dict[str, Any]) -> dict[str, Any]:
    """Toglie i valori economici dal payload delle spedizioni.

    Serve per gli utenti con ruolo 'claims' (magazzino): devono poter cercare
    e selezionare le spedizioni per aprire un claim, ma non devono vedere
    costi attivi/passivi ne' margini. I campi restano presenti ma vuoti, cosi'
    l'interfaccia non si rompe.
    """
    if not isinstance(payload, dict):
        return payload

    for group_rows in (payload.get("groups") or {}).values():
        if not isinstance(group_rows, list):
            continue
        for row in group_rows:
            for bucket in ("display", "raw"):
                data = row.get(bucket)
                if isinstance(data, dict):
                    for key in list(data.keys()):
                        if _is_economic_field(key):
                            data[key] = ""

    stats = payload.get("stats")
    if isinstance(stats, dict):
        for key in list(stats.keys()):
            if _is_economic_field(key):
                stats[key] = ""

    kpis = payload.get("kpis")
    if isinstance(kpis, dict):
        for key in list(kpis.keys()):
            if _is_economic_field(key):
                kpis[key] = 0

    # ATTENZIONE: le chiavi vanno svuotate, NON rimosse: app.js legge
    # data.paths.vtech in fase di avvio e senza la chiave si blocca.
    if isinstance(payload.get("paths"), dict):
        payload["paths"] = {key: "" for key in payload["paths"]}
    if "billingColumns" in payload:
        payload["billingColumns"] = []

    return payload


def claims_payload(db_path: Path = DB_PATH) -> dict[str, Any]:
    """Payload completo per il frontend."""
    claims = list_claims(db_path)
    open_count = sum(1 for claim in claims if clean_text(claim.get("status")) in OPEN_STATUSES)
    return {
        "claims": claims,
        "statuses": CLAIM_STATUSES,
        "reasons": CLAIM_REASONS,
        "reasonGroups": CLAIM_REASON_GROUPS,
        "origins": CLAIM_ORIGINS,
        "attachmentKinds": ATTACHMENT_KINDS,
        "openCount": open_count,
        "total": len(claims),
    }
