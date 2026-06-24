from __future__ import annotations

import csv
import json
import math
import os
import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import pypdf
from openpyxl import load_workbook


PROVINCE_TO_REGION = {
    "AG": "SICILIA",
    "AL": "PIEMONTE",
    "AN": "MARCHE",
    "AO": "VAL D'AOSTA",
    "AP": "MARCHE",
    "AQ": "ABRUZZO",
    "AR": "TOSCANA",
    "AT": "PIEMONTE",
    "AV": "CAMPANIA",
    "BA": "PUGLIA",
    "BG": "LOMBARDIA",
    "BI": "PIEMONTE",
    "BL": "VENETO",
    "BN": "CAMPANIA",
    "BO": "EMILIA ROMAGNA",
    "BR": "PUGLIA",
    "BS": "LOMBARDIA",
    "BT": "PUGLIA",
    "BZ": "TRENTINO A.A.",
    "CA": "SARDEGNA",
    "CB": "MOLISE",
    "CE": "CAMPANIA",
    "CH": "ABRUZZO",
    "CI": "SARDEGNA",
    "CL": "SICILIA",
    "CN": "PIEMONTE",
    "CO": "LOMBARDIA",
    "CR": "LOMBARDIA",
    "CS": "CALABRIA",
    "CT": "SICILIA",
    "CZ": "CALABRIA",
    "EN": "SICILIA",
    "FC": "EMILIA ROMAGNA",
    "FE": "EMILIA ROMAGNA",
    "FG": "PUGLIA",
    "FI": "TOSCANA",
    "FM": "MARCHE",
    "FR": "LAZIO",
    "GE": "LIGURIA",
    "GO": "FRIULI VG",
    "GR": "TOSCANA",
    "IM": "LIGURIA",
    "IS": "MOLISE",
    "KR": "CALABRIA",
    "LC": "LOMBARDIA",
    "LE": "PUGLIA",
    "LI": "TOSCANA",
    "LO": "LOMBARDIA",
    "LT": "LAZIO",
    "LU": "TOSCANA",
    "MB": "LOMBARDIA",
    "MC": "MARCHE",
    "ME": "SICILIA",
    "MI": "LOMBARDIA",
    "MN": "LOMBARDIA",
    "MO": "EMILIA ROMAGNA",
    "MS": "TOSCANA",
    "MT": "BASILICATA",
    "NA": "CAMPANIA",
    "NO": "PIEMONTE",
    "NU": "SARDEGNA",
    "OG": "SARDEGNA",
    "OR": "SARDEGNA",
    "OT": "SARDEGNA",
    "PA": "SICILIA",
    "PC": "EMILIA ROMAGNA",
    "PD": "VENETO",
    "PE": "ABRUZZO",
    "PG": "UMBRIA",
    "PI": "TOSCANA",
    "PN": "FRIULI VG",
    "PO": "TOSCANA",
    "PR": "EMILIA ROMAGNA",
    "PT": "TOSCANA",
    "PU": "MARCHE",
    "PV": "LOMBARDIA",
    "PZ": "BASILICATA",
    "RA": "EMILIA ROMAGNA",
    "RC": "CALABRIA",
    "RE": "EMILIA ROMAGNA",
    "RG": "SICILIA",
    "RI": "LAZIO",
    "RM": "LAZIO",
    "RN": "EMILIA ROMAGNA",
    "RO": "VENETO",
    "SA": "CAMPANIA",
    "SI": "TOSCANA",
    "SO": "LOMBARDIA",
    "SP": "LIGURIA",
    "SR": "SICILIA",
    "SS": "SARDEGNA",
    "SU": "SARDEGNA",
    "SV": "LIGURIA",
    "TA": "PUGLIA",
    "TE": "ABRUZZO",
    "TN": "TRENTINO A.A.",
    "TO": "PIEMONTE",
    "TP": "SICILIA",
    "TR": "UMBRIA",
    "TS": "FRIULI VG",
    "TV": "VENETO",
    "UD": "FRIULI VG",
    "VA": "LOMBARDIA",
    "VB": "PIEMONTE",
    "VC": "PIEMONTE",
    "VE": "VENETO",
    "VI": "VENETO",
    "VR": "VENETO",
    "VS": "SARDEGNA",
    "VT": "LAZIO",
    "VV": "CALABRIA",
}

BRT_REGIONS = set(PROVINCE_TO_REGION.values())
BRT_VOLUMETRIC_KG_PER_M3 = 250.0
BANCALI_EPAL_WEIGHT_KG = 25.0
BANCALI_A_PERDERE_WEIGHT_KG = 15.0
BANCALI_REPORT_BASE_TARE_KG = 10.0
BANCALI_EPAL_EXTRA_WEIGHT_KG = BANCALI_EPAL_WEIGHT_KG - BANCALI_REPORT_BASE_TARE_KG
BANCALI_A_PERDERE_EXTRA_WEIGHT_KG = BANCALI_A_PERDERE_WEIGHT_KG - BANCALI_REPORT_BASE_TARE_KG
BANCALI_SLOT_VOLUME_M3 = 1.20 * 0.80 * 1.20
BANCALI_DECIMAL_EPSILON = 0.001
ACTIVE_ETS_REGIONS = {"SICILIA", "SARDEGNA"}
SLA_CUTOFF = time(12, 30)
SLA_PREP_WITHIN_CUTOFF_HOURS = 48
SLA_PREP_AFTER_CUTOFF_HOURS = 72
DEFAULT_PASSIVE_FUEL_RATE = 0.02
DEFAULT_CARRIER_TARIFFS_PATH = Path(__file__).resolve().parent / "data" / "carrier_tariffs.csv"

BRT_EXTRA_FLAG_COLUMNS = [
    "Shipment",
    "Isole Minori",
    "Dirottamento",
    "Localita Disagiata",
    "Consegna Disagiata",
    "Supermercati GDO",
    "Priority",
    "Servizio 10:30",
    "Contrassegno Valore",
    "ZTL",
    "Fuori Misura",
    "Appuntamento",
    "POD Image",
    "Ricerca Documenti",
    "ORM Commissionato",
    "Recapito Contrassegni",
    "Bancali Rendere",
    "Giacenza Dossier",
    "Riconsegna Giacenza",
]

TRUE_VALUES = {"1", "SI", "S", "YES", "Y", "TRUE", "X"}
LEGAL_SUFFIX_TOKENS = {
    "SRL",
    "Srl",
    "SPA",
    "S",
    "R",
    "L",
    "P",
    "A",
    "SOCIETA",
    "RESPONSABILITA",
    "LIMITATA",
    "UNIPERSONALE",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_registry_text(value: Any) -> str:
    text = clean_text(value).upper()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact_registry_text(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize_registry_text(value))


def customer_match_keys(value: Any) -> set[str]:
    normalized = normalize_registry_text(value)
    if not normalized:
        return set()
    tokens = [token for token in normalized.split() if token not in LEGAL_SUFFIX_TOKENS]
    stripped = " ".join(tokens)
    keys = {normalized, compact_registry_text(normalized)}
    if stripped:
        keys.add(stripped)
        keys.add(compact_registry_text(stripped))
    return {key for key in keys if key}


def iter_registry_values(value: Any) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    parts = [
        part.strip()
        for part in re.split(r"\s*(?:\||\n|\r)\s*", text)
        if part.strip()
    ]
    return [text, *parts] if text not in parts else parts


def default_carrier_tariffs_path() -> Path | None:
    env_data_dir = clean_text(os.environ.get("VTECH_DATA_DIR"))
    candidates = []
    if env_data_dir:
        candidates.append(Path(env_data_dir) / "carrier_tariffs.csv")
    candidates.append(DEFAULT_CARRIER_TARIFFS_PATH)
    for path in candidates:
        if path.exists():
            return path
    return None


@dataclass(frozen=True)
class GdoCustomerMatch:
    label: str
    active_tail_lift: bool = False


@dataclass(frozen=True)
class GdoCustomerRegistry:
    customer_keys: dict[str, GdoCustomerMatch]
    address_keys: dict[str, GdoCustomerMatch]


def load_gdo_customer_registry(path: Path | None) -> GdoCustomerRegistry | None:
    if not path or not path.exists():
        return None

    customer_keys: dict[str, GdoCustomerMatch] = {}
    address_keys: dict[str, GdoCustomerMatch] = {}
    customer_columns = [
        "Ragione Sociale",
        "RAGIONESOCIALE",
        "Cliente",
        "Route to Customer",
        "Customer",
    ]
    address_columns = [
        "Indirizzo Consegna",
        "DESTINAZIONEMERCE",
        "Route To Address",
        "Address",
    ]

    with path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            enabled = clean_text(row.get("GDO") or row.get("gdo") or "SI").upper()
            if enabled and enabled not in TRUE_VALUES:
                continue

            customer = next((clean_text(row.get(column)) for column in customer_columns if clean_text(row.get(column))), "")
            address = next((clean_text(row.get(column)) for column in address_columns if clean_text(row.get(column))), "")
            label = customer or address
            notes = " | ".join(
                clean_text(row.get(column))
                for column in ("Note Scarico", "NOTEPERLOSCARICO", "Shipping Information", "SHIPPINGINFORMATION")
                if clean_text(row.get(column))
            )
            match = GdoCustomerMatch(
                label=label,
                active_tail_lift="SPONDA" in normalize_registry_text(notes),
            )

            def add_match(index: dict[str, GdoCustomerMatch], key: str) -> None:
                existing = index.get(key)
                if existing is None or (match.active_tail_lift and not existing.active_tail_lift):
                    index[key] = match

            for key in customer_match_keys(customer):
                add_match(customer_keys, key)
            for key in customer_match_keys(address):
                add_match(address_keys, key)

    return GdoCustomerRegistry(customer_keys=customer_keys, address_keys=address_keys)


def match_gdo_customer(
    shipment_row: dict[str, Any],
    registry: GdoCustomerRegistry | None,
) -> GdoCustomerMatch | None:
    if not registry:
        return None

    customer_values = [
        shipment_row.get("Route to Customer"),
        shipment_row.get("Customer"),
        shipment_row.get("Cliente"),
    ]
    for value in customer_values:
        for candidate in iter_registry_values(value):
            for key in customer_match_keys(candidate):
                if key in registry.customer_keys:
                    return registry.customer_keys[key]

    address_values = [
        shipment_row.get("Route To Address"),
        shipment_row.get("Indirizzo Consegna"),
        shipment_row.get("Address"),
    ]
    for value in address_values:
        for candidate in iter_registry_values(value):
            for key in customer_match_keys(candidate):
                if key in registry.address_keys:
                    return registry.address_keys[key]

    return None


def classify_gdo_customer(
    shipment_row: dict[str, Any],
    registry: GdoCustomerRegistry | None,
) -> str:
    match = match_gdo_customer(shipment_row, registry)
    return match.label if match else ""


def to_float(value: Any) -> float | None:
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    text = clean_text(value).replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_euro(value: str) -> float:
    return float(value.replace(".", "").replace(",", "."))


def parse_hours(value: Any) -> float | None:
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, int | float):
        return float(value)
    match = re.search(r"\d+(?:[,.]\d+)?", clean_text(value))
    if not match:
        return None
    return float(match.group(0).replace(",", "."))


def parse_datetime_value(value: Any) -> datetime | None:
    if value is None or clean_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)

    text = clean_text(value)
    text = text.replace("T", " ")
    text = re.sub(
        r"(\d{1,2}/\d{1,2}/\d{2,4})\s+(\d{1,2})\.(\d{2})\.(\d{2})",
        r"\1 \2:\3:\4",
        text,
    )
    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d/%m/%y %H:%M:%S",
        "%d/%m/%y %H:%M",
        "%d/%m/%y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_sla_date(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d") if value else ""


def next_business_moment(value: datetime) -> datetime:
    if value.weekday() < 5:
        return value
    days_to_monday = 7 - value.weekday()
    return datetime.combine(value.date() + timedelta(days=days_to_monday), time.min)


def add_business_hours(start_at: datetime, hours: float) -> datetime:
    current = next_business_moment(start_at)
    remaining_hours = max(0.0, float(hours))

    while remaining_hours > 0:
        current = next_business_moment(current)
        next_midnight = datetime.combine(current.date() + timedelta(days=1), time.min)
        available_hours = (next_midnight - current).total_seconds() / 3600
        if remaining_hours <= available_hours:
            return current + timedelta(hours=remaining_hours)
        remaining_hours -= available_hours
        current = next_midnight

    return current


def normalize_region(value: str) -> str:
    cleaned = re.sub(r"\s+", " ", value.replace(".", " ")).strip().upper()
    cleaned = cleaned.replace("FRIULI V G", "FRIULI VG")
    cleaned = cleaned.replace("TRENTINO A A", "TRENTINO A.A.")
    cleaned = cleaned.replace("EMILIA ROMAGNA", "EMILIA ROMAGNA")
    return cleaned


def qli(weight_kg: float) -> int:
    return max(1, math.ceil(weight_kg / 100))


def brt_base_billable_weight(weight_kg: float) -> float:
    """Round BRT base transport weight as stated in the passive tariff PDF."""
    if weight_kg <= 100:
        return float(max(1, math.ceil(weight_kg)))
    return float(math.ceil(weight_kg / 100) * 100)


def capped(value: float, minimum: float | None = None, maximum: float | None = None) -> float:
    result = value
    if minimum is not None:
        result = max(minimum, result)
    if maximum is not None:
        result = min(maximum, result)
    return round(result, 2)


def brt_taxable_weight(weight_kg: Any, volume_m3: Any = None) -> tuple[float, float | None, float | None, float | None] | None:
    actual_weight = to_float(weight_kg)
    volume = to_float(volume_m3)
    volumetric_weight = volume * BRT_VOLUMETRIC_KG_PER_M3 if volume is not None else None
    candidates = [
        value
        for value in (actual_weight, volumetric_weight)
        if value is not None and value > 0
    ]
    if not candidates:
        return None
    return max(1.0, max(candidates)), actual_weight, volume, volumetric_weight


def parse_month_key_from_row(row: dict[str, Any] | None) -> str:
    if not row:
        return ""
    for column in ("Late Ship Date", "Data Consegna", "Early Delivery Date", "Integration Date"):
        parsed = parse_datetime_value(row.get(column))
        if parsed:
            return f"{parsed.year:04d}-{parsed.month:02d}"
    return ""


def parse_fuel_rate(value: Any) -> float | None:
    if value is None or clean_text(value) == "":
        return None
    number = to_float(clean_text(value).replace("%", ""))
    if number is None:
        return None
    if number < 0:
        number = 0
    return number / 100


def first_decimal_number(text: str) -> float | None:
    match = re.search(r"\d+(?:[.,]\d+)?", text)
    if not match:
        return None
    return to_float(match.group(0))


def first_percent_value(text: str) -> float | None:
    match = re.search(r"([+-]?\d+(?:[.,]\d+)?)\s*%", text)
    if not match:
        return None
    number = to_float(match.group(1))
    if number is None:
        return None
    return max(0.0, number / 100)


def parse_active_extra_settings(wb: Any) -> ActiveExtraSettings:
    settings = ActiveExtraSettings()
    if "Additional extra charges" not in wb.sheetnames:
        return settings
    ws = wb["Additional extra charges"]
    for row in ws.iter_rows(values_only=True):
        text = clean_text(row[0] if row else "")
        if not text:
            continue
        normalized = text.lower()
        amount = first_decimal_number(text)
        percent = first_percent_value(text)
        if "tail lift" in normalized and amount is not None:
            settings.tail_lift_groupage = amount
        elif "fixed time slot" in normalized and "for amazon" in normalized:
            if percent is not None:
                settings.fixed_time_slot_percent = percent
            amazon_match = re.search(r"amazon\s+([+-]?\d+(?:[.,]\d+)?)\s*%", normalized)
            if amazon_match:
                amazon_percent = to_float(amazon_match.group(1))
                if amazon_percent is not None:
                    settings.amazon_fixed_time_slot_percent = max(0.0, amazon_percent / 100)
        elif "phone preadvise" in normalized and amount is not None:
            settings.phone_preadvise = amount
        elif "waiting time" in normalized and amount is not None:
            settings.waiting_time_hour = amount
        elif "ets" in normalized and percent is not None:
            settings.ets_sicily_sardinia_percent = percent
        elif "fix day delivery" in normalized and "amazon" in normalized and percent is not None:
            settings.amazon_bkl_percent = percent
        elif "remote area" in normalized and percent is not None:
            settings.remote_area_percent = percent
        elif "gdo delivery" in normalized and percent is not None:
            settings.gdo_percent = percent
        elif "urgent delivery" in normalized and amount is not None:
            settings.urgent_per_pallet = amount
        elif "2nd delivery" in normalized and percent is not None:
            settings.second_delivery_returns_percent = percent
    return settings


def load_fuel_settings(path: Path | None) -> dict[str, dict[str, float]]:
    if not path or not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    settings: dict[str, dict[str, float]] = {}
    for month, values in raw.items():
        if not re.fullmatch(r"\d{4}-\d{2}", clean_text(month)):
            continue
        if not isinstance(values, dict):
            continue
        month_settings: dict[str, float] = {}
        active_rate = parse_fuel_rate(values.get("active"))
        passive_rate = parse_fuel_rate(values.get("passive"))
        if active_rate is not None:
            month_settings["active"] = active_rate
        if passive_rate is not None:
            month_settings["passive"] = passive_rate
        if month_settings:
            settings[clean_text(month)] = month_settings
    return settings


def fuel_settings_for_row(
    row: dict[str, Any] | None,
    settings: dict[str, dict[str, float]],
) -> tuple[float | None, float]:
    month_key = parse_month_key_from_row(row)
    month_settings = settings.get(month_key, {})
    active_rate = month_settings.get("active")
    passive_rate = month_settings.get("passive", DEFAULT_PASSIVE_FUEL_RATE)
    return active_rate, passive_rate


def format_percent(rate: float) -> str:
    value = rate * 100
    if abs(value - round(value)) < 0.001:
        return f"{value:.0f}%"
    text = f"{value:.2f}".rstrip("0").rstrip(".")
    return f"{text}%"


def has_freight_code(row: dict[str, Any] | None, code: str) -> bool:
    if not row:
        return False
    expected = code.upper()
    haystack = " | ".join(
        clean_text(row.get(column)).upper()
        for column in (
            "Freight Code",
            "Prenotazione Scarico",
            "Note Text",
            "Wave",
        )
        if clean_text(row.get(column))
    )
    if expected in re.split(r"[^A-Z0-9]+", haystack):
        return True
    if expected == "BKV" and "PRENOTATA DAL CLIENTE" in haystack:
        return True
    return False


def is_amazon_customer(row: dict[str, Any] | None) -> bool:
    if not row:
        return False
    return "AMAZON" in clean_text(row.get("Route to Customer")).upper()


def requires_epal_pallets(row: dict[str, Any] | None) -> bool:
    if is_amazon_customer(row):
        return True
    if not row:
        return False
    notes = clean_text(row.get("Note Text")).upper()
    return "EPAL" in notes and not re.search(r"\b(NO|NON)\s+EPAL\b", notes)


def pallet_support_breakdown(row: dict[str, Any] | None) -> tuple[float, float, int, int, str]:
    if not row:
        return 0.0, 0.0, 0, 0, ""
    pallets = to_float(row.get("Theoretical Pallets"))
    if pallets is None:
        pallets = to_float(row.get("Pallet Fatturati"))
    if pallets is None or pallets <= 0:
        return 0.0, 0.0, 0, 0, ""

    pallets = max(0.0, pallets)
    if requires_epal_pallets(row):
        epal_count = max(1, math.ceil(pallets - BANCALI_DECIMAL_EPSILON))
        disposable_count = 0
        reason = "Amazon/EPAL richiesto"
    else:
        rounded = round(pallets)
        if abs(pallets - rounded) < BANCALI_DECIMAL_EPSILON:
            full_pallets = int(rounded)
            has_fraction = False
        else:
            full_pallets = int(math.floor(pallets))
            has_fraction = True
        epal_count = max(0, full_pallets)
        disposable_count = 1 if has_fraction else 0
        reason = "Full pallet EPAL + frazione a perdere" if disposable_count else "Full pallet EPAL"

    pallet_weight = (epal_count * BANCALI_EPAL_EXTRA_WEIGHT_KG) + (
        disposable_count * BANCALI_A_PERDERE_EXTRA_WEIGHT_KG
    )
    pallet_volume = (epal_count + disposable_count) * BANCALI_SLOT_VOLUME_M3
    return pallet_weight, pallet_volume, epal_count, disposable_count, reason


def apply_pallet_weight_to_shipment(row: dict[str, Any]) -> None:
    goods_weight = to_float(row.get("Peso Merce Kg"))
    current_weight = to_float(row.get("Grand Total Shipment Ftp Wgt Kg"))
    if goods_weight is None:
        previous_pallet_weight = to_float(row.get("Peso Bancali Kg"))
        if current_weight is not None and previous_pallet_weight is not None:
            goods_weight = max(0.0, current_weight - previous_pallet_weight)
        else:
            goods_weight = current_weight

    goods_volume = to_float(row.get("Volume Merce m3"))
    current_volume = to_float(row.get("Grand Total Shipment Ftp Vol m3"))
    if goods_volume is None:
        previous_pallet_volume = to_float(row.get("Volume Bancali m3"))
        if current_volume is not None and previous_pallet_volume is not None:
            goods_volume = max(0.0, current_volume - previous_pallet_volume)
        else:
            goods_volume = current_volume

    if goods_weight is None and goods_volume is None:
        row["Peso Merce Kg"] = ""
        row["Peso Bancali Kg"] = ""
        row["Peso Totale con Bancali Kg"] = ""
        row["Volume Merce m3"] = ""
        row["Volume Bancali m3"] = ""
        row["Volume Totale con Bancali m3"] = ""
        row["Dettaglio Peso Bancali"] = ""
        return

    pallet_weight, pallet_volume, epal_count, disposable_count, reason = pallet_support_breakdown(row)
    if goods_weight is not None:
        total_weight = round(goods_weight + pallet_weight, 3)
        row["Peso Merce Kg"] = round(goods_weight, 3)
        row["Peso Bancali Kg"] = round(pallet_weight, 3)
        row["Peso Totale con Bancali Kg"] = total_weight
        row["Grand Total Shipment Ftp Wgt Kg"] = total_weight
    else:
        row["Peso Merce Kg"] = ""
        row["Peso Bancali Kg"] = ""
        row["Peso Totale con Bancali Kg"] = ""

    if goods_volume is not None:
        total_volume = round(max(goods_volume, pallet_volume), 3)
        row["Volume Merce m3"] = round(goods_volume, 3)
        row["Volume Bancali m3"] = round(pallet_volume, 3)
        row["Volume Totale con Bancali m3"] = total_volume
        row["Grand Total Shipment Ftp Vol m3"] = total_volume
    else:
        row["Volume Merce m3"] = ""
        row["Volume Bancali m3"] = ""
        row["Volume Totale con Bancali m3"] = ""

    details: list[str] = []
    if epal_count:
        details.append(
            f"{epal_count} EPAL: +{BANCALI_EPAL_EXTRA_WEIGHT_KG:.0f} kg cad. / {BANCALI_SLOT_VOLUME_M3:.3f} m3 cad."
        )
    if disposable_count:
        details.append(
            f"{disposable_count} a perdere: +{BANCALI_A_PERDERE_EXTRA_WEIGHT_KG:.0f} kg cad. / {BANCALI_SLOT_VOLUME_M3:.3f} m3 cad."
        )
    row["Dettaglio Peso Bancali"] = f"{' + '.join(details)} ({reason})" if details else ""


def is_gdo_customer(row: dict[str, Any] | None) -> bool:
    if not row:
        return False
    return clean_text(row.get("Cliente GDO")).upper() in TRUE_VALUES or clean_text(
        row.get("Supermercati GDO")
    ).upper() in TRUE_VALUES


def is_active_urgent(row: dict[str, Any] | None) -> bool:
    if not row:
        return False
    return clean_text(row.get("Attiva Urgente")).upper() in TRUE_VALUES


def is_groupage_active_shipment(row: dict[str, Any] | None) -> bool:
    if not row:
        return False
    manual_level = clean_text(row.get("Service Level Manuale")).upper()
    original_carrier = clean_text(row.get("Carrier Originale")).upper()
    if manual_level == "LTL" and original_carrier and original_carrier != "BRT":
        return False
    return clean_text(row.get("Tipo Servizio")) == "Groupage - BRT LTL"


def requires_active_tail_lift(row: dict[str, Any] | None, is_gdo: bool, is_amazon: bool) -> bool:
    if not is_groupage_active_shipment(row) or is_amazon:
        return False
    if not is_gdo:
        return True
    return clean_text((row or {}).get("Attiva Sponda")).upper() in TRUE_VALUES


def row_has_any_true(row: dict[str, Any] | None, keys: tuple[str, ...]) -> bool:
    if not row:
        return False
    return any(clean_text(row.get(key)).upper() in TRUE_VALUES for key in keys)


@dataclass
class ActiveExtraSettings:
    tail_lift_groupage: float = 10.00
    fixed_time_slot_percent: float = 0.20
    amazon_fixed_time_slot_percent: float = 0.40
    phone_preadvise: float = 5.00
    waiting_time_hour: float = 0.0
    ets_sicily_sardinia_percent: float = 0.05
    amazon_bkl_percent: float = 0.40
    remote_area_percent: float = 0.40
    gdo_percent: float = 0.20
    urgent_per_pallet: float = 15.00
    second_delivery_returns_percent: float = 1.00


DEFAULT_ACTIVE_EXTRA_SETTINGS = ActiveExtraSettings()


def active_extra_charges(
    province_code: str,
    shipment_row: dict[str, Any] | None,
    base_cost: float,
    billed_pallets: int,
    active_fuel_rate: float | None = None,
    settings: ActiveExtraSettings = DEFAULT_ACTIVE_EXTRA_SETTINGS,
) -> list[tuple[str, str, float]]:
    extras: list[tuple[str, str, float]] = []
    is_amazon = is_amazon_customer(shipment_row)
    is_gdo = is_gdo_customer(shipment_row)

    def add_percent(name: str, percent: float) -> None:
        extras.append((name, f"+{percent * 100:.0f}%", round(base_cost * percent, 2)))

    def add_fixed(name: str, amount: float) -> None:
        extras.append((name, "", round(amount, 2)))

    region = PROVINCE_TO_REGION.get(province_code, "")
    if region in ACTIVE_ETS_REGIONS:
        add_percent("ETS Sicilia/Sardegna", settings.ets_sicily_sardinia_percent)

    if active_fuel_rate is not None and active_fuel_rate > 0:
        add_percent("Fuel attivo mese", active_fuel_rate)

    if is_gdo:
        add_percent("GDO delivery and time slot booking", settings.gdo_percent)

    if row_has_any_true(shipment_row, ("Remote Area", "Area Remota", "Localita Disagiata", "Consegna Disagiata")):
        add_percent("Remote area", settings.remote_area_percent)

    if row_has_any_true(shipment_row, ("2nd delivery", "Seconda Consegna", "Returns", "Reso", "Riconsegna")):
        add_percent("2nd delivery and returns", settings.second_delivery_returns_percent)

    if has_freight_code(shipment_row, "BKV") and not is_gdo:
        if is_amazon:
            add_percent("BKV Amazon / fixed time slot", settings.amazon_fixed_time_slot_percent)
        else:
            add_percent("BKV fixed time slot", settings.fixed_time_slot_percent)

    if has_freight_code(shipment_row, "BKL") and is_amazon:
        add_percent("Amazon BKL fix day delivery and time slot booking", settings.amazon_bkl_percent)

    if requires_active_tail_lift(shipment_row, is_gdo, is_amazon):
        add_fixed("Sponda idraulica groupage", settings.tail_lift_groupage)

    if has_freight_code(shipment_row, "BKL") and not is_gdo and not is_amazon:
        add_fixed("Phone preadvise BKL", settings.phone_preadvise)

    if is_active_urgent(shipment_row):
        add_fixed("Urgent delivery", settings.urgent_per_pallet * max(1, billed_pallets))

    return extras


def flag_value(row: dict[str, Any] | None, flags: dict[str, Any], key: str) -> str:
    if key in flags and clean_text(flags.get(key)):
        return clean_text(flags.get(key))
    if row and key in row and clean_text(row.get(key)):
        return clean_text(row.get(key))
    return ""


def flag_enabled(row: dict[str, Any] | None, flags: dict[str, Any], key: str) -> bool:
    value = flag_value(row, flags, key).upper()
    return value in TRUE_VALUES


def flag_number(row: dict[str, Any] | None, flags: dict[str, Any], key: str) -> float | None:
    value = flag_value(row, flags, key)
    number = to_float(value)
    return number


def load_brt_extra_flags(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not path.exists():
        return {}
    with path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        return {
            clean_text(row.get("Shipment")): row
            for row in reader
            if clean_text(row.get("Shipment"))
        }


@dataclass
class ActiveRateResult:
    cost: float
    billed_pallets: int
    tariff_label: str
    extra_cost: float = 0.0
    extras: list[str] | None = None


@dataclass
class ContractSlaResult:
    status: str
    detail: str
    prep_hours: int | None = None
    transit_hours: float | None = None
    min_ship_at: datetime | None = None
    min_delivery_at: datetime | None = None
    transit_type: str = ""
    breach: bool = False


class ActiveRateCard:
    def __init__(
        self,
        rates_by_province: dict[str, dict[str, Any]],
        extra_settings: ActiveExtraSettings | None = None,
    ) -> None:
        self.rates_by_province = rates_by_province
        self.extra_settings = extra_settings or ActiveExtraSettings()

    @classmethod
    def from_excel(cls, path: Path) -> "ActiveRateCard":
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["Rate card+LT for transportation"]
        extra_settings = parse_active_extra_settings(wb)
        rows = ws.iter_rows(min_row=4, values_only=True)
        headers = [clean_text(value) for value in next(rows)]
        rates: dict[str, dict[str, Any]] = {}

        for values in rows:
            row = {
                headers[col_idx]: values[col_idx] if col_idx < len(values) else None
                for col_idx in range(len(headers))
            }
            province_code = clean_text(row.get("Province code")).upper()
            if not province_code:
                continue
            rates[province_code] = row

        wb.close()
        return cls(rates, extra_settings)

    def calculate(
        self,
        province: str,
        pallets: Any,
        shipment_row: dict[str, Any] | None = None,
        active_fuel_rate: float | None = None,
    ) -> ActiveRateResult | None:
        province_code = clean_text(province).upper()
        row = self.rates_by_province.get(province_code)
        pallet_value = to_float(pallets)
        if row is None or pallet_value is None:
            return None

        billed_pallets = max(1, math.ceil(pallet_value))
        if billed_pallets <= 8:
            column = f"{billed_pallets} pallet"
        elif billed_pallets <= 15:
            column = "Small truck up to 15 pallet"
        else:
            column = "FTL"

        cost = to_float(row.get(column))
        if cost is None:
            return None

        base_cost = round(cost, 2)
        extra_amounts = active_extra_charges(
            province_code,
            shipment_row,
            base_cost,
            billed_pallets,
            active_fuel_rate=active_fuel_rate,
            settings=self.extra_settings,
        )
        total_extra = round(sum(amount for _name, _rate_label, amount in extra_amounts), 2)
        final_cost = round(base_cost + total_extra, 2)

        label = f"Attiva {province_code}: {column}, {billed_pallets} pallet fatturati; base EUR {base_cost:.2f}"
        if extra_amounts:
            extras_label = ", ".join(
                f"{f'{name} {rate_label}'.strip()} EUR {amount:.2f}"
                for name, rate_label, amount in extra_amounts
            )
            label = f"{label}; extra attiva: {extras_label}; totale EUR {final_cost:.2f}"
        return ActiveRateResult(
            cost=final_cost,
            billed_pallets=billed_pallets,
            tariff_label=label,
            extra_cost=total_extra,
            extras=[
                f"{f'{name} {rate_label}'.strip()}: EUR {amount:.2f}"
                for name, rate_label, amount in extra_amounts
            ],
        )

    def transit_hours(self, province: str, is_groupage_brt: bool) -> tuple[float, str] | None:
        province_code = clean_text(province).upper()
        row = self.rates_by_province.get(province_code)
        if row is None:
            return None

        column = "Groupage up to 8 pallets" if is_groupage_brt else "Direct"
        hours = parse_hours(row.get(column))
        if hours is None:
            return None
        return hours, "groupage" if is_groupage_brt else "dedicato"

    def contract_sla(self, row: dict[str, Any]) -> ContractSlaResult:
        province = clean_text(row.get("Provincia")).upper()
        integration_at = parse_datetime_value(row.get("Integration Date"))
        is_groupage_brt = clean_text(row.get("Tipo Servizio")) == "Groupage - BRT LTL"
        transit = self.transit_hours(province, is_groupage_brt)

        missing: list[str] = []
        if integration_at is None:
            missing.append("Integration Date")
        if transit is None:
            missing.append("transit time provincia")
        if missing:
            return ContractSlaResult(
                status="Dati SLA mancanti",
                detail=f"Impossibile verificare SLA: manca {', '.join(missing)}.",
            )

        cutoff_at = integration_at.replace(
            hour=SLA_CUTOFF.hour,
            minute=SLA_CUTOFF.minute,
            second=0,
            microsecond=0,
        )
        prep_hours = (
            SLA_PREP_WITHIN_CUTOFF_HOURS
            if integration_at <= cutoff_at
            else SLA_PREP_AFTER_CUTOFF_HOURS
        )
        transit_hours, transit_type = transit
        min_ship_at = add_business_hours(integration_at, prep_hours)
        min_delivery_at = add_business_hours(min_ship_at, transit_hours)

        late_ship_at = parse_datetime_value(row.get("Late Ship Date"))
        delivery_column = (
            "Data Consegna Tassativa"
            if clean_text(row.get("Data Consegna Tassativa"))
            else "Early Delivery Date"
        )
        early_delivery_at = parse_datetime_value(row.get(delivery_column))
        delivery_label = "Data tassativa" if delivery_column == "Data Consegna Tassativa" else "Early delivery"
        breaches: list[str] = []
        missing_dates: list[str] = []

        if late_ship_at is None:
            missing_dates.append("Late Ship Date")
        elif late_ship_at.date() < min_ship_at.date():
            breaches.append(
                f"Late ship {late_ship_at.strftime('%d/%m/%Y')} prima del minimo {min_ship_at.strftime('%d/%m/%Y')}"
            )

        if early_delivery_at is None:
            missing_dates.append(delivery_column)
        elif early_delivery_at.date() < min_delivery_at.date():
            breaches.append(
                f"{delivery_label} {early_delivery_at.strftime('%d/%m/%Y')} prima del minimo {min_delivery_at.strftime('%d/%m/%Y')}"
            )

        detail_prefix = (
            f"Inserito {integration_at.strftime('%d/%m/%Y %H:%M')}; "
            f"preparazione {prep_hours}h lavorative; transit {transit_hours:g}h lavorative {transit_type}; "
            f"ship minima {min_ship_at.strftime('%d/%m/%Y')}; "
            f"consegna minima {min_delivery_at.strftime('%d/%m/%Y')}."
        )
        if breaches:
            return ContractSlaResult(
                status="SLA non rispettato",
                detail=f"{detail_prefix} {' | '.join(breaches)}.",
                prep_hours=prep_hours,
                transit_hours=transit_hours,
                min_ship_at=min_ship_at,
                min_delivery_at=min_delivery_at,
                transit_type=transit_type,
                breach=True,
            )
        if missing_dates:
            return ContractSlaResult(
                status="Dati SLA mancanti",
                detail=f"{detail_prefix} Mancano {', '.join(missing_dates)}.",
                prep_hours=prep_hours,
                transit_hours=transit_hours,
                min_ship_at=min_ship_at,
                min_delivery_at=min_delivery_at,
                transit_type=transit_type,
            )
        return ContractSlaResult(
            status="SLA OK",
            detail=detail_prefix,
            prep_hours=prep_hours,
            transit_hours=transit_hours,
            min_ship_at=min_ship_at,
            min_delivery_at=min_delivery_at,
            transit_type=transit_type,
        )


SLA_COLUMNS = [
    "SLA Contratto",
    "Preparazione SLA h",
    "Transit SLA h",
    "Tipo Transit SLA",
    "Data Ship Minima SLA",
    "Prima Consegna SLA",
    "Dettaglio SLA",
]


def set_contract_sla_fields(row: dict[str, Any], active_card: ActiveRateCard | None) -> None:
    for column in SLA_COLUMNS:
        row[column] = ""
    if not active_card:
        return

    sla_result = active_card.contract_sla(row)
    row["SLA Contratto"] = sla_result.status
    row["Preparazione SLA h"] = sla_result.prep_hours if sla_result.prep_hours is not None else ""
    row["Transit SLA h"] = sla_result.transit_hours if sla_result.transit_hours is not None else ""
    row["Tipo Transit SLA"] = sla_result.transit_type
    row["Data Ship Minima SLA"] = format_sla_date(sla_result.min_ship_at)
    row["Prima Consegna SLA"] = format_sla_date(sla_result.min_delivery_at)
    row["Dettaglio SLA"] = sla_result.detail


def apply_contract_sla_to_shipments(
    shipment_rows: list[dict[str, Any]],
    active_rates_path: Path | None = None,
) -> None:
    active_card = ActiveRateCard.from_excel(active_rates_path) if active_rates_path else None
    for row in shipment_rows:
        set_contract_sla_fields(row, active_card)


@dataclass
class PassiveRateResult:
    carrier: str
    cost: float
    taxable_weight_kg: float
    tariff_label: str
    base_cost: float = 0.0
    extra_cost: float = 0.0
    extras: list[str] | None = None


@dataclass(frozen=True)
class CarrierPalletTariff:
    carrier: str
    region: str
    pallet_from: int
    pallet_to: int
    rate_per_pallet: float


class CarrierPalletRateCard:
    def __init__(self, tariffs: list[CarrierPalletTariff]) -> None:
        self.by_carrier_region: dict[tuple[str, str], list[CarrierPalletTariff]] = {}
        for tariff in tariffs:
            key = (tariff.carrier, tariff.region)
            self.by_carrier_region.setdefault(key, []).append(tariff)
        for rules in self.by_carrier_region.values():
            rules.sort(key=lambda item: (item.pallet_from, item.pallet_to))

    @classmethod
    def from_csv(cls, path: Path | None) -> "CarrierPalletRateCard | None":
        if not path or not path.exists():
            return None

        tariffs: list[CarrierPalletTariff] = []
        with path.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            for row in reader:
                active = clean_text(row.get("Active") or "SI").upper()
                if active and active not in TRUE_VALUES:
                    continue
                carrier = clean_text(row.get("Carrier")).upper()
                region = normalize_region(clean_text(row.get("Region")))
                pallet_from = to_float(row.get("Pallets From"))
                pallet_to = to_float(row.get("Pallets To"))
                rate = to_float(row.get("Rate Per Pallet"))
                if not carrier or not region or pallet_from is None or pallet_to is None or rate is None:
                    continue
                tariffs.append(
                    CarrierPalletTariff(
                        carrier=carrier,
                        region=region,
                        pallet_from=max(1, int(math.ceil(pallet_from))),
                        pallet_to=max(1, int(math.ceil(pallet_to))),
                        rate_per_pallet=round(rate, 4),
                    )
                )
        return cls(tariffs) if tariffs else None

    def calculate(
        self,
        carrier: str,
        province: str,
        pallets: Any,
    ) -> PassiveRateResult | None:
        carrier_code = clean_text(carrier).upper()
        province_code = clean_text(province).upper()
        region = PROVINCE_TO_REGION.get(province_code)
        pallet_number = to_float(pallets)
        if not carrier_code or not region or pallet_number is None:
            return None
        billed_pallets = max(1, int(math.ceil(pallet_number)))
        rules = self.by_carrier_region.get((carrier_code, region), [])
        for rule in rules:
            if rule.pallet_from <= billed_pallets <= rule.pallet_to:
                cost = round(billed_pallets * rule.rate_per_pallet, 2)
                label = (
                    f"{carrier_code} {region}: {billed_pallets} pallet x EUR {rule.rate_per_pallet:.2f}; "
                    f"fascia {rule.pallet_from}-{rule.pallet_to} pallet"
                )
                return PassiveRateResult(
                    carrier=carrier_code,
                    cost=cost,
                    taxable_weight_kg=0,
                    tariff_label=label,
                    base_cost=cost,
                    extra_cost=0,
                    extras=[],
                )
        return None

    def calculate_all(self, province: str, pallets: Any) -> list[PassiveRateResult]:
        province_code = clean_text(province).upper()
        region = PROVINCE_TO_REGION.get(province_code)
        if not region:
            return []
        carriers = sorted(
            carrier
            for carrier, carrier_region in self.by_carrier_region
            if carrier_region == region
        )
        results = [self.calculate(carrier, province, pallets) for carrier in carriers]
        return [result for result in results if result is not None]


def best_passive_results(results: list[PassiveRateResult]) -> list[PassiveRateResult]:
    best_by_carrier: dict[str, PassiveRateResult] = {}
    for result in results:
        carrier = clean_text(result.carrier).upper()
        if not carrier:
            continue
        current = best_by_carrier.get(carrier)
        if current is None or result.cost < current.cost:
            best_by_carrier[carrier] = result
    return sorted(best_by_carrier.values(), key=lambda item: (item.cost, item.carrier))


def set_best_carrier_fields(row: dict[str, Any], results: list[PassiveRateResult]) -> None:
    best_results = best_passive_results(results)[:3]
    fields = ["Miglior Vettore", "Secondo Vettore", "Terzo Vettore"]
    for field, result in zip(fields, best_results, strict=False):
        row[field] = f"{result.carrier} EUR {result.cost:.2f}"


class BrtPassiveRateCard:
    fixed_ranges = [
        (2, "0/2 kg", "0_2"),
        (5, "3/5 kg", "3_5"),
        (10, "6/10 kg", "6_10"),
        (20, "11/20 kg", "11_20"),
        (30, "21/30 kg", "21_30"),
        (50, "31/50 kg", "31_50"),
        (100, "51/100 kg", "51_100"),
    ]

    def __init__(self, rates_by_region: dict[str, dict[str, float]]) -> None:
        self.rates_by_region = rates_by_region

    @classmethod
    def from_pdf(cls, path: Path) -> "BrtPassiveRateCard":
        reader = pypdf.PdfReader(str(path))
        rates: dict[str, dict[str, float]] = {}

        for page in reader.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                values = re.findall(r"\d+,\d{3}", line)
                if len(values) not in {3, 6}:
                    continue

                region_text = line.split(" - ")[0]
                region = normalize_region(region_text)
                if region not in BRT_REGIONS:
                    continue

                row = rates.setdefault(region, {})
                parsed = [parse_euro(value) for value in values]
                if len(parsed) == 6:
                    keys = ["0_2", "3_5", "6_10", "11_20", "21_30", "31_50"]
                else:
                    keys = ["51_100", "101_500_qle", "oltre_500_qle"]
                row.update(dict(zip(keys, parsed, strict=True)))

        return cls(rates)

    def base_calculate(
        self,
        province: str,
        weight_kg: Any,
        volume_m3: Any = None,
    ) -> tuple[str, float, float, float, str] | None:
        province_code = clean_text(province).upper()
        region = PROVINCE_TO_REGION.get(province_code)
        taxable = brt_taxable_weight(weight_kg, volume_m3)
        if region is None or taxable is None:
            return None
        taxable_weight, actual_weight, volume, volumetric_weight = taxable
        billable_weight = brt_base_billable_weight(taxable_weight)

        rates = self.rates_by_region.get(region)
        if not rates:
            return None

        weight_parts: list[str] = []
        if actual_weight is not None:
            weight_parts.append(f"peso reale {actual_weight:.3f} kg")
        if volume is not None and volumetric_weight is not None:
            weight_parts.append(
                f"peso volumetrico {volumetric_weight:.3f} kg ({volume:.3f} m3 x {BRT_VOLUMETRIC_KG_PER_M3:.0f})"
            )
        weight_parts.append(f"peso tassabile {taxable_weight:.3f} kg")
        if billable_weight != taxable_weight:
            if taxable_weight <= 100:
                weight_parts.append(f"arrotondato BRT al kg superiore {billable_weight:.0f} kg")
            else:
                weight_parts.append(f"arrotondato BRT ai 100 kg superiori {billable_weight:.0f} kg")
        weight_label = ", ".join(weight_parts)

        for ceiling, label, key in self.fixed_ranges:
            if billable_weight <= ceiling:
                cost = rates.get(key)
                if cost is None:
                    return None
                return region, cost, taxable_weight, billable_weight, f"BRT {region}: {label} a spedizione; {weight_label}"

        quintals = billable_weight / 100
        key = "101_500_qle" if billable_weight <= 500 else "oltre_500_qle"
        cost_per_quintal = rates.get(key)
        if cost_per_quintal is None:
            return None

        label = "101/500 kg" if billable_weight <= 500 else "oltre 500 kg"
        return (
            region,
            round(cost_per_quintal * quintals, 2),
            taxable_weight,
            billable_weight,
            f"BRT {region}: {label}, {quintals:.3f} q.li x {cost_per_quintal:.3f}; {weight_label}",
        )

    def calculate(
        self,
        province: str,
        weight_kg: Any,
        shipment_row: dict[str, Any] | None = None,
        extra_flags: dict[str, Any] | None = None,
        volume_m3: Any = None,
        passive_fuel_rate: float = DEFAULT_PASSIVE_FUEL_RATE,
    ) -> PassiveRateResult | None:
        if volume_m3 is None and shipment_row is not None:
            volume_m3 = shipment_row.get("Grand Total Shipment Ftp Vol m3")
        base = self.base_calculate(province, weight_kg, volume_m3)
        if base is None:
            return None

        region, base_cost, taxable_weight, billable_weight, base_label = base
        flags = extra_flags or {}
        extras: list[tuple[str, float]] = []
        quintals = qli(taxable_weight)
        address = clean_text((shipment_row or {}).get("Route To Address")).upper()
        service_level = clean_text((shipment_row or {}).get("Service Level")).upper()
        freight_code = clean_text((shipment_row or {}).get("Freight Code")).upper()

        if passive_fuel_rate > 0:
            fuel_label = (
                "Fuel surcharge 2% minimo"
                if abs(passive_fuel_rate - DEFAULT_PASSIVE_FUEL_RATE) < 0.0001
                else f"Fuel surcharge passivo {format_percent(passive_fuel_rate)}"
            )
            extras.append((fuel_label, capped(base_cost * passive_fuel_rate)))

        if region in {"SICILIA", "SARDEGNA"}:
            extras.append(("Traghetti", capped(taxable_weight * 0.030, minimum=1.90, maximum=24.00)))

        if "LIVIGNO" in address:
            extras.append(("Zona franca Livigno", capped(29.90 * quintals)))
        elif "CAMPIONE" in address:
            extras.append(("Zona franca Campione d'Italia", capped(39.90 * quintals)))
        elif flag_enabled(shipment_row, flags, "Isole Minori"):
            extras.append(("Isole minori / zone franche", capped(15.90 * quintals)))

        if flag_enabled(shipment_row, flags, "Dirottamento"):
            extras.append(("Dirottamento", capped(5.90 * quintals)))

        if flag_enabled(shipment_row, flags, "Localita Disagiata"):
            extras.append(("Localita disagiata", capped(3.90 * quintals)))

        if flag_enabled(shipment_row, flags, "Consegna Disagiata"):
            extras.append(("Consegna disagiata", capped(10.00 * quintals)))

        if flag_enabled(shipment_row, flags, "Supermercati GDO") or is_amazon_customer(shipment_row):
            extras.append(("Consegna supermercati/GDO", capped(7.90 * quintals, maximum=20.00)))

        if flag_enabled(shipment_row, flags, "Priority") or "PRIORITY" in service_level:
            extras.append(("Servizio Priority 20%", capped(base_cost * 0.20)))

        if flag_enabled(shipment_row, flags, "Servizio 10:30") or "10:30" in service_level:
            extras.append(("Servizio 10:30", capped(base_cost * 0.30, minimum=4.90)))

        cod_value = flag_number(shipment_row, flags, "Contrassegno Valore")
        if cod_value and cod_value > 0:
            extras.append(("Provvigione contrassegno", capped(cod_value * 0.01, minimum=3.90)))

        if flag_enabled(shipment_row, flags, "ZTL"):
            extras.append(("ZTL", capped(base_cost * 0.10, minimum=2.90)))

        appointment_auto = "BKL" in {code.strip() for code in freight_code.split("|") if code.strip()}
        if appointment_auto and not is_amazon_customer(shipment_row) and not flag_enabled(shipment_row, flags, "Appuntamento"):
            extras.append(("Consegna per appuntamento da BKL", 3.90))

        fixed_extras = [
            ("Fuori Misura", "Fuori misura", 6.90),
            ("Appuntamento", "Consegna per appuntamento", 3.90),
            ("POD Image", "P.O.D. image", 1.25),
            ("Ricerca Documenti", "Ricerca archivio/documenti", 6.20),
            ("ORM Commissionato", "O.R.M. commissionato", 4.00),
            ("Recapito Contrassegni", "Recapito contrassegni", 0.25),
            ("Giacenza Dossier", "Giacenza - spese dossier", 10.00),
        ]
        for flag_key, label, amount in fixed_extras:
            if flag_enabled(shipment_row, flags, flag_key):
                extras.append((label, amount))

        pallet_count = flag_number(shipment_row, flags, "Bancali Rendere")
        if pallet_count and pallet_count > 0:
            extras.append(("Gestione bancali a rendere", capped(3.90 * pallet_count)))
        elif flag_enabled(shipment_row, flags, "Bancali Rendere"):
            fallback_pallets = to_float((shipment_row or {}).get("Pallet Fatturati")) or to_float(
                (shipment_row or {}).get("Theoretical Pallets")
            ) or 1
            extras.append(("Gestione bancali a rendere", capped(3.90 * max(1, math.ceil(fallback_pallets)))))

        if flag_enabled(shipment_row, flags, "Riconsegna Giacenza"):
            extras.append(("Riconsegna giacenza", capped(3.90 * quintals)))

        extra_cost = round(sum(amount for _, amount in extras), 2)
        extras_label = ", ".join(f"{name} EUR {amount:.2f}" for name, amount in extras)
        return PassiveRateResult(
            carrier="BRT",
            cost=round(base_cost + extra_cost, 2),
            taxable_weight_kg=billable_weight,
            tariff_label=f"{base_label}; extra: {extras_label}" if extras_label else base_label,
            base_cost=round(base_cost, 2),
            extra_cost=extra_cost,
            extras=[f"{name}: EUR {amount:.2f}" for name, amount in extras],
        )


def apply_tariffs_to_shipments(
    shipment_rows: list[dict[str, Any]],
    active_rates_path: Path | None = None,
    brt_passive_pdf_path: Path | None = None,
    brt_extra_flags_path: Path | None = None,
    gdo_customers_path: Path | None = None,
    fuel_settings_path: Path | None = None,
) -> None:
    active_card = ActiveRateCard.from_excel(active_rates_path) if active_rates_path else None
    brt_card = BrtPassiveRateCard.from_pdf(brt_passive_pdf_path) if brt_passive_pdf_path else None
    carrier_rate_card = CarrierPalletRateCard.from_csv(default_carrier_tariffs_path())
    brt_extra_flags = load_brt_extra_flags(brt_extra_flags_path)
    gdo_registry = load_gdo_customer_registry(gdo_customers_path)
    fuel_settings = load_fuel_settings(fuel_settings_path)

    for row in shipment_rows:
        apply_pallet_weight_to_shipment(row)
        province = clean_text(row.get("Provincia"))
        shipment = clean_text(row.get("Shipment"))
        selected_carrier = clean_text(row.get("Carrier Scelto")).upper()
        active_fuel_rate, passive_fuel_rate = fuel_settings_for_row(row, fuel_settings)

        gdo_match = match_gdo_customer(row, gdo_registry)
        if gdo_match:
            row["Cliente GDO"] = "SI"
            row["Cliente GDO Fonte"] = gdo_match.label
            row["Supermercati GDO"] = "SI"
            row["Attiva Sponda"] = "SI" if gdo_match.active_tail_lift else ""
        else:
            row["Cliente GDO"] = ""
            row["Cliente GDO Fonte"] = ""
            row["Supermercati GDO"] = ""
            row["Attiva Sponda"] = ""

        for column in [
            "Costo Attivo",
            "Pallet Fatturati",
            "Extra Attivi Totale",
            "Extra Attivi Applicati",
            "Tariffa Attiva Applicata",
        ]:
            row[column] = ""

        active_result = (
            active_card.calculate(
                province,
                row.get("Theoretical Pallets"),
                shipment_row=row,
                active_fuel_rate=active_fuel_rate,
            )
            if active_card
            else None
        )
        if active_result:
            row["Costo Attivo"] = round(active_result.cost, 2)
            row["Pallet Fatturati"] = active_result.billed_pallets
            row["Extra Attivi Totale"] = round(active_result.extra_cost, 2)
            row["Extra Attivi Applicati"] = " | ".join(active_result.extras or [])
            row["Tariffa Attiva Applicata"] = active_result.tariff_label

        for column in [
            "Costo Passivo",
            "Costo Passivo Base BRT",
            "Extra BRT Totale",
            "Extra BRT Applicati",
            "Peso Tariffabile BRT Kg",
            "Tariffa Passiva Applicata",
            "Miglior Vettore",
            "Secondo Vettore",
            "Terzo Vettore",
            "Margine",
            *SLA_COLUMNS,
        ]:
            row[column] = ""

        set_contract_sla_fields(row, active_card)

        passive_results: list[PassiveRateResult] = []
        recommendation_results: list[PassiveRateResult] = []
        is_groupage_brt = clean_text(row.get("Tipo Servizio")) == "Groupage - BRT LTL"
        use_brt = selected_carrier == "BRT" or (is_groupage_brt and selected_carrier in {"", "BRT"})
        if not is_groupage_brt:
            if brt_card:
                brt_recommendation = brt_card.calculate(
                    province,
                    row.get("Grand Total Shipment Ftp Wgt Kg"),
                    shipment_row=row,
                    extra_flags=brt_extra_flags.get(shipment, {}),
                    volume_m3=row.get("Grand Total Shipment Ftp Vol m3"),
                    passive_fuel_rate=passive_fuel_rate,
                )
                if brt_recommendation:
                    recommendation_results.append(brt_recommendation)
            if carrier_rate_card:
                recommendation_results.extend(
                    carrier_rate_card.calculate_all(
                        province,
                        row.get("Pallet Fatturati") or row.get("Theoretical Pallets"),
                    )
                )
            set_best_carrier_fields(row, recommendation_results)

        if brt_card and use_brt:
            brt_result = brt_card.calculate(
                province,
                row.get("Grand Total Shipment Ftp Wgt Kg"),
                shipment_row=row,
                extra_flags=brt_extra_flags.get(shipment, {}),
                volume_m3=row.get("Grand Total Shipment Ftp Vol m3"),
                passive_fuel_rate=passive_fuel_rate,
            )
            if brt_result:
                passive_results.append(brt_result)

        if carrier_rate_card and selected_carrier and selected_carrier != "BRT":
            carrier_result = carrier_rate_card.calculate(
                selected_carrier,
                province,
                row.get("Pallet Fatturati") or row.get("Theoretical Pallets"),
            )
            if carrier_result:
                passive_results.append(carrier_result)

        passive_results.sort(key=lambda result: result.cost)

        if passive_results:
            chosen = passive_results[0]
            row["Carrier Scelto"] = chosen.carrier
            row["Costo Passivo"] = round(chosen.cost, 2)
            row["Costo Passivo Base BRT"] = round(chosen.base_cost, 2)
            row["Extra BRT Totale"] = round(chosen.extra_cost, 2)
            row["Extra BRT Applicati"] = " | ".join(chosen.extras or [])
            row["Peso Tariffabile BRT Kg"] = round(chosen.taxable_weight_kg, 3)
            row["Tariffa Passiva Applicata"] = chosen.tariff_label
            if not recommendation_results:
                set_best_carrier_fields(row, passive_results)

        active_cost = to_float(row.get("Costo Attivo"))
        passive_cost = to_float(row.get("Costo Passivo"))
        if active_cost is not None and passive_cost is not None:
            margin = round(active_cost - passive_cost, 2)
            row["Margine"] = margin
            row["Esito Margine"] = "Guadagno" if margin >= 0 else "Perdita"
        elif active_card or brt_card:
            if selected_carrier and selected_carrier != "BRT":
                row["Esito Margine"] = f"In attesa passiva {selected_carrier}"
            elif not is_groupage_brt and brt_card:
                row["Esito Margine"] = "In attesa scelta vettore"
            else:
                row["Esito Margine"] = "Tariffa incompleta"
