from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tariff_engine import apply_tariffs_to_shipments


HEADER_ROW = 8

REQUIRED_COLUMNS = [
    "Integration Date",
    "Order Type",
    "Order",
    "Order Ordered Qty",
    "Note Text",
    "Carrier",
    "Service Level",
    "Wave",
    "Shipment",
    "Freight Code",
    "Item Description",
    "Item",
    "Grand Total Shipment Ftp Wgt Kg",
    "Grand Total Shipment Ftp Vol m3",
    "Route To Address",
    "Late Ship Date",
    "Early Delivery Date",
    "Units per Pallet",
    "Theoretical Pallets",
    "Grand Total Shipment Ftp CS Wgt Kg",
    "Route to Customer",
]

SHIPMENT_COLUMNS = [
    "Shipment",
    "Orders",
    "Integration Date",
    "Order Type",
    "Order Ordered Qty",
    "Note Text",
    "Carrier Originale",
    "Carrier Scelto",
    "Service Level",
    "Service Level Manuale",
    "Tipo Servizio",
    "Freight Code",
    "Prenotazione Scarico",
    "Attiva Urgente",
    "Attiva Sponda",
    "Wave",
    "Data Partenza Wave",
    "Vettore Wave",
    "Tipo Wave",
    "Route to Customer",
    "Cliente GDO",
    "Route To Address",
    "Provincia",
    "Late Ship Date",
    "Early Delivery Date",
    "Data Scarico Prenotato",
    "SLA Contratto",
    "Preparazione SLA h",
    "Transit SLA h",
    "Tipo Transit SLA",
    "Data Ship Minima SLA",
    "Prima Consegna SLA",
    "Dettaglio SLA",
    "Righe Articolo",
    "Articoli Diversi",
    "Grand Total Shipment Ftp Wgt Kg",
    "Grand Total Shipment Ftp Vol m3",
    "Grand Total Shipment Ftp CS Wgt Kg",
    "Theoretical Pallets",
    "Pallet Originali",
    "Pallet Manuali",
    "Pallet Fatturati",
    "Peso Tariffabile BRT Kg",
    "Costo Attivo",
    "Extra Attivi Totale",
    "Costo Passivo Base BRT",
    "Extra BRT Totale",
    "Costo Passivo Manuale",
    "Costo Passivo",
    "Margine",
    "Esito Margine",
    "Tariffa Attiva Applicata",
    "Extra Attivi Applicati",
    "Tariffa Passiva Applicata",
    "Extra BRT Applicati",
    "Miglior Vettore",
    "Secondo Vettore",
    "Terzo Vettore",
]

DATE_ONLY_COLUMNS = {"Late Ship Date", "Early Delivery Date", "Data Ship Minima SLA", "Prima Consegna SLA", "Data Partenza Wave"}

ITALIAN_PROVINCES = {
    "AG",
    "AL",
    "AN",
    "AO",
    "AP",
    "AQ",
    "AR",
    "AT",
    "AV",
    "BA",
    "BG",
    "BI",
    "BL",
    "BN",
    "BO",
    "BR",
    "BS",
    "BT",
    "BZ",
    "CA",
    "CB",
    "CE",
    "CH",
    "CI",
    "CL",
    "CN",
    "CO",
    "CR",
    "CS",
    "CT",
    "CZ",
    "EN",
    "FC",
    "FE",
    "FG",
    "FI",
    "FM",
    "FR",
    "GE",
    "GO",
    "GR",
    "IM",
    "IS",
    "KR",
    "LC",
    "LE",
    "LI",
    "LO",
    "LT",
    "LU",
    "MB",
    "MC",
    "ME",
    "MI",
    "MN",
    "MO",
    "MS",
    "MT",
    "NA",
    "NO",
    "NU",
    "OG",
    "OR",
    "OT",
    "PA",
    "PC",
    "PD",
    "PE",
    "PG",
    "PI",
    "PN",
    "PO",
    "PR",
    "PT",
    "PU",
    "PV",
    "PZ",
    "RA",
    "RC",
    "RE",
    "RG",
    "RI",
    "RM",
    "RN",
    "RO",
    "SA",
    "SI",
    "SO",
    "SP",
    "SR",
    "SS",
    "SU",
    "SV",
    "TA",
    "TE",
    "TN",
    "TO",
    "TP",
    "TR",
    "TS",
    "TV",
    "UD",
    "VA",
    "VB",
    "VC",
    "VE",
    "VI",
    "VR",
    "VS",
    "VT",
    "VV",
}


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def clean_text(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_key(value: Any) -> str:
    return clean_text(value)


def to_float(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, int | float):
        return float(value)
    text = clean_text(value).replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def serialize(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return round(value, 6)
    return value


def serialize_cell(column: str, value: Any) -> Any:
    if column in DATE_ONLY_COLUMNS:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, date):
            return value.strftime("%d/%m/%Y")

        text = clean_text(value)
        date_part = text.replace("T", " ").split(" ")[0]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(date_part, fmt).strftime("%d/%m/%Y")
            except ValueError:
                continue
        return date_part
    return serialize(value)


def set_if_present(row: dict[str, Any], column: str, value: Any) -> None:
    if column in row and is_blank(row[column]) and not is_blank(value):
        row[column] = value


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    header_values = [cell.value for cell in ws[HEADER_ROW]]
    column_indexes: list[tuple[int, str]] = []
    seen: dict[str, int] = defaultdict(int)

    for idx, header in enumerate(header_values):
        if is_blank(header):
            continue
        name = clean_text(header)
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name} ({seen[name]})"
        column_indexes.append((idx, name))

    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        row = {name: values[idx] if idx < len(values) else None for idx, name in column_indexes}
        if any(not is_blank(value) for value in row.values()):
            rows.append(row)

    wb.close()
    return rows


def index_first(rows: list[dict[str, Any]], key_columns: tuple[str, ...]) -> dict[tuple[str, ...], dict[str, Any]]:
    indexed: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in rows:
        key = tuple(clean_key(row.get(column)) for column in key_columns)
        if all(key) and key not in indexed:
            indexed[key] = row
    return indexed


def build_notes_by_order(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    notes: dict[str, dict[str, Any]] = {}
    texts_by_order: dict[str, list[str]] = defaultdict(list)

    for row in rows:
        order = clean_key(row.get("Order"))
        if not order:
            continue
        current = notes.setdefault(order, {})
        for column in [
            "Integration Date",
            "Order Type",
            "Order Ordered Qty",
            "Carrier",
            "Service Level",
            "Wave",
            "Shipment",
        ]:
            if is_blank(current.get(column)) and not is_blank(row.get(column)):
                current[column] = row.get(column)

        note_text = clean_text(row.get("Note Text"))
        if note_text and note_text not in texts_by_order[order]:
            texts_by_order[order].append(note_text)

    for order, texts in texts_by_order.items():
        notes.setdefault(order, {})["Note Text"] = " | ".join(texts)

    return notes


def unique_values(rows: list[dict[str, Any]], column: str) -> list[str]:
    values: list[str] = []
    for row in rows:
        value = clean_text(row.get(column))
        if value and value not in values:
            values.append(value)
    return values


def joined_unique(rows: list[dict[str, Any]], column: str) -> str:
    return " | ".join(unique_values(rows, column))


def first_present(rows: list[dict[str, Any]], column: str) -> Any:
    for row in rows:
        value = row.get(column)
        if not is_blank(value):
            return value
    return ""


def first_number(rows: list[dict[str, Any]], column: str) -> float | str:
    for row in rows:
        value = to_float(row.get(column))
        if value is not None:
            return value
    return ""


def sum_numbers(rows: list[dict[str, Any]], column: str) -> float:
    total = 0.0
    for row in rows:
        value = to_float(row.get(column))
        if value is not None:
            total += value
    return total


def sum_order_quantities(rows: list[dict[str, Any]]) -> float:
    qty_by_order: dict[str, float] = {}
    for row in rows:
        order = clean_key(row.get("Order"))
        qty = to_float(row.get("Order Ordered Qty"))
        if order and qty is not None and order not in qty_by_order:
            qty_by_order[order] = qty
    return sum(qty_by_order.values())


def extract_province(address: Any) -> str:
    text = clean_text(address).upper()
    if not text:
        return ""

    parts = [part.strip().upper() for part in text.split(",") if part.strip()]
    for part in parts:
        if part in ITALIAN_PROVINCES:
            return part

    match = re.search(r"\b([A-Z]{2})\b\s*,?\s*\d{5}\b", text)
    if match and match.group(1) in ITALIAN_PROVINCES:
        return match.group(1)

    tokens = re.findall(r"\b[A-Z]{2}\b", text)
    for token in tokens:
        if token in ITALIAN_PROVINCES:
            return token

    return ""


def classify_service(carrier: Any, service_level: Any) -> str:
    carrier_text = clean_text(carrier).upper()
    service_text = clean_text(service_level).upper()

    if "BRT" in carrier_text and "LTL" in service_text:
        return "Groupage - BRT LTL"
    if ("KN" in carrier_text or "KUEHNE" in carrier_text) and ("LTL" in service_text or "FTL" in service_text):
        return "Diretta - KN LTL/FTL"
    return "Altro"


WAVE_CARRIER_ALIASES = {
    "KUEHNE NAGEL": "KN",
    "KUEHNE+NAGEL": "KN",
    "BARTOLINI": "BRT",
    "GRENDI": "GRENDI",
    "FERCAM": "FERCAM",
    "DACHSER": "DACHSER",
    "GEODIS": "GEODIS",
    "DHL": "DHL",
    "DSV": "DSV",
    "BRT": "BRT",
    "KN": "KN",
}


def normalized_wave_text(wave: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(wave).upper()).strip()


def parse_wave_departure_date(wave: Any) -> str:
    text = normalized_wave_text(wave)
    for match in re.finditer(r"\b(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?\b", text):
        day_text, month_text, year_text = match.groups()
        year = date.today().year
        if year_text:
            year = int(year_text)
            if year < 100:
                year += 2000
        try:
            return date(year, int(month_text), int(day_text)).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def infer_wave_carrier(wave: Any) -> str:
    text = normalized_wave_text(wave)
    padded = f" {re.sub(r'[^A-Z0-9+]+', ' ', text)} "
    for alias, carrier in sorted(WAVE_CARRIER_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        alias_pattern = f" {re.sub(r'[^A-Z0-9+]+', ' ', alias).strip()} "
        if alias_pattern in padded:
            return carrier
    return ""


def infer_wave_service_level(wave: Any) -> str:
    text = normalized_wave_text(wave)
    if re.search(r"\bFTL\b", text):
        return "FTL"
    if re.search(r"\bLTL\b", text) or "GROUPAGE" in text:
        return "LTL"
    return ""


def wave_is_groupage(wave: Any, carrier: Any, service_level: Any) -> bool:
    text = normalized_wave_text(wave)
    wave_carrier = infer_wave_carrier(text) or clean_text(carrier).upper()
    wave_level = infer_wave_service_level(text) or clean_text(service_level).upper()
    return "GROUPAGE" in text or (wave_carrier == "BRT" and wave_level == "LTL")


def apply_wave_decision(wave: Any, carrier: Any, service_level: Any) -> dict[str, str]:
    wave_carrier = infer_wave_carrier(wave)
    wave_level = infer_wave_service_level(wave)
    selected_carrier = wave_carrier or clean_text(carrier).upper()
    selected_level = wave_level or clean_text(service_level).upper()
    if wave_is_groupage(wave, selected_carrier, selected_level):
        selected_carrier = "BRT"
        selected_level = "LTL"
        tipo_servizio = "Groupage - BRT LTL"
        tipo_wave = "BRT GROUPAGE"
    else:
        tipo_servizio = classify_service(selected_carrier, selected_level)
        if tipo_servizio == "Altro" and selected_level in {"LTL", "FTL"}:
            tipo_servizio = "Diretta - KN LTL/FTL"
        tipo_wave = " ".join(part for part in (selected_carrier, selected_level) if part)
    return {
        "carrier": selected_carrier or clean_text(carrier),
        "service_level": selected_level or clean_text(service_level),
        "tipo_servizio": tipo_servizio,
        "wave_carrier": wave_carrier,
        "wave_type": tipo_wave,
        "wave_departure": parse_wave_departure_date(wave),
    }


def classify_unloading_booking(freight_code: Any) -> str:
    codes = {code.strip().upper() for code in clean_text(freight_code).split("|") if code.strip()}
    if "BKL" in codes:
        return "Da prenotare"
    if "BKV" in codes:
        return "Gia prenotato da loro"
    if not codes:
        return "Non necessaria"
    return "Verificare"


def extract_vtech_rows(path: Path) -> list[dict[str, Any]]:
    order_lines = load_rows(path, "Outbound Order Lines")
    order_notes = load_rows(path, "Order Notes (VAS)")
    outbound_orders = load_rows(path, "Outbound Orders")
    shipments = load_rows(path, "Outbound Shipments")
    transport = load_rows(path, "Outbound Transport")
    insights = load_rows(path, "Theoretical Insights by Order")

    notes_by_order = build_notes_by_order(order_notes)
    orders_by_order = index_first(outbound_orders, ("Order",))
    shipments_by_ordlin = index_first(shipments, ("Ord+Lin",))
    transport_by_ordlin = index_first(transport, ("Ord+Lin",))
    insights_by_order_shipment = index_first(insights, ("Order", "Shipment"))

    extracted: list[dict[str, Any]] = []

    for source in order_lines:
        row = {column: "" for column in REQUIRED_COLUMNS}
        order = clean_key(source.get("Order"))
        order_line = clean_key(source.get("Order Line"))
        shipment = clean_key(source.get("Shipment"))
        ordlin = clean_key(source.get("Ord+Lin"))

        row.update(
            {
                "Integration Date": source.get("Integration Date"),
                "Order Type": source.get("Order Type"),
                "Order": order,
                "Carrier": source.get("Carrier"),
                "Service Level": source.get("Service Level"),
                "Wave": source.get("Wave"),
                "Shipment": shipment,
                "Item Description": source.get("Item Description"),
                "Item": source.get("Item"),
                "Route to Customer": source.get("Route to Customer"),
            }
        )

        note = notes_by_order.get(order, {})
        for column in ["Integration Date", "Order Type", "Order Ordered Qty", "Note Text", "Carrier", "Service Level", "Wave", "Shipment"]:
            set_if_present(row, column, note.get(column))

        order_header = orders_by_order.get((order,), {})
        set_if_present(row, "Order Ordered Qty", order_header.get("Ordered Qty"))

        shipment_row = shipments_by_ordlin.get((ordlin,), {})
        if shipment_row:
            set_if_present(row, "Integration Date", shipment_row.get("Order Integration Date"))
            set_if_present(row, "Item Description", shipment_row.get("Item Description"))
            set_if_present(row, "Item", shipment_row.get("Item"))
            set_if_present(row, "Grand Total Shipment Ftp Wgt Kg", shipment_row.get("Grand Total Shipment Ftp Wgt Kg"))
            set_if_present(row, "Grand Total Shipment Ftp Vol m3", shipment_row.get("Grand Total Shipment Ftp Vol m3"))
            set_if_present(row, "Late Ship Date", shipment_row.get("Late Shipping Date"))
            set_if_present(row, "Early Delivery Date", shipment_row.get("Early Delivery Date"))
            set_if_present(row, "Route to Customer", shipment_row.get("Route to Customer"))

        transport_row = transport_by_ordlin.get((ordlin,), {})
        if transport_row:
            set_if_present(row, "Route To Address", transport_row.get("Route To Address"))
            set_if_present(row, "Freight Code", transport_row.get("Freight Code"))
            set_if_present(row, "Late Ship Date", transport_row.get("Late Ship Date"))
            set_if_present(row, "Early Delivery Date", transport_row.get("Early Delivery Date"))
            set_if_present(row, "Units per Pallet", transport_row.get("Units per Pallet"))
            set_if_present(row, "Theoretical Pallets", transport_row.get("Theoretical Pallets"))
            set_if_present(row, "Grand Total Shipment Ftp CS Wgt Kg", transport_row.get("Grand Total Shipment Ftp CS Wgt Kg"))

        shipment = clean_key(row.get("Shipment"))
        insight = insights_by_order_shipment.get((order, shipment), {})
        set_if_present(row, "Theoretical Pallets", insight.get("Theoretical Pallets"))

        if not clean_key(row.get("Shipment")):
            continue

        extracted.append(row)

    return extracted


def build_shipment_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        shipment = clean_key(row.get("Shipment"))
        if not shipment:
            continue
        grouped[shipment].append(row)

    shipment_rows: list[dict[str, Any]] = []
    for shipment in sorted(grouped):
        group = grouped[shipment]
        carrier = joined_unique(group, "Carrier")
        service_level = joined_unique(group, "Service Level")
        freight_code = joined_unique(group, "Freight Code")
        address = first_present(group, "Route To Address")
        total_weight = first_number(group, "Grand Total Shipment Ftp Wgt Kg")
        total_volume = first_number(group, "Grand Total Shipment Ftp Vol m3")
        total_cs_weight = first_number(group, "Grand Total Shipment Ftp CS Wgt Kg")
        total_pallets = sum_numbers(group, "Theoretical Pallets")
        wave = joined_unique(group, "Wave")
        wave_decision = apply_wave_decision(wave, carrier, service_level)
        tipo_servizio = wave_decision["tipo_servizio"]

        shipment_rows.append(
            {
                "Shipment": shipment,
                "Orders": joined_unique(group, "Order"),
                "Integration Date": first_present(group, "Integration Date"),
                "Order Type": joined_unique(group, "Order Type"),
                "Order Ordered Qty": sum_order_quantities(group),
                "Note Text": joined_unique(group, "Note Text"),
                "Carrier Originale": carrier,
                "Carrier Scelto": wave_decision["carrier"],
                "Service Level": wave_decision["service_level"],
                "Tipo Servizio": tipo_servizio,
                "Freight Code": freight_code,
                "Prenotazione Scarico": classify_unloading_booking(freight_code),
                "Attiva Urgente": "",
                "Attiva Sponda": "",
                "Wave": wave,
                "Data Partenza Wave": wave_decision["wave_departure"],
                "Vettore Wave": wave_decision["wave_carrier"],
                "Tipo Wave": wave_decision["wave_type"],
                "Route to Customer": joined_unique(group, "Route to Customer"),
                "Cliente GDO": "",
                "Route To Address": address,
                "Provincia": extract_province(address),
                "Late Ship Date": first_present(group, "Late Ship Date"),
                "Early Delivery Date": first_present(group, "Early Delivery Date"),
                "Righe Articolo": len(group),
                "Articoli Diversi": len(set(unique_values(group, "Item"))),
                "Grand Total Shipment Ftp Wgt Kg": total_weight,
                "Grand Total Shipment Ftp Vol m3": total_volume,
                "Grand Total Shipment Ftp CS Wgt Kg": total_cs_weight,
                "Theoretical Pallets": total_pallets,
                "Pallet Fatturati": "",
                "Peso Tariffabile BRT Kg": "",
                "Costo Attivo": "",
                "Extra Attivi Totale": "",
                "Costo Passivo Base BRT": "",
                "Extra BRT Totale": "",
                "Costo Passivo": "",
                "Margine": "",
                "Esito Margine": "In attesa tariffari",
                "Tariffa Attiva Applicata": "",
                "Extra Attivi Applicati": "",
                "Tariffa Passiva Applicata": "",
                "Extra BRT Applicati": "",
                "Miglior Vettore": "",
                "Secondo Vettore": "",
                "Terzo Vettore": "",
            }
        )

    return shipment_rows


def write_csv(rows: list[dict[str, Any]], output_path: Path, columns: list[str]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: serialize_cell(column, row.get(column)) for column in columns})


def build_summary(rows: list[dict[str, Any]], shipment_rows: list[dict[str, Any]]) -> dict[str, Any]:
    completeness = {
        column: sum(1 for row in rows if not is_blank(row.get(column)))
        for column in REQUIRED_COLUMNS
    }
    shipment_completeness = {
        column: sum(1 for row in shipment_rows if not is_blank(row.get(column)))
        for column in SHIPMENT_COLUMNS
    }
    return {
        "detail_rows": len(rows),
        "shipment_rows": len(shipment_rows),
        "orders": len({clean_key(row.get("Order")) for row in rows if clean_key(row.get("Order"))}),
        "shipments": len({clean_key(row.get("Shipment")) for row in rows if clean_key(row.get("Shipment"))}),
        "detail_columns": REQUIRED_COLUMNS,
        "shipment_columns": SHIPMENT_COLUMNS,
        "detail_completeness": completeness,
        "shipment_completeness": shipment_completeness,
        "provinces": sorted({clean_key(row.get("Provincia")) for row in shipment_rows if clean_key(row.get("Provincia"))}),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa il report outbound V-Tech e crea tabelle pulite.")
    parser.add_argument("input", type=Path, help="Percorso del file Excel V-Tech.")
    parser.add_argument("--output", type=Path, default=Path("outputs/vtech_spedizioni_estratte.csv"))
    parser.add_argument("--shipments-output", type=Path, default=Path("outputs/vtech_spedizioni_riepilogo.csv"))
    parser.add_argument("--summary", type=Path, default=Path("outputs/vtech_spedizioni_summary.json"))
    parser.add_argument("--active-rates", type=Path, help="File Excel con le tariffe attive.")
    parser.add_argument("--brt-passive-pdf", type=Path, help="PDF con le tariffe passive Bartolini/BRT.")
    parser.add_argument("--brt-extra-flags", type=Path, help="CSV opzionale con extra BRT manuali per shipment.")
    parser.add_argument("--gdo-customers", type=Path, help="CSV anagrafica clienti da classificare come GDO.")
    parser.add_argument("--fuel-settings", type=Path, help="JSON con fuel mensile attivo/passivo.")
    args = parser.parse_args()

    rows = extract_vtech_rows(args.input)
    shipment_rows = build_shipment_rows(rows)
    apply_tariffs_to_shipments(
        shipment_rows,
        active_rates_path=args.active_rates,
        brt_passive_pdf_path=args.brt_passive_pdf,
        brt_extra_flags_path=args.brt_extra_flags,
        gdo_customers_path=args.gdo_customers,
        fuel_settings_path=args.fuel_settings,
    )
    write_csv(rows, args.output, REQUIRED_COLUMNS)
    write_csv(shipment_rows, args.shipments_output, SHIPMENT_COLUMNS)

    summary = build_summary(rows, shipment_rows)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.summary.write_text(json.dumps(summary, indent=2, ensure_ascii=False, default=serialize), encoding="utf-8")

    print(
        json.dumps(
            {
                "detail_output": str(args.output),
                "shipments_output": str(args.shipments_output),
                "summary": summary,
            },
            ensure_ascii=False,
            default=serialize,
        )
    )


if __name__ == "__main__":
    main()
