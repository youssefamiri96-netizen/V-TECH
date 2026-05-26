from __future__ import annotations

import argparse
import calendar
import copy
import csv
import html as html_lib
import json
import math
import os
import re
import shutil
import sqlite3
import subprocess
import threading
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog, ttk
except Exception:
    class _TkFallback:
        class Tk:
            def __init__(self, *_args: Any, **_kwargs: Any) -> None:
                raise RuntimeError("Interfaccia desktop Tk non disponibile in questo ambiente.")

    tk = _TkFallback()  # type: ignore[assignment]
    filedialog = messagebox = simpledialog = ttk = None  # type: ignore[assignment]

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from tariff_engine import BRT_EXTRA_FLAG_COLUMNS, customer_match_keys, iter_registry_values, to_float
from vtech_importer import (
    REQUIRED_COLUMNS,
    SHIPMENT_COLUMNS,
    build_shipment_rows,
    build_summary,
    classify_unloading_booking,
    clean_text,
    extract_vtech_rows,
    parse_wave_departure_date,
    serialize,
    write_csv,
)
from tariff_engine import apply_contract_sla_to_shipments, apply_tariffs_to_shipments, default_carrier_tariffs_path


APP_DIR = Path(__file__).resolve().parent


def env_flag(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "si", "on"}


def cloud_runtime_enabled() -> bool:
    return env_flag(os.environ.get("VTECH_CLOUD_MODE")) or bool(
        os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RAILWAY_SERVICE_NAME")
    )


def runtime_data_default() -> Path:
    railway_volume = clean_text(os.environ.get("RAILWAY_VOLUME_MOUNT_PATH"))
    if railway_volume:
        return Path(railway_volume)
    if cloud_runtime_enabled():
        return Path("/data")
    return APP_DIR / "data"


DATA_DIR = Path(os.environ.get("VTECH_DATA_DIR") or runtime_data_default()).resolve()
OUTPUT_DIR = Path(os.environ.get("VTECH_OUTPUT_DIR") or (DATA_DIR / "outputs" if cloud_runtime_enabled() else APP_DIR / "outputs")).resolve()
DOWNLOADS_DIR = Path(os.environ.get("VTECH_DOWNLOADS_DIR") or (DATA_DIR / "downloads" if cloud_runtime_enabled() else Path.home() / "Downloads")).resolve()
XML_DIR = Path(os.environ.get("VTECH_XML_DIR") or str(DOWNLOADS_DIR)).resolve()
SEED_DIR = APP_DIR / "seed"
DB_PATH = DATA_DIR / "vtech_trasporti.db"
SETTINGS_PATH = DATA_DIR / "settings.json"
BRT_EXTRA_FLAGS_PATH = DATA_DIR / "brt_extra_flags.csv"
GDO_CUSTOMERS_PATH = DATA_DIR / "gdo_customers.csv"
CUSTOMER_REGISTRY_PATH = DATA_DIR / "customer_registry.csv"
FUEL_SETTINGS_PATH = DATA_DIR / "fuel_settings.json"
DELIVERY_XML_TEMPLATE_PATH = DATA_DIR / "kn_delivery_template.xml"
DOWNLOAD_SCAN_MS = 15000
DOWNLOAD_STABLE_SECONDS = 5
SHIPMENTS_CACHE_LOCK = threading.Lock()
SHIPMENTS_CACHE: dict[str, Any] = {"key": None, "rows": None}
STATUS_IMPORTED = "Importata"
STATUS_PLANNED = "Pianificata"
STATUS_DEPARTED = "Partita"
STATUS_CONFIRMED = "FTL Confermata"
STATUS_DELIVERED = "Consegnata"
AVAILABLE_CARRIERS = ["BRT", "KN", "GRENDI", "GD TRASPORTI", "DB", "FERCAM", "DACHSER", "DHL", "GEODIS", "DSV", "ALTRO"]
EMPTY_FILTER_VALUE = "__EMPTY__"
DATE_ONLY_COLUMNS = {
    "Late Ship Date",
    "Early Delivery Date",
    "Early Delivery Date Originale",
    "Data Consegna Tassativa",
    "Data Ship Minima SLA",
    "Prima Consegna SLA",
    "Data Pianifica",
    "Data Partenza",
    "Data Partenza Wave",
    "Data Scarico Prenotato",
}
GROUPAGE_MAIL_EXCLUDED_COLUMNS = {
    "Cliente GDO",
    "Pallet Fatturati",
    "Costo Attivo",
    "Extra Attivi Totale",
    "Extra Attivi Applicati",
    "Extra BRT Totale",
    "Costo Passivo",
    "Margine",
    "Miglior Vettore",
    "SLA Contratto",
    "Prima Consegna SLA",
    "Data Pianifica",
    "Pallet Manuali",
    "Data Scarico Prenotato",
    "Ora Scarico Prenotato",
    "Riferimento Booking Scarico",
    "Booking Scarico",
}
GROUPAGE_MAIL_ONLY_COLUMNS = ["Grand Total Shipment Ftp Vol m3"]
WAREHOUSE_CONTACT_MARKER = "CONTATTI MAGAZZINO"
WAREHOUSE_CONTACT_COLUMNS = [
    ("Responsabile Scarico", "Resp."),
    ("Mail", "Mail"),
    ("Telefono", "Tel."),
    ("Shipping Information", "Info scarico"),
    ("Note Scarico", "Note scarico"),
]


def ensure_seed_data_files() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if SEED_DIR.exists():
        for source in SEED_DIR.rglob("*"):
            if not source.is_file():
                continue
            relative = source.relative_to(SEED_DIR)
            target = DATA_DIR / relative
            target.parent.mkdir(parents=True, exist_ok=True)
            if not target.exists():
                shutil.copyfile(source, target)
    for filename in ("gdo_customers.csv", "customer_registry.csv", "carrier_tariffs.csv", "kn_delivery_template.xml"):
        source = APP_DIR / "data" / filename
        target = DATA_DIR / filename
        if source.exists() and source.resolve() != target.resolve() and not target.exists():
            shutil.copyfile(source, target)


ensure_seed_data_files()

DEFAULT_VTECH_PATH = DOWNLOADS_DIR / "006 - Outbound Report (PROD) (1).xlsx"
DEFAULT_ACTIVE_PATH = (
    Path(os.environ.get("LOCALAPPDATA", ""))
    / "Temp"
    / "07_Kuehne+Nagel_ Distribution rate card & leadtime.xlsx"
)
DEFAULT_BRT_PATH = (
    Path(os.environ.get("LOCALAPPDATA", ""))
    / "Temp"
    / "TARIFFA_123163_167410_0939666.PDF"
)


def available_gdo_customers_path() -> Path | None:
    return GDO_CUSTOMERS_PATH if GDO_CUSTOMERS_PATH.exists() else None


def available_fuel_settings_path() -> Path | None:
    return FUEL_SETTINGS_PATH if FUEL_SETTINGS_PATH.exists() else None


def strip_warehouse_contact_note(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    pattern = rf"\s*(?:\|\s*)?(?:\[)?{re.escape(WAREHOUSE_CONTACT_MARKER)}(?:\])?:?.*$"
    return re.sub(pattern, "", text, flags=re.IGNORECASE | re.DOTALL).strip(" |")


def clean_contact_value(value: Any) -> str:
    text = clean_text(value)
    if text.upper() in {"0", "0.0", "-", "--", "N/A", "NA", "N.D.", "ND", "NONE", "NULL"}:
        return ""
    return text


def build_warehouse_contact_note(row: dict[str, Any]) -> str:
    pieces: list[str] = []
    has_contact_field = False
    customer = clean_text(row.get("Ragione Sociale"))
    if customer:
        pieces.append(f"Magazzino: {customer}")
    for column, label in WAREHOUSE_CONTACT_COLUMNS:
        value = clean_contact_value(row.get(column))
        if value:
            pieces.append(f"{label} {value}")
            has_contact_field = True

    unique_pieces: list[str] = []
    seen: set[str] = set()
    for piece in pieces:
        key = piece.upper()
        if key not in seen:
            unique_pieces.append(piece)
            seen.add(key)
    if not has_contact_field:
        return ""
    return " | ".join(unique_pieces)


def load_warehouse_contact_registry(
    path: Path = CUSTOMER_REGISTRY_PATH,
) -> dict[str, Any]:
    registry: dict[str, Any] = {
        "customer_index": {},
        "address_index": {},
        "customer_entries": [],
        "address_entries": [],
    }
    if not path.exists():
        return registry

    def add_keys(target: str, value: Any, note: str) -> None:
        if not note:
            return
        index: dict[str, str] = registry[f"{target}_index"]
        entries: list[tuple[str, str]] = registry[f"{target}_entries"]
        for candidate in iter_registry_values(value):
            for key in customer_match_keys(candidate):
                if not key:
                    continue
                if key not in index:
                    index[key] = note
                    entries.append((key, note))

    with path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            contact_note = build_warehouse_contact_note(row)
            if not contact_note:
                continue
            add_keys("customer", row.get("Ragione Sociale"), contact_note)
            add_keys("customer", row.get("Customer Match Key"), contact_note)
            add_keys("address", row.get("Indirizzo Consegna"), contact_note)
            add_keys("address", row.get("Address Match Key"), contact_note)
    return registry


def match_warehouse_contact_note(row: dict[str, Any], registry: dict[str, Any]) -> str:
    def find_match(kind: str, values: list[Any]) -> str:
        index: dict[str, str] = registry.get(f"{kind}_index", {})
        entries: list[tuple[str, str]] = registry.get(f"{kind}_entries", [])
        candidate_keys: list[str] = []
        for value in values:
            for candidate in iter_registry_values(value):
                candidate_keys.extend(customer_match_keys(candidate))

        seen: set[str] = set()
        ordered_keys: list[str] = []
        for key in sorted(candidate_keys, key=len, reverse=True):
            if key and key not in seen:
                ordered_keys.append(key)
                seen.add(key)

        for key in ordered_keys:
            if key in index:
                return index[key]

        for key in ordered_keys:
            if len(key) < 8:
                continue
            for registry_key, contact_note in entries:
                if len(registry_key) < 8:
                    continue
                if key in registry_key or registry_key in key:
                    return contact_note
        return ""

    address_match = find_match(
        "address",
        [
            row.get("Route To Address"),
            row.get("Indirizzo Consegna"),
            row.get("Address"),
        ],
    )
    if address_match:
        return address_match

    return find_match(
        "customer",
        [
            row.get("Route to Customer"),
            row.get("Customer"),
            row.get("Cliente"),
        ],
    )


def apply_warehouse_contacts_to_shipments(rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    registry = load_warehouse_contact_registry()
    if not registry.get("customer_index") and not registry.get("address_index"):
        return

    for row in rows:
        base_note = strip_warehouse_contact_note(row.get("Note Text"))
        contact_note = match_warehouse_contact_note(row, registry)
        row["Contatti Magazzino"] = contact_note
        if contact_note:
            suffix = f"[{WAREHOUSE_CONTACT_MARKER}] {contact_note}"
            row["Note Text"] = f"{base_note} | {suffix}" if base_note else suffix
        else:
            row["Note Text"] = base_note

APP_BG = "#0b1220"
MAIN_BG = "#0f1b2d"
SIDEBAR_BG = "#07111f"
SIDEBAR_HOVER = "#102033"
SURFACE_BG = "#12243a"
SURFACE_SOFT = "#162b45"
TABLE_BG = "#0d1828"
TABLE_ALT = "#102033"
LINE_COLOR = "#263b55"
TEXT_COLOR = "#f8fafc"
MUTED_COLOR = "#94a3b8"
ACCENT_BLUE = "#38bdf8"
ACTION_BLUE = "#2563eb"
SUCCESS_GREEN = "#22c55e"
WARNING_AMBER = "#f59e0b"
DANGER_RED = "#ef4444"

DISPLAY_COLUMNS = [
    "Shipment",
    "Orders",
    "Provincia",
    "Route to Customer",
    "Cliente GDO",
    "Route To Address",
    "Service Level",
    "Carrier Scelto",
    "Freight Code",
    "Wave",
    "Data Pianifica",
    "Data Partenza Wave",
    "Attiva Urgente",
    "Prenotazione Scarico",
    "Note Text",
    "Late Ship Date",
    "Early Delivery Date",
    "Data Consegna Tassativa",
    "SLA Contratto",
    "Prima Consegna SLA",
    "Booking Scarico",
    "Theoretical Pallets",
    "Pallet Manuali",
    "Pallet Fatturati",
    "Grand Total Shipment Ftp Wgt Kg",
    "Costo Attivo",
    "Extra Attivi Totale",
    "Extra BRT Totale",
    "Costo Passivo",
    "Margine",
    "Miglior Vettore",
]

COLUMN_TITLES = {
    "Shipment": "Shipment",
    "Orders": "Ordine",
    "Provincia": "Prov.",
    "Route to Customer": "Cliente",
    "Cliente GDO": "GDO",
    "Route To Address": "Indirizzo consegna",
    "Service Level": "Servizio",
    "Carrier Scelto": "Vettore",
    "Freight Code": "Freight",
    "Freight Code Manuale": "Freight manuale",
    "Wave": "Wave",
    "Data Pianifica": "Partenza pianificata",
    "Data Partenza Wave": "Partenza wave",
    "Vettore Wave": "Vettore wave",
    "Tipo Wave": "Tipo wave",
    "Attiva Urgente": "Urgente",
    "Prenotazione Scarico": "Prenotazione",
    "Note Text": "Note",
    "Late Ship Date": "Late ship",
    "Early Delivery Date": "Early delivery",
    "Early Delivery Date Originale": "Early originale",
    "Data Consegna Tassativa": "Consegna tassativa",
    "Data Scarico Prenotato": "Scarico prenotato",
    "Ora Scarico Prenotato": "Ora scarico",
    "Riferimento Booking Scarico": "Rif. booking scarico",
    "Booking Scarico": "Booking scarico",
    "SLA Contratto": "SLA",
    "Prima Consegna SLA": "Min. consegna SLA",
    "Data Partenza": "Data spedito",
    "Data Consegna": "Data consegna",
    "Data Eliminazione": "Data eliminazione",
    "XML Consegna": "XML consegna",
    "Data Ship Minima SLA": "Min. ship SLA",
    "Preparazione SLA h": "Prep SLA h",
    "Transit SLA h": "Transit SLA h",
    "Tipo Transit SLA": "Transit SLA",
    "Dettaglio SLA": "Dettaglio SLA",
    "Stato": "Stato",
    "Theoretical Pallets": "Pallet",
    "Pallet Manuali": "Pallet manuali",
    "Pallet Fatturati": "Pallet fatt.",
    "Grand Total Shipment Ftp Wgt Kg": "Peso kg",
    "Grand Total Shipment Ftp Vol m3": "Volume m3",
    "Costo Attivo": "Attivo",
    "Extra Attivi Totale": "Extra attivi",
    "Extra Attivi Applicati": "Extra attivi applicati",
    "Extra BRT Totale": "Extra BRT",
    "Costo Passivo Manuale": "Passiva manuale",
    "Costo Passivo": "Passivo",
    "Margine": "Margine",
    "Miglior Vettore": "Miglior vettore",
    "Secondo Vettore": "Secondo vettore",
    "Terzo Vettore": "Terzo vettore",
    "Esito Margine": "Esito",
}

COLUMN_WIDTHS = {
    "Shipment": 100,
    "Orders": 90,
    "Provincia": 55,
    "Route to Customer": 300,
    "Cliente GDO": 70,
    "Route To Address": 340,
    "Service Level": 70,
    "Carrier Scelto": 80,
    "Freight Code": 80,
    "Freight Code Manuale": 100,
    "Wave": 180,
    "Data Pianifica": 120,
    "Data Partenza Wave": 120,
    "Attiva Urgente": 80,
    "Prenotazione Scarico": 145,
    "Note Text": 190,
    "Late Ship Date": 130,
    "Early Delivery Date": 130,
    "Early Delivery Date Originale": 140,
    "Data Consegna Tassativa": 145,
    "Data Scarico Prenotato": 135,
    "Ora Scarico Prenotato": 90,
    "Riferimento Booking Scarico": 150,
    "Booking Scarico": 190,
    "SLA Contratto": 130,
    "Prima Consegna SLA": 140,
    "Data Ship Minima SLA": 130,
    "Preparazione SLA h": 90,
    "Transit SLA h": 90,
    "Tipo Transit SLA": 90,
    "Dettaglio SLA": 260,
    "Stato": 95,
    "Theoretical Pallets": 75,
    "Pallet Manuali": 90,
    "Pallet Fatturati": 85,
    "Grand Total Shipment Ftp Wgt Kg": 90,
    "Costo Attivo": 80,
    "Extra Attivi Totale": 90,
    "Extra BRT Totale": 80,
    "Costo Passivo": 80,
    "Margine": 80,
    "Miglior Vettore": 120,
    "Secondo Vettore": 120,
    "Terzo Vettore": 120,
    "Esito Margine": 95,
}


def format_number(value: Any) -> str:
    number = to_float(value)
    if number is None:
        return clean_text(value)
    return f"{number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def ceil_pallets(value: Any) -> int:
    number = to_float(value)
    if number is None:
        return 0
    return max(0, math.ceil(number))


def format_date_only(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    text = clean_text(value)
    if not text:
        return ""

    date_part = text.replace("T", " ").split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_part, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return date_part


def parse_date_value(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    date_part = text.replace("T", " ").split(" ")[0].replace(".", "/")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_part, fmt).date()
        except ValueError:
            continue
    return None


def iso_date_text(value: Any) -> str:
    parsed = parse_date_value(value)
    if parsed:
        return parsed.strftime("%Y-%m-%d")
    return clean_text(value)


def wave_departure_iso(row: dict[str, Any]) -> str:
    departure = iso_date_text(row.get("Data Partenza Wave"))
    if departure:
        return departure
    return parse_wave_departure_date(row.get("Wave"))


def day_after_iso(value: Any | None = None) -> str:
    parsed = parse_date_value(value) if value else None
    base_date = parsed or date.today()
    return (base_date + timedelta(days=1)).strftime("%Y-%m-%d")


def billing_reference_date(row: dict[str, Any]) -> date | None:
    for column in ("Late Ship Date", "Data Consegna", "Data Consegna Tassativa", "Early Delivery Date", "Integration Date"):
        parsed = parse_date_value(row.get(column))
        if parsed:
            return parsed
    return None


def display_value(row: dict[str, Any], column: str) -> str:
    if column in {
        "Costo Attivo",
        "Costo Passivo",
        "Costo Passivo Base BRT",
        "Costo Passivo Manuale",
        "Margine",
        "Extra Attivi Totale",
        "Extra BRT Totale",
    }:
        value = format_number(row.get(column))
        return value if not value else f"EUR {value}"
    if column in {"Theoretical Pallets", "Pallet Originali", "Pallet Manuali", "Grand Total Shipment Ftp Wgt Kg", "Grand Total Shipment Ftp Vol m3"}:
        return format_number(row.get(column))
    if column in {"Preparazione SLA h", "Transit SLA h"}:
        number = to_float(row.get(column))
        if number is None:
            return clean_text(row.get(column))
        return str(int(number)) if number.is_integer() else format_number(number)
    if column == "Booking Scarico":
        return clean_text(row.get(column)) or build_unload_booking_text(
            row.get("Data Scarico Prenotato"),
            row.get("Ora Scarico Prenotato"),
            row.get("Riferimento Booking Scarico"),
        )
    if column in DATE_ONLY_COLUMNS:
        return format_date_only(row.get(column))
    return clean_text(row.get(column))


def serialize_shipment_payload(row: dict[str, Any]) -> str:
    payload: dict[str, Any] = {}
    for key in SHIPMENT_COLUMNS:
        if key in DATE_ONLY_COLUMNS:
            payload[key] = format_date_only(row.get(key))
        else:
            payload[key] = serialize(row.get(key))
    return json.dumps(payload, ensure_ascii=False)


def apply_manual_passive_to_row(row: dict[str, Any]) -> None:
    manual_cost = to_float(row.get("Costo Passivo Manuale"))
    if manual_cost is None:
        row["Costo Passivo Manuale"] = ""
        return

    manual_cost = round(manual_cost, 2)
    carrier = clean_text(row.get("Carrier Scelto")) or clean_text(row.get("Carrier Originale")) or "vettore"
    row["Costo Passivo Manuale"] = manual_cost
    row["Costo Passivo"] = manual_cost
    row["Costo Passivo Base BRT"] = manual_cost
    row["Extra BRT Totale"] = 0
    row["Extra BRT Applicati"] = "Passiva manuale: extra inclusi nel prezzo concordato."
    row["Peso Tariffabile BRT Kg"] = ""
    row["Tariffa Passiva Applicata"] = f"Passiva manuale concordata con {carrier}: EUR {format_number(manual_cost)}"
    row["Miglior Vettore"] = f"{carrier} EUR {manual_cost:.2f}"
    row["Secondo Vettore"] = ""
    row["Terzo Vettore"] = ""

    active_cost = to_float(row.get("Costo Attivo"))
    if active_cost is not None:
        margin = round(active_cost - manual_cost, 2)
        row["Margine"] = margin
        row["Esito Margine"] = "Guadagno" if margin >= 0 else "Perdita"


def clear_manual_passive_from_row(row: dict[str, Any]) -> None:
    row["Costo Passivo Manuale"] = ""


def apply_manual_pallets_to_row(row: dict[str, Any], pallets: Any) -> float:
    pallet_value = to_float(pallets)
    if pallet_value is None:
        raise ValueError("Inserisci un numero di bancali valido.")
    if pallet_value <= 0:
        raise ValueError("I bancali manuali devono essere maggiori di zero.")

    pallet_value = round(pallet_value, 3)
    if not clean_text(row.get("Pallet Originali")):
        row["Pallet Originali"] = row.get("Theoretical Pallets")
    row["Pallet Manuali"] = pallet_value
    row["Theoretical Pallets"] = pallet_value
    return pallet_value


def clear_manual_pallets_from_row(row: dict[str, Any]) -> None:
    original = row.get("Pallet Originali")
    if clean_text(original):
        row["Theoretical Pallets"] = original
    row["Pallet Manuali"] = ""


def normalize_unload_date(value: Any) -> str:
    parsed = parse_date_value(value)
    if parsed is None:
        raise ValueError("Inserisci una data scarico valida.")
    return parsed.strftime("%Y-%m-%d")


def normalize_unload_time(value: Any) -> str:
    text = clean_text(value).replace(".", ":")
    if not text:
        return ""
    for fmt in ("%H:%M", "%H%M", "%H"):
        try:
            parsed = datetime.strptime(text, fmt).time()
            return parsed.strftime("%H:%M")
        except ValueError:
            continue
    raise ValueError("Inserisci un'ora scarico valida.")


def build_unload_booking_text(unload_date: Any, unload_time: Any = "", booking_ref: Any = "") -> str:
    date_text = format_date_only(unload_date)
    time_text = normalize_unload_time(unload_time) if clean_text(unload_time) else ""
    ref_text = clean_text(booking_ref)
    parts = []
    if date_text:
        parts.append(date_text)
    if time_text:
        parts.append(f"ore {time_text}")
    if ref_text:
        parts.append(f"rif. {ref_text}")
    return " - ".join(parts)


def sync_unload_booking_fields(row: dict[str, Any]) -> None:
    row["Ora Scarico Prenotato"] = normalize_unload_time(row.get("Ora Scarico Prenotato")) if clean_text(row.get("Ora Scarico Prenotato")) else ""
    row["Riferimento Booking Scarico"] = clean_text(row.get("Riferimento Booking Scarico"))
    row["Booking Scarico"] = build_unload_booking_text(
        row.get("Data Scarico Prenotato"),
        row.get("Ora Scarico Prenotato"),
        row.get("Riferimento Booking Scarico"),
    )


def has_unload_booking(row: dict[str, Any]) -> bool:
    return any(
        clean_text(row.get(column))
        for column in (
            "Data Scarico Prenotato",
            "Ora Scarico Prenotato",
            "Riferimento Booking Scarico",
            "Booking Scarico",
        )
    )


def apply_unload_booking_status(row: dict[str, Any], persisted_status: Any = "") -> str:
    if has_unload_booking(row):
        status = "Prenotato"
    else:
        status = clean_text(persisted_status) or classify_unloading_booking(row.get("Freight Code"))
    row["Prenotazione Scarico"] = status
    return status


def is_brt_groupage(row: dict[str, Any]) -> bool:
    return clean_text(row.get("Tipo Servizio")) == "Groupage - BRT LTL"


def active_urgent_enabled(value: Any) -> bool:
    return clean_text(value).upper() in {"1", "SI", "S", "YES", "Y", "TRUE", "X"}


def apply_active_urgent_to_row(row: dict[str, Any], enabled: bool) -> None:
    row["Attiva Urgente"] = "SI" if enabled else ""


def normalize_service_level(value: Any) -> str:
    service_level = clean_text(value).upper()
    if service_level not in {"LTL", "FTL"}:
        raise ValueError("Scegli un service level valido: LTL o FTL.")
    return service_level


def apply_manual_service_level_to_row(row: dict[str, Any], service_level: Any) -> None:
    level = normalize_service_level(service_level)
    row["Service Level Manuale"] = level
    row["Service Level"] = level
    if level == "LTL":
        row["Carrier Scelto"] = "BRT"
        row["Tipo Servizio"] = "Groupage - BRT LTL"
    else:
        current_carrier = clean_text(row.get("Carrier Scelto")).upper()
        if current_carrier == "BRT":
            row["Carrier Scelto"] = clean_text(row.get("Carrier Originale")) or "KN"
        row["Tipo Servizio"] = "Diretta - KN LTL/FTL"


def normalize_manual_freight_code(value: Any) -> str:
    freight_code = clean_text(value).upper()
    if freight_code not in {"DKL", "DKV"}:
        raise ValueError("Scegli un freight valido: DKL o DKV.")
    return freight_code


def normalize_required_delivery_date(value: Any) -> str:
    parsed = parse_date_value(value)
    if parsed is None:
        raise ValueError("Inserisci una data tassativa valida.")
    return parsed.strftime("%Y-%m-%d")


def freight_code_tokens(row: dict[str, Any]) -> set[str]:
    return {
        token
        for token in re.split(r"[^A-Z0-9]+", clean_text(row.get("Freight Code")).upper())
        if token
    }


def is_dkv_order(row: dict[str, Any]) -> bool:
    return "DKV" in freight_code_tokens(row) or clean_text(row.get("Freight Code Manuale")).upper() == "DKV"


def restore_early_delivery_if_possible(row: dict[str, Any]) -> None:
    original = clean_text(row.get("Early Delivery Date Originale"))
    if original:
        row["Early Delivery Date"] = original
    elif not clean_text(row.get("Early Delivery Date")):
        row["Early Delivery Date"] = clean_text(row.get("Late Delivery Date"))


def hide_early_delivery_for_dkv(row: dict[str, Any]) -> None:
    current_early = clean_text(row.get("Early Delivery Date"))
    if current_early and not clean_text(row.get("Early Delivery Date Originale")):
        row["Early Delivery Date Originale"] = current_early
    row["Early Delivery Date"] = ""


def apply_required_delivery_date_to_row(row: dict[str, Any], required_date: Any, clear: bool = False) -> str:
    if clear:
        if is_dkv_order(row):
            raise ValueError("Una DKV deve avere una data tassativa. Cambiala in DKL per toglierla.")
        row["Data Consegna Tassativa"] = ""
        return ""
    normalized = normalize_required_delivery_date(required_date)
    row["Data Consegna Tassativa"] = normalized
    if is_dkv_order(row):
        hide_early_delivery_for_dkv(row)
    return normalized


def apply_manual_freight_code_to_row(
    row: dict[str, Any],
    freight_code: Any,
    required_date: Any = "",
    clear_required_for_dkl: bool = True,
) -> tuple[str, str]:
    manual_freight_code = normalize_manual_freight_code(freight_code)
    row["Freight Code Manuale"] = manual_freight_code
    row["Freight Code"] = manual_freight_code
    row["Prenotazione Scarico"] = classify_unloading_booking(manual_freight_code)

    if manual_freight_code == "DKV":
        normalized_required_date = normalize_required_delivery_date(
            required_date or row.get("Data Consegna Tassativa")
        )
        row["Data Consegna Tassativa"] = normalized_required_date
        hide_early_delivery_for_dkv(row)
    else:
        if clear_required_for_dkl:
            normalized_required_date = ""
            row["Data Consegna Tassativa"] = ""
        else:
            normalized_required_date = (
                normalize_required_delivery_date(required_date)
                if clean_text(required_date)
                else clean_text(row.get("Data Consegna Tassativa"))
            )
            row["Data Consegna Tassativa"] = normalized_required_date
        restore_early_delivery_if_possible(row)
    return manual_freight_code, normalized_required_date


def load_settings() -> dict[str, Any]:
    if not SETTINGS_PATH.exists():
        return {}
    try:
        settings = json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    if not isinstance(settings, dict):
        return {}
    replacements = {
        "{APP_DIR}": str(APP_DIR),
        "{DATA_DIR}": str(DATA_DIR),
        "{OUTPUT_DIR}": str(OUTPUT_DIR),
        "{DOWNLOADS_DIR}": str(DOWNLOADS_DIR),
    }
    for key in ("vtech_path", "active_rates_path", "brt_passive_path"):
        value = settings.get(key)
        if not isinstance(value, str):
            continue
        for token, replacement in replacements.items():
            value = value.replace(token, replacement)
        settings[key] = value
    return settings


def save_settings(settings: dict[str, Any]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    SETTINGS_PATH.write_text(json.dumps(settings, indent=2, ensure_ascii=False), encoding="utf-8")


def parse_percent_value(value: Any, default: float | None = None) -> float | None:
    text = clean_text(value).replace("%", "").replace(",", ".")
    if not text:
        return default
    try:
        number = float(text)
    except ValueError:
        return default
    return max(0.0, round(number, 3))


def load_monthly_fuel_settings() -> dict[str, dict[str, float]]:
    if not FUEL_SETTINGS_PATH.exists():
        return {}
    try:
        raw = json.loads(FUEL_SETTINGS_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    settings: dict[str, dict[str, float]] = {}
    for month, values in raw.items():
        if not re.fullmatch(r"\d{4}-\d{2}", clean_text(month)) or not isinstance(values, dict):
            continue
        month_settings: dict[str, float] = {}
        active = parse_percent_value(values.get("active"))
        passive = parse_percent_value(values.get("passive"))
        if active is not None:
            month_settings["active"] = active
        if passive is not None:
            month_settings["passive"] = passive
        if month_settings:
            settings[clean_text(month)] = month_settings
    return settings


def save_monthly_fuel_settings(month: str, active: Any, passive: Any) -> dict[str, dict[str, float]]:
    month_key = clean_text(month)
    if not re.fullmatch(r"\d{4}-\d{2}", month_key):
        raise ValueError("Mese fuel non valido.")
    settings = load_monthly_fuel_settings()
    active_value = parse_percent_value(active, default=0.0)
    passive_value = parse_percent_value(passive, default=2.0)
    settings[month_key] = {
        "active": active_value if active_value is not None else 0.0,
        "passive": passive_value if passive_value is not None else 2.0,
    }
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    FUEL_SETTINGS_PATH.write_text(json.dumps(settings, indent=2, ensure_ascii=False), encoding="utf-8")
    return settings


def ensure_brt_extra_flags_template(shipments: list[dict[str, Any]]) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    existing: dict[str, dict[str, str]] = {}
    if BRT_EXTRA_FLAGS_PATH.exists():
        with BRT_EXTRA_FLAGS_PATH.open("r", newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            for row in reader:
                shipment = clean_text(row.get("Shipment"))
                if shipment:
                    existing[shipment] = {column: clean_text(row.get(column)) for column in BRT_EXTRA_FLAG_COLUMNS}

    changed = False
    for row in shipments:
        shipment = clean_text(row.get("Shipment"))
        if shipment and shipment not in existing:
            existing[shipment] = {column: "" for column in BRT_EXTRA_FLAG_COLUMNS}
            existing[shipment]["Shipment"] = shipment
            changed = True

    if changed or not BRT_EXTRA_FLAGS_PATH.exists():
        with BRT_EXTRA_FLAGS_PATH.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=BRT_EXTRA_FLAG_COLUMNS)
            writer.writeheader()
            for shipment in sorted(existing):
                writer.writerow(existing[shipment])


def init_db(db_path: Path = DB_PATH) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS shipments (
                shipment TEXT PRIMARY KEY,
                orders_text TEXT,
                service_type TEXT,
                carrier_originale TEXT,
                carrier_scelto TEXT,
                manual_carrier TEXT,
                manual_service_level TEXT,
                manual_freight_code TEXT,
                service_level TEXT,
                freight_code TEXT,
                booking_status TEXT,
                province TEXT,
                customer TEXT,
                active_cost REAL,
                passive_cost REAL,
                manual_passive_cost REAL,
                manual_pallets REAL,
                margin REAL,
                status TEXT NOT NULL DEFAULT 'Importata',
                planned_at TEXT,
                departed_at TEXT,
                unload_date TEXT,
                unload_time TEXT,
                unload_booking_ref TEXT,
                required_delivery_date TEXT,
                delivered_at TEXT,
                delivery_xml_path TEXT,
                source_file TEXT,
                imported_at TEXT NOT NULL,
                payload_json TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS imported_files (
                file_path TEXT PRIMARY KEY,
                size_bytes INTEGER NOT NULL,
                mtime_ns INTEGER NOT NULL,
                imported_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS deleted_shipments (
                shipment TEXT PRIMARY KEY,
                deleted_at TEXT NOT NULL,
                payload_json TEXT
            )
            """
        )

        existing_columns = {
            row[1] for row in conn.execute("PRAGMA table_info(shipments)").fetchall()
        }
        if "freight_code" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN freight_code TEXT")
        if "booking_status" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN booking_status TEXT")
        if "delivery_xml_path" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN delivery_xml_path TEXT")
        if "manual_carrier" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN manual_carrier TEXT")
        if "manual_service_level" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN manual_service_level TEXT")
        if "manual_freight_code" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN manual_freight_code TEXT")
        if "departed_at" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN departed_at TEXT")
        if "planned_at" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN planned_at TEXT")
        if "manual_passive_cost" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN manual_passive_cost REAL")
        if "manual_pallets" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN manual_pallets REAL")
        if "unload_date" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN unload_date TEXT")
        if "unload_time" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN unload_time TEXT")
        if "unload_booking_ref" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN unload_booking_ref TEXT")
        if "required_delivery_date" not in existing_columns:
            conn.execute("ALTER TABLE shipments ADD COLUMN required_delivery_date TEXT")


def file_fingerprint(path: Path) -> tuple[int, int]:
    stat = path.stat()
    return stat.st_size, stat.st_mtime_ns


def is_imported_file_current(path: Path, db_path: Path = DB_PATH) -> bool:
    if not path.exists():
        return False
    init_db(db_path)
    size, mtime_ns = file_fingerprint(path)
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            "SELECT size_bytes, mtime_ns FROM imported_files WHERE file_path = ?",
            (str(path),),
        ).fetchone()
    return bool(row and row[0] == size and row[1] == mtime_ns)


def mark_file_imported(path: Path, db_path: Path = DB_PATH) -> None:
    size, mtime_ns = file_fingerprint(path)
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            INSERT INTO imported_files (file_path, size_bytes, mtime_ns, imported_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(file_path) DO UPDATE SET
                size_bytes = excluded.size_bytes,
                mtime_ns = excluded.mtime_ns,
                imported_at = excluded.imported_at
            """,
            (str(path), size, mtime_ns, imported_at),
        )


def file_cache_signature(path: Path | None) -> tuple[str, int, int] | None:
    if path is None:
        return None
    try:
        stat = path.stat()
    except OSError:
        return (str(path), -1, -1)
    return (str(path), stat.st_size, stat.st_mtime_ns)


def shipments_cache_key(db_path: Path = DB_PATH) -> tuple[Any, ...]:
    settings = load_settings()
    active_path_text = settings.get("active_rates_path")
    brt_path_text = settings.get("brt_passive_path")
    active_path = Path(active_path_text) if active_path_text else DEFAULT_ACTIVE_PATH
    brt_path = Path(brt_path_text) if brt_path_text else DEFAULT_BRT_PATH
    return (
        file_cache_signature(db_path),
        file_cache_signature(SETTINGS_PATH),
        file_cache_signature(active_path),
        file_cache_signature(brt_path),
        file_cache_signature(BRT_EXTRA_FLAGS_PATH),
        file_cache_signature(available_gdo_customers_path()),
        file_cache_signature(available_fuel_settings_path()),
        file_cache_signature(default_carrier_tariffs_path()),
    )


def clear_shipments_cache() -> None:
    with SHIPMENTS_CACHE_LOCK:
        SHIPMENTS_CACHE["key"] = None
        SHIPMENTS_CACHE["rows"] = None


def load_manual_carriers(db_path: Path = DB_PATH) -> dict[str, str]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        return {
            shipment: carrier
            for shipment, carrier in conn.execute(
                "SELECT shipment, manual_carrier FROM shipments WHERE manual_carrier IS NOT NULL AND manual_carrier <> ''"
            ).fetchall()
        }


def normalized_order_key(value: Any) -> str:
    text = clean_text(value).upper()
    if not text:
        return ""
    tokens = [token for token in re.split(r"[^A-Z0-9]+", text) if token]
    return "|".join(sorted(tokens))


def load_manual_carriers_by_order(db_path: Path = DB_PATH) -> dict[str, str]:
    init_db(db_path)
    carriers: dict[str, str] = {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT orders_text, manual_carrier
            FROM shipments
            WHERE manual_carrier IS NOT NULL AND manual_carrier <> ''
            """
        ).fetchall()
    for orders_text, carrier in rows:
        key = normalized_order_key(orders_text)
        if key and key not in carriers:
            carriers[key] = clean_text(carrier)
    return carriers


def load_manual_passive_costs(db_path: Path = DB_PATH) -> dict[str, float]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        return {
            shipment: float(cost)
            for shipment, cost in conn.execute(
                "SELECT shipment, manual_passive_cost FROM shipments WHERE manual_passive_cost IS NOT NULL"
            ).fetchall()
            if clean_text(shipment) and cost is not None
        }


def load_manual_passive_costs_by_order(db_path: Path = DB_PATH) -> dict[str, float]:
    init_db(db_path)
    costs: dict[str, float] = {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT orders_text, manual_passive_cost
            FROM shipments
            WHERE manual_passive_cost IS NOT NULL
            """
        ).fetchall()
    for orders_text, cost in rows:
        key = normalized_order_key(orders_text)
        if key and key not in costs and cost is not None:
            costs[key] = float(cost)
    return costs


def load_manual_pallets(db_path: Path = DB_PATH) -> dict[str, float]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        return {
            shipment: float(pallets)
            for shipment, pallets in conn.execute(
                "SELECT shipment, manual_pallets FROM shipments WHERE manual_pallets IS NOT NULL"
            ).fetchall()
            if clean_text(shipment) and pallets is not None
        }


def load_manual_pallets_by_order(db_path: Path = DB_PATH) -> dict[str, float]:
    init_db(db_path)
    pallets_by_order: dict[str, float] = {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT orders_text, manual_pallets
            FROM shipments
            WHERE manual_pallets IS NOT NULL
            """
        ).fetchall()
    for orders_text, pallets in rows:
        key = normalized_order_key(orders_text)
        if key and key not in pallets_by_order and pallets is not None:
            pallets_by_order[key] = float(pallets)
    return pallets_by_order


def load_manual_service_levels(db_path: Path = DB_PATH) -> dict[str, str]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        return {
            shipment: service_level
            for shipment, service_level in conn.execute(
                "SELECT shipment, manual_service_level FROM shipments WHERE manual_service_level IS NOT NULL AND manual_service_level <> ''"
            ).fetchall()
            if clean_text(shipment) and clean_text(service_level)
        }


def load_manual_service_levels_by_order(db_path: Path = DB_PATH) -> dict[str, str]:
    init_db(db_path)
    service_levels: dict[str, str] = {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT orders_text, manual_service_level
            FROM shipments
            WHERE manual_service_level IS NOT NULL AND manual_service_level <> ''
            """
        ).fetchall()
    for orders_text, service_level in rows:
        key = normalized_order_key(orders_text)
        if key and key not in service_levels:
            service_levels[key] = clean_text(service_level).upper()
    return service_levels


def load_manual_freight_codes(db_path: Path = DB_PATH) -> dict[str, tuple[str, str]]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT shipment, manual_freight_code, required_delivery_date
            FROM shipments
            WHERE manual_freight_code IS NOT NULL AND manual_freight_code <> ''
            """
        ).fetchall()
    return {
        clean_text(shipment): (clean_text(freight_code).upper(), clean_text(required_date))
        for shipment, freight_code, required_date in rows
        if clean_text(shipment) and clean_text(freight_code)
    }


def load_manual_freight_codes_by_order(db_path: Path = DB_PATH) -> dict[str, tuple[str, str]]:
    init_db(db_path)
    freight_codes: dict[str, tuple[str, str]] = {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT orders_text, manual_freight_code, required_delivery_date
            FROM shipments
            WHERE manual_freight_code IS NOT NULL AND manual_freight_code <> ''
            """
        ).fetchall()
    for orders_text, freight_code, required_date in rows:
        key = normalized_order_key(orders_text)
        if key and key not in freight_codes:
            freight_codes[key] = (clean_text(freight_code).upper(), clean_text(required_date))
    return freight_codes


def load_active_urgent_flags(db_path: Path = DB_PATH) -> dict[str, bool]:
    init_db(db_path)
    flags: dict[str, bool] = {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute("SELECT shipment, payload_json FROM shipments").fetchall()
    for shipment, payload_json in rows:
        try:
            payload = json.loads(payload_json or "{}")
        except json.JSONDecodeError:
            continue
        if clean_text(shipment) and active_urgent_enabled(payload.get("Attiva Urgente")):
            flags[clean_text(shipment)] = True
    return flags


def load_active_urgent_flags_by_order(db_path: Path = DB_PATH) -> dict[str, bool]:
    init_db(db_path)
    flags: dict[str, bool] = {}
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute("SELECT orders_text, payload_json FROM shipments").fetchall()
    for orders_text, payload_json in rows:
        try:
            payload = json.loads(payload_json or "{}")
        except json.JSONDecodeError:
            continue
        key = normalized_order_key(orders_text or payload.get("Orders"))
        if key and key not in flags and active_urgent_enabled(payload.get("Attiva Urgente")):
            flags[key] = True
    return flags


STATUS_PRIORITY = {
    STATUS_IMPORTED: 0,
    STATUS_PLANNED: 1,
    STATUS_DEPARTED: 2,
    STATUS_CONFIRMED: 3,
    STATUS_DELIVERED: 4,
}


def merge_duplicate_order_shipments(db_path: Path = DB_PATH) -> int:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT shipment, orders_text, status, planned_at, departed_at, unload_date, unload_time, unload_booking_ref,
                   delivered_at, delivery_xml_path,
                   source_file, imported_at, payload_json, manual_carrier, manual_passive_cost, manual_service_level,
                   manual_freight_code, required_delivery_date, manual_pallets
            FROM shipments
            """
        ).fetchall()

        groups: dict[str, list[dict[str, Any]]] = {}
        for item in rows:
            key = normalized_order_key(item[1])
            if not key:
                continue
            groups.setdefault(key, []).append(
                {
                    "shipment": clean_text(item[0]),
                    "orders_text": clean_text(item[1]),
                    "status": clean_text(item[2]) or STATUS_IMPORTED,
                    "planned_at": clean_text(item[3]),
                    "departed_at": clean_text(item[4]),
                    "unload_date": clean_text(item[5]),
                    "unload_time": clean_text(item[6]),
                    "unload_booking_ref": clean_text(item[7]),
                    "delivered_at": clean_text(item[8]),
                    "delivery_xml_path": clean_text(item[9]),
                    "source_file": clean_text(item[10]),
                    "imported_at": clean_text(item[11]),
                    "payload_json": item[12],
                    "manual_carrier": clean_text(item[13]),
                    "manual_passive_cost": item[14],
                    "manual_service_level": clean_text(item[15]),
                    "manual_freight_code": clean_text(item[16]),
                    "required_delivery_date": clean_text(item[17]),
                    "manual_pallets": item[18],
                }
            )

        merged = 0
        for duplicates in groups.values():
            if len(duplicates) <= 1:
                continue
            duplicates.sort(key=lambda row: (row["imported_at"], row["shipment"]), reverse=True)
            keep = duplicates[0]
            status_row = max(
                duplicates,
                key=lambda row: (STATUS_PRIORITY.get(row["status"], 0), row["imported_at"]),
            )
            manual_carrier = next((row["manual_carrier"] for row in duplicates if row["manual_carrier"]), "")
            manual_passive_cost = next(
                (row["manual_passive_cost"] for row in duplicates if row["manual_passive_cost"] is not None),
                None,
            )
            manual_service_level = next(
                (row["manual_service_level"] for row in duplicates if row["manual_service_level"]),
                "",
            )
            manual_freight_code, required_delivery_date = next(
                (
                    (row["manual_freight_code"], row["required_delivery_date"])
                    for row in duplicates
                    if row["manual_freight_code"]
                ),
                ("", ""),
            )
            manual_pallets = next(
                (row["manual_pallets"] for row in duplicates if row["manual_pallets"] is not None),
                None,
            )
            try:
                payload = json.loads(keep["payload_json"] or "{}")
            except json.JSONDecodeError:
                payload = {}
            payload["Shipment"] = keep["shipment"]
            payload["Stato"] = status_row["status"] or STATUS_IMPORTED
            payload["Data Pianifica"] = status_row["planned_at"]
            payload["Data Partenza"] = status_row["departed_at"]
            payload["Data Scarico Prenotato"] = status_row["unload_date"]
            payload["Ora Scarico Prenotato"] = status_row["unload_time"]
            payload["Riferimento Booking Scarico"] = status_row["unload_booking_ref"]
            sync_unload_booking_fields(payload)
            payload["Data Consegna"] = status_row["delivered_at"]
            payload["XML Consegna"] = status_row["delivery_xml_path"]
            if manual_carrier:
                payload["Carrier Scelto"] = manual_carrier
            if manual_passive_cost is not None:
                payload["Costo Passivo Manuale"] = manual_passive_cost
                apply_manual_passive_to_row(payload)
            if manual_service_level:
                apply_manual_service_level_to_row(payload, manual_service_level)
            if manual_freight_code:
                apply_manual_freight_code_to_row(
                    payload,
                    manual_freight_code,
                    required_delivery_date,
                    clear_required_for_dkl=not bool(required_delivery_date),
                )
            if manual_pallets is not None:
                apply_manual_pallets_to_row(payload, manual_pallets)
            booking_status = apply_unload_booking_status(payload)

            conn.execute(
                """
                UPDATE shipments
                SET status = ?,
                    planned_at = ?,
                    departed_at = ?,
                    unload_date = ?,
                    unload_time = ?,
                    unload_booking_ref = ?,
                    delivered_at = ?,
                    delivery_xml_path = ?,
                    manual_carrier = ?,
                    manual_passive_cost = ?,
                    manual_service_level = ?,
                    manual_freight_code = ?,
                    required_delivery_date = ?,
                    manual_pallets = ?,
                    booking_status = ?,
                    payload_json = ?
                WHERE shipment = ?
                """,
                (
                    payload["Stato"],
                    status_row["planned_at"] or None,
                    status_row["departed_at"] or None,
                    status_row["unload_date"] or None,
                    status_row["unload_time"] or None,
                    status_row["unload_booking_ref"] or None,
                    status_row["delivered_at"] or None,
                    status_row["delivery_xml_path"] or None,
                    manual_carrier,
                    manual_passive_cost,
                    manual_service_level,
                    manual_freight_code,
                    required_delivery_date or None,
                    manual_pallets,
                    booking_status,
                    serialize_shipment_payload(payload),
                    keep["shipment"],
                ),
            )

            for duplicate in duplicates[1:]:
                conn.execute("DELETE FROM shipments WHERE shipment = ?", (duplicate["shipment"],))
                merged += 1
    return merged


def save_shipments_to_db(shipments: list[dict[str, Any]], source_file: Path, db_path: Path = DB_PATH) -> dict[str, int]:
    init_db(db_path)
    merge_duplicate_order_shipments(db_path)
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    counters = {"inserted": 0, "updated": 0, "total": 0, "skipped_deleted": 0, "merged_by_order": 0}
    with sqlite3.connect(db_path) as conn:
        deleted_rows = conn.execute("SELECT shipment, payload_json FROM deleted_shipments").fetchall()
        deleted_shipments = {clean_text(row[0]) for row in deleted_rows if clean_text(row[0])}
        deleted_orders: set[str] = set()
        for _shipment, payload_json in deleted_rows:
            try:
                payload = json.loads(payload_json or "{}")
            except json.JSONDecodeError:
                payload = {}
            order_key = normalized_order_key(payload.get("Orders"))
            if order_key:
                deleted_orders.add(order_key)

        existing_rows = conn.execute(
            """
            SELECT shipment, orders_text, status, delivered_at, manual_carrier,
                   manual_passive_cost, manual_service_level, manual_freight_code,
                   required_delivery_date, manual_pallets, unload_date, unload_time,
                   unload_booking_ref, planned_at
            FROM shipments
            """
        ).fetchall()
        existing_by_shipment = {clean_text(row[0]): row for row in existing_rows if clean_text(row[0])}
        existing_by_order: dict[str, tuple[Any, ...]] = {}
        for existing_row in existing_rows:
            order_key = normalized_order_key(existing_row[1])
            if order_key and order_key not in existing_by_order:
                existing_by_order[order_key] = existing_row

        for row in shipments:
            shipment = clean_text(row.get("Shipment"))
            order_key = normalized_order_key(row.get("Orders"))
            if not shipment:
                continue
            if shipment in deleted_shipments or (order_key and order_key in deleted_orders):
                counters["skipped_deleted"] += 1
                continue
            counters["total"] += 1

            existing = existing_by_shipment.get(shipment)
            existing_shipment = shipment
            if existing is None and order_key:
                existing = existing_by_order.get(order_key)
                if existing:
                    existing_shipment = clean_text(existing[0])
                    if existing_shipment and existing_shipment != shipment:
                        conn.execute("DELETE FROM shipments WHERE shipment = ?", (existing_shipment,))
                        counters["merged_by_order"] += 1

            if existing:
                counters["updated"] += 1
            else:
                counters["inserted"] += 1
            status = existing[2] if existing and existing[2] else STATUS_IMPORTED
            delivered_at = existing[3] if existing else None
            manual_carrier = existing[4] if existing and existing[4] else ""
            manual_passive_cost = (
                existing[5]
                if existing and existing[5] is not None
                else to_float(row.get("Costo Passivo Manuale"))
            )
            manual_service_level = existing[6] if existing and clean_text(existing[6]) else ""
            manual_freight_code = existing[7] if existing and clean_text(existing[7]) else clean_text(row.get("Freight Code Manuale"))
            required_delivery_date = existing[8] if existing and clean_text(existing[8]) else clean_text(row.get("Data Consegna Tassativa"))
            manual_pallets = (
                existing[9]
                if existing and existing[9] is not None
                else to_float(row.get("Pallet Manuali"))
            )
            unload_date = existing[10] if existing and clean_text(existing[10]) else clean_text(row.get("Data Scarico Prenotato"))
            unload_time = existing[11] if existing and clean_text(existing[11]) else clean_text(row.get("Ora Scarico Prenotato"))
            unload_booking_ref = existing[12] if existing and clean_text(existing[12]) else clean_text(row.get("Riferimento Booking Scarico"))
            planned_at = existing[13] if existing and clean_text(existing[13]) else clean_text(row.get("Data Pianifica"))
            if status == STATUS_PLANNED and not clean_text(planned_at):
                planned_at = wave_departure_iso(row) or date.today().strftime("%Y-%m-%d")
            row["Data Pianifica"] = planned_at
            if manual_carrier:
                row["Carrier Scelto"] = manual_carrier
            if manual_service_level:
                apply_manual_service_level_to_row(row, manual_service_level)
            if manual_freight_code:
                manual_freight_code, required_delivery_date = apply_manual_freight_code_to_row(
                    row,
                    manual_freight_code,
                    required_delivery_date,
                    clear_required_for_dkl=not bool(required_delivery_date),
                )
            elif required_delivery_date:
                required_delivery_date = apply_required_delivery_date_to_row(row, required_delivery_date)
            if manual_pallets is not None:
                manual_pallets = apply_manual_pallets_to_row(row, manual_pallets)
            if unload_date:
                row["Data Scarico Prenotato"] = unload_date
            row["Ora Scarico Prenotato"] = unload_time
            row["Riferimento Booking Scarico"] = unload_booking_ref
            sync_unload_booking_fields(row)
            booking_status = apply_unload_booking_status(row)
            if manual_passive_cost is not None:
                row["Costo Passivo Manuale"] = manual_passive_cost
                apply_manual_passive_to_row(row)

            conn.execute(
                """
                INSERT INTO shipments (
                    shipment, orders_text, service_type, carrier_originale, carrier_scelto, manual_carrier,
                    manual_service_level, manual_freight_code, service_level, freight_code, booking_status, province, customer, active_cost, passive_cost,
                    manual_passive_cost, manual_pallets, margin,
                    status, planned_at, unload_date, unload_time, unload_booking_ref, required_delivery_date, delivered_at, source_file, imported_at, payload_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(shipment) DO UPDATE SET
                    orders_text = excluded.orders_text,
                    service_type = excluded.service_type,
                    carrier_originale = excluded.carrier_originale,
                    carrier_scelto = excluded.carrier_scelto,
                    manual_carrier = excluded.manual_carrier,
                    manual_service_level = excluded.manual_service_level,
                    manual_freight_code = excluded.manual_freight_code,
                    service_level = excluded.service_level,
                    freight_code = excluded.freight_code,
                    booking_status = excluded.booking_status,
                    province = excluded.province,
                    customer = excluded.customer,
                    active_cost = excluded.active_cost,
                    passive_cost = excluded.passive_cost,
                    manual_passive_cost = excluded.manual_passive_cost,
                    manual_pallets = excluded.manual_pallets,
                    margin = excluded.margin,
                    status = excluded.status,
                    planned_at = excluded.planned_at,
                    unload_date = excluded.unload_date,
                    unload_time = excluded.unload_time,
                    unload_booking_ref = excluded.unload_booking_ref,
                    required_delivery_date = excluded.required_delivery_date,
                    delivered_at = excluded.delivered_at,
                    source_file = excluded.source_file,
                    imported_at = excluded.imported_at,
                    payload_json = excluded.payload_json
                """,
                (
                    shipment,
                    clean_text(row.get("Orders")),
                    clean_text(row.get("Tipo Servizio")),
                    clean_text(row.get("Carrier Originale")),
                    clean_text(row.get("Carrier Scelto")),
                    manual_carrier,
                    manual_service_level,
                    manual_freight_code,
                    clean_text(row.get("Service Level")),
                    clean_text(row.get("Freight Code")),
                    booking_status,
                    clean_text(row.get("Provincia")),
                    clean_text(row.get("Route to Customer")),
                    to_float(row.get("Costo Attivo")),
                    to_float(row.get("Costo Passivo")),
                    manual_passive_cost,
                    manual_pallets,
                    to_float(row.get("Margine")),
                    status,
                    planned_at or None,
                    unload_date or None,
                    unload_time or None,
                    unload_booking_ref or None,
                    required_delivery_date or None,
                    delivered_at,
                    str(source_file),
                    imported_at,
                    serialize_shipment_payload(row),
                ),
            )

            current_row = (
                shipment,
                clean_text(row.get("Orders")),
                status,
                delivered_at,
                manual_carrier,
                manual_passive_cost,
                manual_service_level,
                manual_freight_code,
                required_delivery_date,
                manual_pallets,
                unload_date,
                unload_time,
                unload_booking_ref,
                planned_at,
            )
            existing_by_shipment[shipment] = current_row
            if order_key:
                existing_by_order[order_key] = current_row
    return counters


def remove_brt_extra_flags(shipments: list[str], flags_path: Path = BRT_EXTRA_FLAGS_PATH) -> None:
    shipment_set = {clean_text(shipment) for shipment in shipments if clean_text(shipment)}
    if not shipment_set or not flags_path.exists():
        return

    with flags_path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        fieldnames = reader.fieldnames or BRT_EXTRA_FLAG_COLUMNS
        rows = [
            row
            for row in reader
            if clean_text(row.get("Shipment")) not in shipment_set
        ]

    with flags_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def delete_shipments_permanently(
    shipments: list[str],
    db_path: Path = DB_PATH,
    flags_path: Path | None = BRT_EXTRA_FLAGS_PATH,
) -> int:
    shipment_set = sorted({clean_text(shipment) for shipment in shipments if clean_text(shipment)})
    if not shipment_set:
        return 0

    init_db(db_path)
    deleted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    deleted = 0
    with sqlite3.connect(db_path) as conn:
        for shipment in shipment_set:
            row = conn.execute(
                "SELECT payload_json FROM shipments WHERE shipment = ?",
                (shipment,),
            ).fetchone()
            payload_json = row[0] if row else ""
            conn.execute(
                """
                INSERT INTO deleted_shipments (shipment, deleted_at, payload_json)
                VALUES (?, ?, ?)
                ON CONFLICT(shipment) DO UPDATE SET
                    deleted_at = excluded.deleted_at,
                    payload_json = excluded.payload_json
                """,
                (shipment, deleted_at, payload_json),
            )
            cursor = conn.execute("DELETE FROM shipments WHERE shipment = ?", (shipment,))
            deleted += cursor.rowcount

    if flags_path is not None:
        remove_brt_extra_flags(shipment_set, flags_path)
    return deleted


def count_deleted_shipments(db_path: Path = DB_PATH) -> int:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        row = conn.execute("SELECT COUNT(*) FROM deleted_shipments").fetchone()
    return int(row[0] or 0) if row else 0


def load_deleted_shipments(db_path: Path = DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT shipment, deleted_at, payload_json
            FROM deleted_shipments
            ORDER BY deleted_at DESC, shipment ASC
            """
        ).fetchall()

    deleted_rows: list[dict[str, Any]] = []
    for shipment, deleted_at, payload_json in rows:
        try:
            payload = json.loads(payload_json or "{}")
        except json.JSONDecodeError:
            payload = {}
        payload["Shipment"] = clean_text(payload.get("Shipment")) or clean_text(shipment)
        payload["Stato"] = "Eliminata"
        payload["Data Eliminazione"] = clean_text(deleted_at)
        deleted_rows.append(payload)
    return deleted_rows


def restore_deleted_shipments(shipments: list[str], db_path: Path = DB_PATH) -> int:
    shipment_set = sorted({clean_text(shipment) for shipment in shipments if clean_text(shipment)})
    if not shipment_set:
        return 0

    init_db(db_path)
    imported_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    restored = 0
    with sqlite3.connect(db_path) as conn:
        for shipment in shipment_set:
            row = conn.execute(
                "SELECT payload_json FROM deleted_shipments WHERE shipment = ?",
                (shipment,),
            ).fetchone()
            if not row:
                continue
            try:
                payload = json.loads(row[0] or "{}")
            except json.JSONDecodeError:
                payload = {}
            payload["Shipment"] = clean_text(payload.get("Shipment")) or shipment
            payload["Stato"] = STATUS_IMPORTED
            payload["Data Partenza"] = ""
            unload_date = clean_text(payload.get("Data Scarico Prenotato"))
            unload_time = clean_text(payload.get("Ora Scarico Prenotato"))
            unload_booking_ref = clean_text(payload.get("Riferimento Booking Scarico"))
            payload["Ora Scarico Prenotato"] = unload_time
            payload["Riferimento Booking Scarico"] = unload_booking_ref
            sync_unload_booking_fields(payload)
            payload["Data Consegna"] = ""
            payload["XML Consegna"] = ""
            carrier_scelto = clean_text(payload.get("Carrier Scelto"))
            manual_passive_cost = to_float(payload.get("Costo Passivo Manuale"))
            manual_service_level = clean_text(payload.get("Service Level Manuale")).upper()
            manual_freight_code = clean_text(payload.get("Freight Code Manuale")).upper()
            required_delivery_date = clean_text(payload.get("Data Consegna Tassativa"))
            if manual_freight_code:
                manual_freight_code, required_delivery_date = apply_manual_freight_code_to_row(
                    payload,
                    manual_freight_code,
                    required_delivery_date,
                    clear_required_for_dkl=not bool(required_delivery_date),
                )
            elif required_delivery_date:
                required_delivery_date = apply_required_delivery_date_to_row(payload, required_delivery_date)
            manual_pallets = to_float(payload.get("Pallet Manuali"))
            booking_status = apply_unload_booking_status(payload)

            conn.execute(
                """
                INSERT INTO shipments (
                    shipment, orders_text, service_type, carrier_originale, carrier_scelto, manual_carrier,
                    manual_service_level, manual_freight_code, service_level, freight_code, booking_status, province, customer, active_cost, passive_cost,
                    manual_passive_cost, manual_pallets, margin,
                    status, departed_at, unload_date, unload_time, unload_booking_ref,
                    required_delivery_date, delivered_at, delivery_xml_path, source_file, imported_at, payload_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?, ?, ?, NULL, NULL, ?, ?, ?)
                ON CONFLICT(shipment) DO UPDATE SET
                    orders_text = excluded.orders_text,
                    service_type = excluded.service_type,
                    carrier_originale = excluded.carrier_originale,
                    carrier_scelto = excluded.carrier_scelto,
                    manual_carrier = excluded.manual_carrier,
                    manual_service_level = excluded.manual_service_level,
                    manual_freight_code = excluded.manual_freight_code,
                    service_level = excluded.service_level,
                    freight_code = excluded.freight_code,
                    booking_status = excluded.booking_status,
                    province = excluded.province,
                    customer = excluded.customer,
                    active_cost = excluded.active_cost,
                    passive_cost = excluded.passive_cost,
                    manual_passive_cost = excluded.manual_passive_cost,
                    manual_pallets = excluded.manual_pallets,
                    margin = excluded.margin,
                    status = excluded.status,
                    departed_at = NULL,
                    unload_date = excluded.unload_date,
                    unload_time = excluded.unload_time,
                    unload_booking_ref = excluded.unload_booking_ref,
                    required_delivery_date = excluded.required_delivery_date,
                    delivered_at = NULL,
                    delivery_xml_path = NULL,
                    source_file = excluded.source_file,
                    imported_at = excluded.imported_at,
                    payload_json = excluded.payload_json
                """,
                (
                    payload["Shipment"],
                    clean_text(payload.get("Orders")),
                    clean_text(payload.get("Tipo Servizio")),
                    clean_text(payload.get("Carrier Originale")),
                    carrier_scelto,
                    carrier_scelto,
                    manual_service_level,
                    manual_freight_code,
                    clean_text(payload.get("Service Level")),
                    clean_text(payload.get("Freight Code")),
                    booking_status,
                    clean_text(payload.get("Provincia")),
                    clean_text(payload.get("Route to Customer")),
                    to_float(payload.get("Costo Attivo")),
                    to_float(payload.get("Costo Passivo")),
                    manual_passive_cost,
                    manual_pallets,
                    to_float(payload.get("Margine")),
                    STATUS_IMPORTED,
                    unload_date or None,
                    unload_time or None,
                    unload_booking_ref or None,
                    required_delivery_date or None,
                    "ripristinata_da_eliminate",
                    imported_at,
                    serialize_shipment_payload(payload),
                ),
            )
            conn.execute("DELETE FROM deleted_shipments WHERE shipment = ?", (shipment,))
            restored += 1
    return restored


def purge_deleted_shipments(shipments: list[str], db_path: Path = DB_PATH) -> int:
    shipment_set = sorted({clean_text(shipment) for shipment in shipments if clean_text(shipment)})
    if not shipment_set:
        return 0

    init_db(db_path)
    purged = 0
    with sqlite3.connect(db_path) as conn:
        for shipment in shipment_set:
            cursor = conn.execute("DELETE FROM deleted_shipments WHERE shipment = ?", (shipment,))
            purged += cursor.rowcount
    return purged


def mark_planned(shipment: str, planned_at: str | None = None, db_path: Path = DB_PATH) -> None:
    row = get_shipment_from_db(shipment, db_path)
    planned_value = (
        iso_date_text(planned_at)
        or (wave_departure_iso(row) if row else "")
        or datetime.now().strftime("%Y-%m-%d")
    )
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "UPDATE shipments SET status = ?, planned_at = ?, departed_at = NULL, delivered_at = NULL, delivery_xml_path = NULL WHERE shipment = ?",
            (STATUS_PLANNED, planned_value, shipment),
        )


def mark_unplanned(shipment: str, db_path: Path = DB_PATH) -> None:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "UPDATE shipments SET status = ?, planned_at = NULL, departed_at = NULL, delivered_at = NULL, delivery_xml_path = NULL WHERE shipment = ?",
            (STATUS_IMPORTED, shipment),
        )


def mark_departed(
    shipment: str,
    departed_at: str | None = None,
    planned_at: str | None = None,
    db_path: Path = DB_PATH,
) -> None:
    departed_value = iso_date_text(departed_at) or datetime.now().strftime("%Y-%m-%d")
    planned_value = iso_date_text(planned_at) or day_after_iso(departed_value)
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "UPDATE shipments SET status = ?, planned_at = ?, departed_at = ?, delivered_at = NULL, delivery_xml_path = NULL WHERE shipment = ?",
            (STATUS_DEPARTED, planned_value, departed_value, shipment),
        )


def mark_confirmed(shipment: str, db_path: Path = DB_PATH) -> None:
    row = get_shipment_from_db(shipment, db_path)
    if not row:
        raise ValueError(f"Spedizione non trovata: {shipment}")
    if clean_text(row.get("Tipo Servizio")) == "Groupage - BRT LTL":
        raise ValueError("FTL confermato si puo usare solo su spedizioni dirette LTL/FTL, non sul groupage BRT.")
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "UPDATE shipments SET status = ?, departed_at = NULL, delivered_at = NULL, delivery_xml_path = NULL WHERE shipment = ?",
            (STATUS_CONFIRMED, shipment),
        )


def get_shipment_from_db(shipment: str, db_path: Path = DB_PATH) -> dict[str, Any] | None:
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            """
            SELECT payload_json, status, planned_at, departed_at, unload_date, unload_time, unload_booking_ref,
                   delivered_at, delivery_xml_path,
                   manual_carrier, manual_passive_cost, manual_service_level, manual_freight_code,
                   required_delivery_date, manual_pallets, booking_status
            FROM shipments
            WHERE shipment = ?
            """,
            (shipment,),
        ).fetchone()
    if not row:
        return None

    (
        payload_json,
        status,
        planned_at,
        departed_at,
        unload_date,
        unload_time,
        unload_booking_ref,
        delivered_at,
        delivery_xml_path,
        manual_carrier,
        manual_passive_cost,
        manual_service_level,
        manual_freight_code,
        required_delivery_date,
        manual_pallets,
        booking_status,
    ) = row
    try:
        payload = json.loads(payload_json)
    except json.JSONDecodeError:
        return None
    payload["Stato"] = status
    payload["Data Pianifica"] = planned_at or ""
    payload["Data Partenza"] = departed_at or ""
    payload["Data Scarico Prenotato"] = unload_date or ""
    payload["Ora Scarico Prenotato"] = unload_time or ""
    payload["Riferimento Booking Scarico"] = unload_booking_ref or ""
    sync_unload_booking_fields(payload)
    payload["Data Consegna"] = delivered_at or ""
    payload["XML Consegna"] = delivery_xml_path or ""
    if manual_carrier:
        payload["Carrier Scelto"] = manual_carrier
    if manual_service_level:
        apply_manual_service_level_to_row(payload, manual_service_level)
    if manual_freight_code:
        apply_manual_freight_code_to_row(
            payload,
            manual_freight_code,
            required_delivery_date,
            clear_required_for_dkl=not bool(required_delivery_date),
        )
    else:
        payload["Data Consegna Tassativa"] = required_delivery_date or clean_text(payload.get("Data Consegna Tassativa"))
        if is_dkv_order(payload) and clean_text(payload.get("Data Consegna Tassativa")):
            hide_early_delivery_for_dkv(payload)
    if not is_dkv_order(payload) and not clean_text(payload.get("Early Delivery Date")):
        payload["Early Delivery Date"] = clean_text(payload.get("Late Delivery Date"))
    if manual_pallets is not None:
        apply_manual_pallets_to_row(payload, manual_pallets)
    if manual_passive_cost is not None:
        payload["Costo Passivo Manuale"] = manual_passive_cost
    apply_unload_booking_status(payload, booking_status)
    apply_warehouse_contacts_to_shipments([payload])
    return payload


def refresh_early_delivery_dates_from_source(vtech_path: Path | None, db_path: Path = DB_PATH) -> int:
    if not vtech_path or not vtech_path.exists():
        return 0

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        db_rows = conn.execute("SELECT shipment, payload_json FROM shipments").fetchall()

    source_fields = ["Early Delivery Date", "Grand Total Shipment Ftp Vol m3"]
    missing_shipments: set[str] = set()
    payloads: dict[str, dict[str, Any]] = {}
    for shipment, payload_json in db_rows:
        try:
            payload = json.loads(payload_json)
        except json.JSONDecodeError:
            continue
        payloads[shipment] = payload
        if any(
            not clean_text(payload.get(field))
            for field in source_fields
            if not (field == "Early Delivery Date" and is_dkv_order(payload))
        ):
            missing_shipments.add(shipment)

    if not missing_shipments:
        return 0

    source_rows = build_shipment_rows(extract_vtech_rows(vtech_path))
    source_by_shipment = {
        clean_text(row.get("Shipment")): row
        for row in source_rows
        if clean_text(row.get("Shipment"))
    }

    updates: list[tuple[str, str]] = []
    for shipment in missing_shipments:
        source_row = source_by_shipment.get(shipment)
        if not source_row:
            continue
        payload = payloads.get(shipment, {})
        changed = False
        for field in source_fields:
            if field == "Early Delivery Date" and is_dkv_order(payload):
                continue
            source_value = source_row.get(field)
            if clean_text(source_value) and not clean_text(payload.get(field)):
                payload[field] = source_value
                changed = True
        if not changed:
            continue
        updates.append(
            (
                serialize_shipment_payload(payload),
                shipment,
            )
        )

    if not updates:
        return 0

    with sqlite3.connect(db_path) as conn:
        conn.executemany("UPDATE shipments SET payload_json = ? WHERE shipment = ?", updates)
    return len(updates)


def normalize_saved_date_columns(db_path: Path = DB_PATH) -> int:
    init_db(db_path)
    updates: list[tuple[str, str]] = []
    with sqlite3.connect(db_path) as conn:
        db_rows = conn.execute("SELECT shipment, payload_json FROM shipments").fetchall()

    for shipment, payload_json in db_rows:
        try:
            payload = json.loads(payload_json)
        except json.JSONDecodeError:
            continue

        changed = False
        for column in DATE_ONLY_COLUMNS:
            current = clean_text(payload.get(column))
            normalized = format_date_only(current)
            if current and normalized and normalized != current:
                payload[column] = normalized
                changed = True

        if changed:
            updates.append(
                (
                    serialize_shipment_payload(payload),
                    shipment,
                )
            )

    if not updates:
        return 0

    with sqlite3.connect(db_path) as conn:
        conn.executemany("UPDATE shipments SET payload_json = ? WHERE shipment = ?", updates)
    return len(updates)


def set_manual_carrier(
    shipment: str,
    carrier: str,
    active_rates_path: Path | None,
    brt_passive_path: Path | None,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")

    carrier_clean = clean_text(carrier).upper()
    row["Carrier Scelto"] = carrier_clean
    apply_tariffs_to_shipments(
        [row],
        active_rates_path=active_rates_path if active_rates_path and active_rates_path.exists() else None,
        brt_passive_pdf_path=brt_passive_path if brt_passive_path and brt_passive_path.exists() else None,
        brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
        gdo_customers_path=available_gdo_customers_path(),
        fuel_settings_path=available_fuel_settings_path(),
    )
    manual_passive_cost = to_float(row.get("Costo Passivo Manuale"))
    if manual_passive_cost is not None:
        apply_manual_passive_to_row(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET carrier_scelto = ?,
                manual_carrier = ?,
                active_cost = ?,
                passive_cost = ?,
                manual_passive_cost = ?,
                margin = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                clean_text(row.get("Carrier Scelto")),
                carrier_clean,
                to_float(row.get("Costo Attivo")),
                to_float(row.get("Costo Passivo")),
                manual_passive_cost,
                to_float(row.get("Margine")),
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def set_manual_service_level(
    shipment: str,
    service_level: str,
    active_rates_path: Path | None,
    brt_passive_path: Path | None,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")

    manual_service_level = normalize_service_level(service_level)
    apply_manual_service_level_to_row(row, manual_service_level)

    apply_tariffs_to_shipments(
        [row],
        active_rates_path=active_rates_path if active_rates_path and active_rates_path.exists() else None,
        brt_passive_pdf_path=brt_passive_path if brt_passive_path and brt_passive_path.exists() else None,
        brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
        gdo_customers_path=available_gdo_customers_path(),
        fuel_settings_path=available_fuel_settings_path(),
    )
    apply_manual_service_level_to_row(row, manual_service_level)
    manual_passive_cost = to_float(row.get("Costo Passivo Manuale"))
    if manual_passive_cost is not None:
        apply_manual_passive_to_row(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET service_level = ?,
                service_type = ?,
                carrier_scelto = ?,
                manual_carrier = ?,
                manual_service_level = ?,
                active_cost = ?,
                passive_cost = ?,
                manual_passive_cost = ?,
                margin = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                clean_text(row.get("Service Level")),
                clean_text(row.get("Tipo Servizio")),
                clean_text(row.get("Carrier Scelto")),
                clean_text(row.get("Carrier Scelto")),
                manual_service_level,
                to_float(row.get("Costo Attivo")),
                to_float(row.get("Costo Passivo")),
                manual_passive_cost,
                to_float(row.get("Margine")),
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def set_manual_freight_code(
    shipment: str,
    freight_code: str,
    required_delivery_date: Any,
    active_rates_path: Path | None,
    brt_passive_path: Path | None,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")

    manual_freight_code, normalized_required_date = apply_manual_freight_code_to_row(
        row,
        freight_code,
        required_delivery_date,
    )

    apply_tariffs_to_shipments(
        [row],
        active_rates_path=active_rates_path if active_rates_path and active_rates_path.exists() else None,
        brt_passive_pdf_path=brt_passive_path if brt_passive_path and brt_passive_path.exists() else None,
        brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
        gdo_customers_path=available_gdo_customers_path(),
        fuel_settings_path=available_fuel_settings_path(),
    )
    apply_manual_freight_code_to_row(row, manual_freight_code, normalized_required_date)
    booking_status = apply_unload_booking_status(row)
    if active_rates_path and active_rates_path.exists():
        apply_contract_sla_to_shipments([row], active_rates_path=active_rates_path)
    manual_passive_cost = to_float(row.get("Costo Passivo Manuale"))
    if manual_passive_cost is not None:
        apply_manual_passive_to_row(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET freight_code = ?,
                booking_status = ?,
                manual_freight_code = ?,
                required_delivery_date = ?,
                active_cost = ?,
                passive_cost = ?,
                manual_passive_cost = ?,
                margin = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                clean_text(row.get("Freight Code")),
                booking_status,
                manual_freight_code,
                normalized_required_date or None,
                to_float(row.get("Costo Attivo")),
                to_float(row.get("Costo Passivo")),
                manual_passive_cost,
                to_float(row.get("Margine")),
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def set_required_delivery_date(
    shipment: str,
    required_delivery_date: Any,
    active_rates_path: Path | None,
    clear: bool = False,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")

    normalized_required_date = apply_required_delivery_date_to_row(
        row,
        required_delivery_date,
        clear=clear,
    )
    if active_rates_path and active_rates_path.exists():
        apply_contract_sla_to_shipments([row], active_rates_path=active_rates_path)
    booking_status = apply_unload_booking_status(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET required_delivery_date = ?,
                booking_status = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                normalized_required_date or None,
                booking_status,
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def set_manual_passive_cost(
    shipment: str,
    passive_cost: Any,
    active_rates_path: Path | None,
    brt_passive_path: Path | None,
    clear: bool = False,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")

    if clear or clean_text(passive_cost) == "":
        manual_passive_cost = None
        clear_manual_passive_from_row(row)
    else:
        passive_text = clean_text(passive_cost).upper().replace("EUR", "").replace("€", "")
        manual_passive_cost = to_float(passive_text)
        if manual_passive_cost is None:
            raise ValueError("Inserisci un costo passivo manuale valido.")
        if manual_passive_cost < 0:
            raise ValueError("Il costo passivo manuale non puo essere negativo.")
        manual_passive_cost = round(manual_passive_cost, 2)
        row["Costo Passivo Manuale"] = manual_passive_cost

    apply_tariffs_to_shipments(
        [row],
        active_rates_path=active_rates_path if active_rates_path and active_rates_path.exists() else None,
        brt_passive_pdf_path=brt_passive_path if brt_passive_path and brt_passive_path.exists() else None,
        brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
        gdo_customers_path=available_gdo_customers_path(),
        fuel_settings_path=available_fuel_settings_path(),
    )
    if manual_passive_cost is not None:
        row["Costo Passivo Manuale"] = manual_passive_cost
        apply_manual_passive_to_row(row)
    else:
        clear_manual_passive_from_row(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET active_cost = ?,
                passive_cost = ?,
                manual_passive_cost = ?,
                margin = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                to_float(row.get("Costo Attivo")),
                to_float(row.get("Costo Passivo")),
                manual_passive_cost,
                to_float(row.get("Margine")),
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def set_manual_pallets(
    shipment: str,
    pallets: Any,
    active_rates_path: Path | None,
    brt_passive_path: Path | None,
    clear: bool = False,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")

    if clear or clean_text(pallets) == "":
        manual_pallets = None
        clear_manual_pallets_from_row(row)
    else:
        manual_pallets = apply_manual_pallets_to_row(row, pallets)

    apply_tariffs_to_shipments(
        [row],
        active_rates_path=active_rates_path if active_rates_path and active_rates_path.exists() else None,
        brt_passive_pdf_path=brt_passive_path if brt_passive_path and brt_passive_path.exists() else None,
        brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
        gdo_customers_path=available_gdo_customers_path(),
        fuel_settings_path=available_fuel_settings_path(),
    )
    if manual_pallets is not None:
        apply_manual_pallets_to_row(row, manual_pallets)
    else:
        clear_manual_pallets_from_row(row)

    manual_passive_cost = to_float(row.get("Costo Passivo Manuale"))
    if manual_passive_cost is not None:
        apply_manual_passive_to_row(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET active_cost = ?,
                passive_cost = ?,
                manual_passive_cost = ?,
                manual_pallets = ?,
                margin = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                to_float(row.get("Costo Attivo")),
                to_float(row.get("Costo Passivo")),
                manual_passive_cost,
                manual_pallets,
                to_float(row.get("Margine")),
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def set_unload_date(
    shipment: str,
    unload_date: Any,
    clear: bool = False,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    return set_unload_booking(shipment, unload_date, "", "", clear=clear, db_path=db_path)


def set_unload_booking(
    shipment: str,
    unload_date: Any,
    unload_time: Any = "",
    booking_ref: Any = "",
    clear: bool = False,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")
    if is_brt_groupage(row):
        raise ValueError("Il booking scarico si imposta sulle spedizioni FTL/LTL, non sul groupage BRT.")

    if not clear and not clean_text(unload_date):
        raise ValueError("Inserisci la data scarico.")
    if not clear and not clean_text(unload_time):
        raise ValueError("Inserisci l'ora scarico.")
    normalized_date = "" if clear else normalize_unload_date(unload_date)
    normalized_time = "" if clear else normalize_unload_time(unload_time)
    normalized_ref = "" if clear else clean_text(booking_ref)
    row["Data Scarico Prenotato"] = normalized_date
    row["Ora Scarico Prenotato"] = normalized_time
    row["Riferimento Booking Scarico"] = normalized_ref
    sync_unload_booking_fields(row)
    booking_status = apply_unload_booking_status(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET unload_date = ?,
                unload_time = ?,
                unload_booking_ref = ?,
                booking_status = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                normalized_date or None,
                normalized_time or None,
                normalized_ref or None,
                booking_status,
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def set_active_urgent(
    shipment: str,
    active_rates_path: Path | None,
    brt_passive_path: Path | None,
    clear: bool = False,
    db_path: Path = DB_PATH,
) -> dict[str, Any]:
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")

    apply_active_urgent_to_row(row, not clear)
    apply_tariffs_to_shipments(
        [row],
        active_rates_path=active_rates_path if active_rates_path and active_rates_path.exists() else None,
        brt_passive_pdf_path=brt_passive_path if brt_passive_path and brt_passive_path.exists() else None,
        brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
        gdo_customers_path=available_gdo_customers_path(),
        fuel_settings_path=available_fuel_settings_path(),
    )
    manual_passive_cost = to_float(row.get("Costo Passivo Manuale"))
    if manual_passive_cost is not None:
        apply_manual_passive_to_row(row)

    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE shipments
            SET active_cost = ?,
                passive_cost = ?,
                margin = ?,
                payload_json = ?
            WHERE shipment = ?
            """,
            (
                to_float(row.get("Costo Attivo")),
                to_float(row.get("Costo Passivo")),
                to_float(row.get("Margine")),
                serialize_shipment_payload(row),
                shipment,
            ),
        )
    return row


def generate_delivery_xml(row: dict[str, Any], delivered_at: str) -> Path:
    XML_DIR.mkdir(parents=True, exist_ok=True)
    shipment = clean_text(row.get("Shipment"))
    if not shipment:
        raise ValueError("Shipment mancante: impossibile creare XML consegna.")
    if not DELIVERY_XML_TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Modello XML non trovato: {DELIVERY_XML_TEMPLATE_PATH}")

    safe_shipment = "".join(char if char.isalnum() or char in "-_" else "_" for char in shipment)
    output_path = XML_DIR / f"KN_{safe_shipment}.xml"

    tree = ET.parse(DELIVERY_XML_TEMPLATE_PATH)
    root = tree.getroot()
    release_xid = root.find("./TransmissionBody/GLogXMLElement/TransOrderStatus/ReleaseGid/Gid/Xid")
    event_date = root.find("./TransmissionBody/GLogXMLElement/TransOrderStatus/EventDt/GLogDate")
    if release_xid is None or event_date is None:
        raise ValueError("Modello XML consegna non valido: ReleaseGid/Xid o EventDt/GLogDate mancanti.")

    release_xid.text = shipment
    event_date.text = f"{datetime.now().strftime('%Y%m%d')}180000"

    tree.write(output_path, encoding="utf-8", xml_declaration=False)
    if not output_path.exists() or output_path.stat().st_size == 0:
        raise OSError(f"XML non creato correttamente: {output_path}")
    return output_path


def mark_delivered(shipment: str, db_path: Path = DB_PATH) -> tuple[str, Path]:
    delivered_at = datetime.now().strftime("%Y-%m-%d")
    row = get_shipment_from_db(shipment, db_path)
    if row is None:
        raise ValueError(f"Spedizione non trovata: {shipment}")
    if clean_text(row.get("Stato")) != STATUS_CONFIRMED:
        raise ValueError("Puoi segnare come consegnate solo le spedizioni in FTL CONFERMATI.")
    xml_path = generate_delivery_xml(row, delivered_at)
    init_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            "UPDATE shipments SET status = ?, delivered_at = ?, delivery_xml_path = ? WHERE shipment = ?",
            (STATUS_DELIVERED, delivered_at, str(xml_path), shipment),
        )
    return delivered_at, xml_path


def _safe_sheet_title(name: str, used_titles: set[str]) -> str:
    invalid_chars = set('[]:*?/\\')
    cleaned = "".join(char if char not in invalid_chars else " " for char in clean_text(name))
    cleaned = " ".join(cleaned.split()) or "Senza vettore"
    base = cleaned[:31]
    title = base
    index = 2
    while title in used_titles:
        suffix = f" {index}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(title)
    return title


def _safe_table_name(prefix: str, name: str, used_names: set[str]) -> str:
    cleaned = "".join(char if char.isalnum() else "_" for char in f"{prefix}_{name}")
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    cleaned = cleaned[:240]
    table_name = cleaned
    index = 2
    while table_name in used_names:
        table_name = f"{cleaned[:230]}_{index}"
        index += 1
    used_names.add(table_name)
    return table_name


def _style_passive_sheet(sheet, money_columns: list[int], numeric_columns: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="0F2742")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D8E3EE")
    border = Border(bottom=thin)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_index in money_columns:
        for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for item in cell:
                item.number_format = '#,##0.00 [$€-it-IT]'
    for column_index in numeric_columns:
        for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for item in cell:
                item.number_format = '#,##0.00'
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 11), 44)


EXTRA_COLUMN_ORDER = [
    "ETS",
    "Fuel",
    "Traghetti",
    "Consegna GDO",
    "Sponda",
    "Appuntamento",
    "Preavviso",
    "Amazon booking",
    "BKV time slot",
    "Urgent delivery",
    "Priority",
    "Servizio 10:30",
    "ZTL",
    "Contrassegno",
    "Fuori misura",
    "POD",
    "Ricerca documenti",
    "ORM",
    "Bancali a rendere",
    "Giacenza",
    "Riconsegna giacenza",
    "Dirottamento",
    "Localita disagiata",
    "Consegna disagiata",
    "Isole minori",
    "Zona franca",
]


def billing_extra_label(raw_label: Any) -> str:
    label = clean_text(raw_label)
    normalized = label.upper()
    if "ETS" in normalized:
        return "ETS"
    if "FUEL" in normalized:
        return "Fuel"
    if "TRAGHETT" in normalized:
        return "Traghetti"
    if "GDO" in normalized or "SUPERMERCAT" in normalized:
        return "Consegna GDO"
    if "SPONDA" in normalized:
        return "Sponda"
    if "PHONE PREADVISE" in normalized or "PREADVISE" in normalized:
        return "Preavviso"
    if "AMAZON" in normalized and ("BKL" in normalized or "TIME SLOT" in normalized or "FIX" in normalized):
        return "Amazon booking"
    if "BKV" in normalized or "FIXED TIME SLOT" in normalized:
        return "BKV time slot"
    if "APPUNTAMENTO" in normalized or "TIME SLOT" in normalized:
        return "Appuntamento"
    if "URGENT" in normalized:
        return "Urgent delivery"
    if "PRIORITY" in normalized:
        return "Priority"
    if "10:30" in normalized:
        return "Servizio 10:30"
    if "CONTRASSEGNO" in normalized:
        return "Contrassegno"
    if "ZTL" in normalized:
        return "ZTL"
    if "FUORI MISURA" in normalized:
        return "Fuori misura"
    if "P.O.D" in normalized or "POD" in normalized:
        return "POD"
    if "RICERCA" in normalized or "DOCUMENT" in normalized:
        return "Ricerca documenti"
    if "O.R.M" in normalized or "ORM" in normalized:
        return "ORM"
    if "BANCALI" in normalized:
        return "Bancali a rendere"
    if "RICONSEGNA" in normalized:
        return "Riconsegna giacenza"
    if "GIACENZA" in normalized:
        return "Giacenza"
    if "DIROTTAMENTO" in normalized:
        return "Dirottamento"
    if "LOCALITA" in normalized or "LOCALITÀ" in normalized:
        return "Localita disagiata"
    if "CONSEGNA DISAGIATA" in normalized:
        return "Consegna disagiata"
    if "ISOLE" in normalized:
        return "Isole minori"
    if "LIVIGNO" in normalized or "CAMPIONE" in normalized or "ZONA FRANCA" in normalized:
        return "Zona franca"
    cleaned = re.sub(r"\+\s*\d+(?:[,.]\d+)?\s*%", "", label)
    cleaned = re.sub(r"\b\d+(?:[,.]\d+)?\s*%\b", "", cleaned)
    return " ".join(cleaned.split()) or "Altro extra"


def parse_billing_extra_amounts(value: Any) -> dict[str, float]:
    text = clean_text(value)
    if not text:
        return {}
    amounts: dict[str, float] = {}
    for part in text.split("|"):
        item = clean_text(part)
        if not item:
            continue
        match = re.match(r"^(.*?):\s*(?:EUR|€)?\s*([-+]?\d[\d.,]*)", item, flags=re.IGNORECASE)
        if not match:
            continue
        label = billing_extra_label(match.group(1))
        amount = to_float(match.group(2))
        if amount is None:
            continue
        amounts[label] = round(amounts.get(label, 0) + amount, 2)
    return amounts


def billing_extra_columns(rows: list[tuple[date, dict[str, Any]]], field: str) -> list[str]:
    labels: set[str] = set()
    for _ref_date, row in rows:
        labels.update(parse_billing_extra_amounts(row.get(field)).keys())
    priority = {label: index for index, label in enumerate(EXTRA_COLUMN_ORDER)}
    return sorted(labels, key=lambda label: (priority.get(label, len(priority)), label.lower()))


def sum_formula(column_index: int, total_row: int) -> str:
    letter = get_column_letter(column_index)
    return f"=SUM({letter}2:{letter}{total_row - 1})"


def export_passive_billing_by_carrier(
    month_key: str | None = None,
    db_path: Path = DB_PATH,
    downloads_dir: Path = DOWNLOADS_DIR,
) -> Path:
    rows = load_shipments_from_db(db_path)
    selected: list[tuple[date, dict[str, Any]]] = []
    normalized_month = clean_text(month_key)
    for row in rows:
        ref_date = billing_reference_date(row)
        if ref_date is None:
            continue
        if normalized_month and f"{ref_date.year:04d}-{ref_date.month:02d}" != normalized_month:
            continue
        passive = to_float(row.get("Costo Passivo"))
        if passive is None or passive <= 0:
            continue
        selected.append((ref_date, row))

    if not selected:
        label = normalized_month or "tutti i mesi"
        raise ValueError(f"Nessuna spedizione con passivo disponibile per {label}.")

    selected.sort(key=lambda item: (clean_text(item[1].get("Carrier Scelto")), item[0], clean_text(item[1].get("Shipment"))))
    by_carrier: dict[str, list[tuple[date, dict[str, Any]]]] = {}
    for ref_date, row in selected:
        carrier = clean_text(row.get("Carrier Scelto")) or clean_text(row.get("Carrier Originale")) or "SENZA VETTORE"
        by_carrier.setdefault(carrier, []).append((ref_date, row))

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Riepilogo"
    workbook.properties.title = "V-Tech passivo vettori"
    workbook.properties.subject = "Dettaglio passivo per vettore"

    summary_headers = ["Vettore", "Spedizioni", "Trasporto", "Extra", "Passivo totale", "Peso kg", "Pallet fatt."]
    summary.append(summary_headers)
    for carrier, carrier_rows in sorted(by_carrier.items()):
        transport = sum(to_float(row.get("Costo Passivo Base BRT")) or 0 for _date, row in carrier_rows)
        extras = sum(to_float(row.get("Extra BRT Totale")) or 0 for _date, row in carrier_rows)
        passive = sum(to_float(row.get("Costo Passivo")) or 0 for _date, row in carrier_rows)
        weight = sum(to_float(row.get("Grand Total Shipment Ftp Wgt Kg")) or 0 for _date, row in carrier_rows)
        pallets = sum(to_float(row.get("Pallet Fatturati")) or 0 for _date, row in carrier_rows)
        summary.append([carrier, len(carrier_rows), transport, extras, passive, weight, pallets])

    total_row = summary.max_row + 1
    summary.append([
        "Totale",
        f"=SUM(B2:B{total_row - 1})",
        f"=SUM(C2:C{total_row - 1})",
        f"=SUM(D2:D{total_row - 1})",
        f"=SUM(E2:E{total_row - 1})",
        f"=SUM(F2:F{total_row - 1})",
        f"=SUM(G2:G{total_row - 1})",
    ])
    for cell in summary[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E0F2FE")
    _style_passive_sheet(summary, money_columns=[3, 4, 5], numeric_columns=[2, 6, 7])
    summary_table = Table(displayName="RiepilogoPassivo", ref=f"A1:G{summary.max_row}")
    summary_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    summary.add_table(summary_table)

    passive_extra_columns = billing_extra_columns(selected, "Extra BRT Applicati")
    passive_base_headers = [
        "Data rif.",
        "Shipment",
        "Ordine",
        "Cliente",
        "Provincia",
        "Indirizzo consegna",
        "Servizio",
        "Vettore",
        "Freight",
        "Pallet fatt.",
        "Peso kg",
        "Volume m3",
    ]
    passive_price_headers = [
        "Trasporto base",
        *passive_extra_columns,
        "Passivo totale",
    ]
    detail_headers = [
        *passive_base_headers,
        *passive_price_headers,
        "Tariffa passiva",
        "Extra applicati",
    ]
    used_titles = {"Riepilogo"}
    used_tables = {"RiepilogoPassivo"}
    for carrier, carrier_rows in sorted(by_carrier.items()):
        sheet = workbook.create_sheet(_safe_sheet_title(carrier, used_titles))
        sheet.append(detail_headers)
        for ref_date, row in carrier_rows:
            extra_amounts = parse_billing_extra_amounts(row.get("Extra BRT Applicati"))
            sheet.append([
                ref_date,
                clean_text(row.get("Shipment")),
                clean_text(row.get("Orders")),
                clean_text(row.get("Route to Customer")),
                clean_text(row.get("Provincia")),
                clean_text(row.get("Route To Address")),
                clean_text(row.get("Service Level")),
                carrier,
                clean_text(row.get("Freight Code")),
                to_float(row.get("Pallet Fatturati")) or 0,
                to_float(row.get("Grand Total Shipment Ftp Wgt Kg")) or 0,
                to_float(row.get("Grand Total Shipment Ftp Vol m3")) or 0,
                to_float(row.get("Costo Passivo Base BRT")) or 0,
                *[extra_amounts.get(label, "") for label in passive_extra_columns],
                to_float(row.get("Costo Passivo")) or 0,
                clean_text(row.get("Tariffa Passiva Applicata")),
                clean_text(row.get("Extra BRT Applicati")),
            ])
        total = sheet.max_row + 1
        total_values: list[Any] = [""] * len(detail_headers)
        total_values[0] = "Totale"
        total_values[7] = carrier
        for column_index in [10, 11, 12]:
            total_values[column_index - 1] = sum_formula(column_index, total)
        price_start = len(passive_base_headers) + 1
        price_end = price_start + len(passive_price_headers) - 1
        for column_index in range(price_start, price_end + 1):
            total_values[column_index - 1] = sum_formula(column_index, total)
        sheet.append(total_values)
        for cell in sheet[total]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="ECFDF5")
        for cell in sheet["A"][1:]:
            if isinstance(cell.value, date):
                cell.number_format = "dd/mm/yyyy"
        _style_passive_sheet(
            sheet,
            money_columns=list(range(price_start, price_end + 1)),
            numeric_columns=[10, 11, 12],
        )
        table_ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        table = Table(displayName=_safe_table_name("Passivo", carrier, used_tables), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)

    downloads_dir.mkdir(parents=True, exist_ok=True)
    suffix = normalized_month if normalized_month else "tutti_i_mesi"
    output_path = downloads_dir / f"VTech_passivo_vettori_{suffix}.xlsx"
    workbook.save(output_path)
    if not output_path.exists() or output_path.stat().st_size == 0:
        raise OSError(f"Excel passivo non creato correttamente: {output_path}")
    return output_path


def export_active_billing_report(
    month_key: str | None = None,
    db_path: Path = DB_PATH,
    downloads_dir: Path = DOWNLOADS_DIR,
) -> Path:
    rows = load_shipments_from_db(db_path)
    selected: list[tuple[date, dict[str, Any]]] = []
    normalized_month = clean_text(month_key)
    for row in rows:
        ref_date = billing_reference_date(row)
        if ref_date is None:
            continue
        if normalized_month and f"{ref_date.year:04d}-{ref_date.month:02d}" != normalized_month:
            continue
        active = to_float(row.get("Costo Attivo"))
        if active is None or active <= 0:
            continue
        selected.append((ref_date, row))

    if not selected:
        label = normalized_month or "tutti i mesi"
        raise ValueError(f"Nessuna spedizione con attivo disponibile per {label}.")

    selected.sort(key=lambda item: (clean_text(item[1].get("Route to Customer")), item[0], clean_text(item[1].get("Shipment"))))
    by_customer: dict[str, list[tuple[date, dict[str, Any]]]] = {}
    for ref_date, row in selected:
        customer = clean_text(row.get("Route to Customer")) or "CLIENTE NON INDICATO"
        by_customer.setdefault(customer, []).append((ref_date, row))

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Riepilogo"
    workbook.properties.title = "V-Tech fatturazione attiva"
    workbook.properties.subject = "Dettaglio fatturazione attiva cliente"

    summary_headers = ["Cliente", "Spedizioni", "Base attiva", "Extra attivi", "Attivo totale", "Pallet fatt.", "Peso kg", "Volume m3"]
    summary.append(summary_headers)
    for customer, customer_rows in sorted(by_customer.items()):
        active = sum(to_float(row.get("Costo Attivo")) or 0 for _date, row in customer_rows)
        extras = sum(to_float(row.get("Extra Attivi Totale")) or 0 for _date, row in customer_rows)
        base = max(0, active - extras)
        pallets = sum(to_float(row.get("Pallet Fatturati")) or 0 for _date, row in customer_rows)
        weight = sum(to_float(row.get("Grand Total Shipment Ftp Wgt Kg")) or 0 for _date, row in customer_rows)
        volume = sum(to_float(row.get("Grand Total Shipment Ftp Vol m3")) or 0 for _date, row in customer_rows)
        summary.append([customer, len(customer_rows), base, extras, active, pallets, weight, volume])

    total_row = summary.max_row + 1
    summary.append([
        "Totale",
        f"=SUM(B2:B{total_row - 1})",
        f"=SUM(C2:C{total_row - 1})",
        f"=SUM(D2:D{total_row - 1})",
        f"=SUM(E2:E{total_row - 1})",
        f"=SUM(F2:F{total_row - 1})",
        f"=SUM(G2:G{total_row - 1})",
        f"=SUM(H2:H{total_row - 1})",
    ])
    for cell in summary[total_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
    _style_passive_sheet(summary, money_columns=[3, 4, 5], numeric_columns=[2, 6, 7, 8])
    summary_table = Table(displayName="RiepilogoAttivo", ref=f"A1:H{summary.max_row}")
    summary_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    summary.add_table(summary_table)

    detail = workbook.create_sheet("Dettaglio attivo")
    active_extra_columns = billing_extra_columns(selected, "Extra Attivi Applicati")
    active_base_headers = [
        "Data rif.",
        "Shipment",
        "Ordine",
        "Cliente",
        "GDO",
        "Provincia",
        "Indirizzo consegna",
        "Servizio",
        "Vettore",
        "Freight",
        "Prenotazione",
        "Pallet fatt.",
        "Peso kg",
        "Volume m3",
    ]
    active_price_headers = [
        "Base attiva",
        *active_extra_columns,
        "Attivo totale",
    ]
    detail_headers = [
        *active_base_headers,
        *active_price_headers,
        "Extra attivi applicati",
        "Tariffa attiva",
        "Stato",
    ]
    detail.append(detail_headers)
    for ref_date, row in selected:
        active = to_float(row.get("Costo Attivo")) or 0
        extras = to_float(row.get("Extra Attivi Totale")) or 0
        base = max(0, active - extras)
        extra_amounts = parse_billing_extra_amounts(row.get("Extra Attivi Applicati"))
        detail.append([
            ref_date,
            clean_text(row.get("Shipment")),
            clean_text(row.get("Orders")),
            clean_text(row.get("Route to Customer")),
            clean_text(row.get("Cliente GDO")),
            clean_text(row.get("Provincia")),
            clean_text(row.get("Route To Address")),
            clean_text(row.get("Service Level")),
            clean_text(row.get("Carrier Scelto")),
            clean_text(row.get("Freight Code")),
            clean_text(row.get("Prenotazione Scarico")),
            to_float(row.get("Pallet Fatturati")) or 0,
            to_float(row.get("Grand Total Shipment Ftp Wgt Kg")) or 0,
            to_float(row.get("Grand Total Shipment Ftp Vol m3")) or 0,
            base,
            *[extra_amounts.get(label, "") for label in active_extra_columns],
            active,
            clean_text(row.get("Extra Attivi Applicati")),
            clean_text(row.get("Tariffa Attiva Applicata")),
            clean_text(row.get("Stato")),
        ])

    total = detail.max_row + 1
    total_values = [""] * len(detail_headers)
    total_values[0] = "Totale"
    for column_index in [12, 13, 14]:
        total_values[column_index - 1] = sum_formula(column_index, total)
    price_start = len(active_base_headers) + 1
    price_end = price_start + len(active_price_headers) - 1
    for column_index in range(price_start, price_end + 1):
        total_values[column_index - 1] = sum_formula(column_index, total)
    detail.append(total_values)
    for cell in detail[total]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="ECFDF5")
    for cell in detail["A"][1:]:
        if isinstance(cell.value, date):
            cell.number_format = "dd/mm/yyyy"
    _style_passive_sheet(
        detail,
        money_columns=list(range(price_start, price_end + 1)),
        numeric_columns=[12, 13, 14],
    )
    detail_ref = f"A1:{get_column_letter(detail.max_column)}{detail.max_row}"
    detail_table = Table(displayName="DettaglioAttivo", ref=detail_ref)
    detail_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    detail.add_table(detail_table)

    downloads_dir.mkdir(parents=True, exist_ok=True)
    suffix = normalized_month if normalized_month else "tutti_i_mesi"
    output_path = downloads_dir / f"VTech_fatturazione_attiva_{suffix}.xlsx"
    workbook.save(output_path)
    if not output_path.exists() or output_path.stat().st_size == 0:
        raise OSError(f"Excel fatturazione attiva non creato correttamente: {output_path}")
    return output_path


def load_shipments_from_db(db_path: Path = DB_PATH) -> list[dict[str, Any]]:
    init_db(db_path)
    cache_key = shipments_cache_key(db_path)
    with SHIPMENTS_CACHE_LOCK:
        if SHIPMENTS_CACHE["key"] == cache_key and SHIPMENTS_CACHE["rows"] is not None:
            return copy.deepcopy(SHIPMENTS_CACHE["rows"])

    rows: list[dict[str, Any]] = []
    with sqlite3.connect(db_path) as conn:
        db_rows = conn.execute(
            """
            SELECT payload_json, status, planned_at, departed_at, unload_date, unload_time, unload_booking_ref,
                   delivered_at, freight_code, booking_status, delivery_xml_path,
                   manual_carrier, manual_passive_cost, manual_service_level, manual_freight_code,
                   required_delivery_date, manual_pallets
            FROM shipments
            ORDER BY imported_at DESC, shipment ASC
            """
        ).fetchall()

    for (
        payload_json,
        status,
        planned_at,
        departed_at,
        unload_date,
        unload_time,
        unload_booking_ref,
        delivered_at,
        freight_code,
        booking_status,
        delivery_xml_path,
        manual_carrier,
        manual_passive_cost,
        manual_service_level,
        manual_freight_code,
        required_delivery_date,
        manual_pallets,
    ) in db_rows:
        try:
            row = json.loads(payload_json)
        except json.JSONDecodeError:
            continue
        if freight_code and not clean_text(row.get("Freight Code")):
            row["Freight Code"] = freight_code
        if booking_status:
            row["Prenotazione Scarico"] = booking_status
        row["Stato"] = status
        row["Data Partenza"] = departed_at or ""
        row["Data Scarico Prenotato"] = unload_date or ""
        row["Ora Scarico Prenotato"] = unload_time or ""
        row["Riferimento Booking Scarico"] = unload_booking_ref or ""
        sync_unload_booking_fields(row)
        row["Data Consegna"] = delivered_at or ""
        row["XML Consegna"] = delivery_xml_path or ""
        if manual_carrier:
            row["Carrier Scelto"] = manual_carrier
        if manual_service_level:
            apply_manual_service_level_to_row(row, manual_service_level)
        if manual_freight_code:
            apply_manual_freight_code_to_row(
                row,
                manual_freight_code,
                required_delivery_date,
                clear_required_for_dkl=not bool(required_delivery_date),
            )
        else:
            row["Data Consegna Tassativa"] = required_delivery_date or clean_text(row.get("Data Consegna Tassativa"))
            if is_dkv_order(row) and clean_text(row.get("Data Consegna Tassativa")):
                hide_early_delivery_for_dkv(row)
        if not is_dkv_order(row) and not clean_text(row.get("Early Delivery Date")):
            row["Early Delivery Date"] = clean_text(row.get("Late Delivery Date"))
        if manual_pallets is not None:
            apply_manual_pallets_to_row(row, manual_pallets)
        if manual_passive_cost is not None:
            row["Costo Passivo Manuale"] = manual_passive_cost
        apply_unload_booking_status(row, booking_status)
        row["Data Pianifica"] = planned_at or ""
        rows.append(row)

    apply_warehouse_contacts_to_shipments(rows)

    settings = load_settings()
    active_path_text = settings.get("active_rates_path")
    brt_path_text = settings.get("brt_passive_path")
    active_path = Path(active_path_text) if active_path_text else DEFAULT_ACTIVE_PATH
    brt_path = Path(brt_path_text) if brt_path_text else DEFAULT_BRT_PATH
    active_rates_path = active_path if active_path.exists() else None
    brt_passive_path = brt_path if brt_path.exists() else None
    carrier_tariffs_path = default_carrier_tariffs_path()
    if rows and (active_rates_path or brt_passive_path or carrier_tariffs_path):
        apply_tariffs_to_shipments(
            rows,
            active_rates_path=active_rates_path,
            brt_passive_pdf_path=brt_passive_path,
            brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
            gdo_customers_path=available_gdo_customers_path(),
            fuel_settings_path=available_fuel_settings_path(),
        )
    for row in rows:
        if to_float(row.get("Costo Passivo Manuale")) is not None:
            apply_manual_passive_to_row(row)
    if rows and active_rates_path:
        apply_contract_sla_to_shipments(rows, active_rates_path=active_rates_path)

    with SHIPMENTS_CACHE_LOCK:
        SHIPMENTS_CACHE["key"] = cache_key
        SHIPMENTS_CACHE["rows"] = copy.deepcopy(rows)
    return rows


def find_outbound_reports(downloads_dir: Path = DOWNLOADS_DIR) -> list[Path]:
    if not downloads_dir.exists():
        return []
    candidates: list[Path] = []
    for path in downloads_dir.iterdir():
        if not path.is_file() or path.name.startswith("~$"):
            continue
        name = path.name.lower()
        if "outbound report (prod)" not in name:
            continue
        if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
            continue
        age_seconds = datetime.now().timestamp() - path.stat().st_mtime
        if age_seconds < DOWNLOAD_STABLE_SECONDS:
            continue
        candidates.append(path)
    return sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)


def run_import(
    vtech_path: Path,
    active_rates_path: Path | None,
    brt_passive_path: Path | None,
    save_db: bool = True,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    detail_rows = extract_vtech_rows(vtech_path)
    shipment_rows = build_shipment_rows(detail_rows)
    ensure_brt_extra_flags_template(shipment_rows)
    manual_carriers = load_manual_carriers() if save_db else {}
    manual_carriers_by_order = load_manual_carriers_by_order() if save_db else {}
    manual_passive_costs = load_manual_passive_costs() if save_db else {}
    manual_passive_costs_by_order = load_manual_passive_costs_by_order() if save_db else {}
    manual_pallets = load_manual_pallets() if save_db else {}
    manual_pallets_by_order = load_manual_pallets_by_order() if save_db else {}
    manual_service_levels = load_manual_service_levels() if save_db else {}
    manual_service_levels_by_order = load_manual_service_levels_by_order() if save_db else {}
    manual_freight_codes = load_manual_freight_codes() if save_db else {}
    manual_freight_codes_by_order = load_manual_freight_codes_by_order() if save_db else {}
    active_urgent_flags = load_active_urgent_flags() if save_db else {}
    active_urgent_flags_by_order = load_active_urgent_flags_by_order() if save_db else {}
    for row in shipment_rows:
        manual_carrier = manual_carriers.get(clean_text(row.get("Shipment")))
        if not manual_carrier:
            manual_carrier = manual_carriers_by_order.get(normalized_order_key(row.get("Orders")))
        if manual_carrier:
            row["Carrier Scelto"] = manual_carrier
        manual_service_level = manual_service_levels.get(clean_text(row.get("Shipment")))
        if not manual_service_level:
            manual_service_level = manual_service_levels_by_order.get(normalized_order_key(row.get("Orders")))
        if manual_service_level:
            apply_manual_service_level_to_row(row, manual_service_level)
        manual_freight = manual_freight_codes.get(clean_text(row.get("Shipment")))
        if not manual_freight:
            manual_freight = manual_freight_codes_by_order.get(normalized_order_key(row.get("Orders")))
        if manual_freight:
            apply_manual_freight_code_to_row(
                row,
                manual_freight[0],
                manual_freight[1],
                clear_required_for_dkl=not bool(manual_freight[1]),
            )
        manual_passive_cost = manual_passive_costs.get(clean_text(row.get("Shipment")))
        if manual_passive_cost is None:
            manual_passive_cost = manual_passive_costs_by_order.get(normalized_order_key(row.get("Orders")))
        if manual_passive_cost is not None:
            row["Costo Passivo Manuale"] = manual_passive_cost
        manual_pallet_value = manual_pallets.get(clean_text(row.get("Shipment")))
        if manual_pallet_value is None:
            manual_pallet_value = manual_pallets_by_order.get(normalized_order_key(row.get("Orders")))
        if manual_pallet_value is not None:
            apply_manual_pallets_to_row(row, manual_pallet_value)
        active_urgent = active_urgent_flags.get(clean_text(row.get("Shipment")))
        if not active_urgent:
            active_urgent = active_urgent_flags_by_order.get(normalized_order_key(row.get("Orders")))
        if active_urgent:
            apply_active_urgent_to_row(row, True)
    apply_warehouse_contacts_to_shipments(shipment_rows)
    apply_tariffs_to_shipments(
        shipment_rows,
        active_rates_path=active_rates_path if active_rates_path and active_rates_path.exists() else None,
        brt_passive_pdf_path=brt_passive_path if brt_passive_path and brt_passive_path.exists() else None,
        brt_extra_flags_path=BRT_EXTRA_FLAGS_PATH,
        gdo_customers_path=available_gdo_customers_path(),
        fuel_settings_path=available_fuel_settings_path(),
    )
    for row in shipment_rows:
        if to_float(row.get("Costo Passivo Manuale")) is not None:
            apply_manual_passive_to_row(row)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_csv(detail_rows, OUTPUT_DIR / "vtech_spedizioni_estratte.csv", REQUIRED_COLUMNS)
    write_csv(shipment_rows, OUTPUT_DIR / "vtech_spedizioni_riepilogo.csv", SHIPMENT_COLUMNS)
    summary = build_summary(detail_rows, shipment_rows)
    (OUTPUT_DIR / "vtech_spedizioni_summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False, default=serialize),
        encoding="utf-8",
    )

    db_result = {"inserted": 0, "updated": 0, "total": len(shipment_rows)}
    if save_db:
        db_result = save_shipments_to_db(shipment_rows, vtech_path)
        mark_file_imported(vtech_path)
    summary["db"] = db_result

    return detail_rows, shipment_rows, summary


class VTechApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("V-Tech Trasporti")
        self.geometry("1500x900")
        self.minsize(1280, 760)

        init_db()
        self.shipments: list[dict[str, Any]] = []
        self.selected_shipment = ""
        self.selected_shipments: list[str] = []
        self.trees: dict[str, ttk.Treeview] = {}
        self.tree_max_heights: dict[str, int] = {}
        self.filters: dict[str, str] = {}
        self.import_running = False
        self.settings = load_settings()
        self.column_order = self._load_column_order()
        self.column_drag_source: str | None = None
        self.column_drag_start_x = 0
        self.column_drag_moved = False

        self.vtech_var = tk.StringVar(value=self._initial_path("vtech_path", DEFAULT_VTECH_PATH))
        self.active_var = tk.StringVar(value=self._initial_path("active_rates_path", DEFAULT_ACTIVE_PATH))
        self.brt_var = tk.StringVar(value=self._initial_path("brt_passive_path", DEFAULT_BRT_PATH))
        self.status_var = tk.StringVar(value="Pronto")
        self.kpi_var = tk.StringVar(value="Nessuna spedizione importata")
        self.kpi_open_var = tk.StringVar(value="0")
        self.kpi_planned_var = tk.StringVar(value="0")
        self.kpi_customer_var = tk.StringVar(value="0")
        self.kpi_delivered_var = tk.StringVar(value="0")
        self.kpi_margin_var = tk.StringVar(value="EUR 0,00")
        self.kpi_filters_var = tk.StringVar(value="Nessun filtro")
        self.groupage_mail_date_var = tk.StringVar(value=date.today().strftime("%d/%m/%y"))
        self.page_title_var = tk.StringVar(value="Spedizioni")
        self.page_subtitle_var = tk.StringVar(value="Da pianificare, margini e assegnazione vettori")
        self.pages: dict[str, ttk.Frame] = {}
        self.page_canvases: dict[str, tk.Canvas] = {}
        self.nav_buttons: dict[str, tk.Button] = {}
        self.metric_cards: list[dict[str, Any]] = []
        self.current_page = "shipments"

        self._configure_style()
        self._build_ui()
        repaired_early_dates = self._refresh_early_delivery_dates_safely()
        self.shipments = load_shipments_from_db()
        self.refresh_tables()
        if self.shipments:
            repair_note = f" Early delivery aggiornate: {repaired_early_dates}." if repaired_early_dates else ""
            self.status_var.set(f"Storico locale caricato: {len(self.shipments)} spedizioni.{repair_note}")
        self.after(2000, self.scan_downloads_for_new_reports)

    def _initial_path(self, key: str, default: Path) -> str:
        saved = self.settings.get(key, "")
        if isinstance(saved, str) and saved and Path(saved).exists():
            return saved
        return str(default) if default.exists() else ""

    def _load_column_order(self) -> list[str]:
        saved = self.settings.get("column_order", [])
        if not isinstance(saved, list):
            return DISPLAY_COLUMNS[:]

        valid_saved = [
            clean_text(column)
            for column in saved
            if clean_text(column) in DISPLAY_COLUMNS
        ]
        column_order = valid_saved[:]
        for column in DISPLAY_COLUMNS:
            if column in column_order:
                continue
            default_index = DISPLAY_COLUMNS.index(column)
            previous_columns = [
                candidate for candidate in DISPLAY_COLUMNS[:default_index]
                if candidate in column_order
            ]
            if previous_columns:
                insert_after = column_order.index(previous_columns[-1]) + 1
                column_order.insert(insert_after, column)
            else:
                column_order.append(column)
        return column_order

    def _save_column_order(self) -> None:
        self.settings["column_order"] = self.column_order[:]
        self._save_current_settings()

    def _apply_column_order(self) -> None:
        for tree in self.trees.values():
            tree.configure(displaycolumns=self.column_order)

    def _refresh_early_delivery_dates_safely(self) -> int:
        try:
            vtech_path = Path(self.vtech_var.get()) if self.vtech_var.get() else None
            refreshed = refresh_early_delivery_dates_from_source(vtech_path)
            normalize_saved_date_columns()
            return refreshed
        except Exception as exc:
            self.status_var.set(f"Aggiornamento Early Delivery non riuscito: {exc}")
            return 0

    def _configure_style(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")
        style.layout("Treeview", [("Treeview.treearea", {"sticky": "nswe"})])
        self.configure(background=APP_BG)
        style.configure("TFrame", background=MAIN_BG)
        style.configure("Surface.TFrame", background=SURFACE_BG)
        style.configure("Header.TFrame", background=APP_BG)
        style.configure("Toolbar.TFrame", background=SURFACE_BG)
        style.configure("PageHead.TFrame", background=MAIN_BG)
        style.configure("Panel.TLabelframe", background=SURFACE_BG, bordercolor=LINE_COLOR, relief="flat")
        style.configure(
            "Panel.TLabelframe.Label",
            font=("Segoe UI Semibold", 10),
            background=MAIN_BG,
            foreground=TEXT_COLOR,
        )
        style.configure("TLabel", background=MAIN_BG, foreground=TEXT_COLOR, font=("Segoe UI", 9))
        style.configure("Panel.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("Segoe UI", 9))
        style.configure("Muted.TLabel", background=MAIN_BG, foreground=MUTED_COLOR, font=("Segoe UI", 9))
        style.configure("HeaderTitle.TLabel", background=APP_BG, foreground=TEXT_COLOR, font=("Segoe UI Semibold", 17))
        style.configure("HeaderSub.TLabel", background=APP_BG, foreground=MUTED_COLOR, font=("Segoe UI", 9))
        style.configure("PageTitle.TLabel", background=MAIN_BG, foreground=TEXT_COLOR, font=("Segoe UI Semibold", 22))
        style.configure("PageSub.TLabel", background=MAIN_BG, foreground=MUTED_COLOR, font=("Segoe UI", 10))
        style.configure("Kpi.TLabel", background=SURFACE_BG, foreground=TEXT_COLOR, font=("Segoe UI Semibold", 10), padding=(14, 8))
        style.configure("TButton", font=("Segoe UI Semibold", 9), padding=(12, 7), borderwidth=0)
        style.configure("Accent.TButton", background=ACTION_BLUE, foreground="#ffffff")
        style.map("Accent.TButton", background=[("active", "#1d4ed8"), ("disabled", "#93c5fd")])
        style.configure("Secondary.TButton", background=SURFACE_SOFT, foreground=TEXT_COLOR)
        style.map("Secondary.TButton", background=[("active", "#1f3655")])
        style.configure("Success.TButton", background=SUCCESS_GREEN, foreground="#052e16")
        style.map("Success.TButton", background=[("active", "#16a34a")])
        style.configure("Warning.TButton", background=WARNING_AMBER, foreground="#111827")
        style.map("Warning.TButton", background=[("active", "#d97706")])
        style.configure("Browse.TButton", background=SURFACE_SOFT, foreground=TEXT_COLOR, padding=(10, 5))
        style.map("Browse.TButton", background=[("active", "#1f3655")])
        style.configure(
            "TEntry",
            padding=(8, 5),
            fieldbackground="#0b1626",
            foreground=TEXT_COLOR,
            insertcolor=TEXT_COLOR,
            bordercolor=LINE_COLOR,
            lightcolor=LINE_COLOR,
            darkcolor=LINE_COLOR,
        )
        style.configure(
            "Treeview",
            background=TABLE_BG,
            fieldbackground=TABLE_BG,
            foreground="#dbeafe",
            font=("Segoe UI", 9),
            rowheight=31,
            borderwidth=0,
            padding=0,
            relief="flat",
            bordercolor=TABLE_BG,
            lightcolor=TABLE_BG,
            darkcolor=TABLE_BG,
        )
        style.configure(
            "Treeview.Heading",
            background="#12243a",
            foreground="#c7d2fe",
            font=("Segoe UI Semibold", 9),
            relief="flat",
            padding=(8, 7),
        )
        style.map("Treeview", background=[("selected", "#1d4ed8")], foreground=[("selected", "#ffffff")])
        style.configure(
            "Vertical.TScrollbar",
            gripcount=0,
            background="#263b55",
            troughcolor=TABLE_BG,
            bordercolor=TABLE_BG,
            arrowcolor=MUTED_COLOR,
            lightcolor="#263b55",
            darkcolor="#263b55",
        )
        style.configure(
            "Horizontal.TScrollbar",
            gripcount=0,
            background="#263b55",
            troughcolor=TABLE_BG,
            bordercolor=TABLE_BG,
            arrowcolor=MUTED_COLOR,
            lightcolor="#263b55",
            darkcolor="#263b55",
        )
        style.configure("TNotebook", background=MAIN_BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(18, 7), font=("Segoe UI Semibold", 9), background=SURFACE_BG, foreground=MUTED_COLOR)
        style.map("TNotebook.Tab", background=[("selected", TABLE_BG)], foreground=[("selected", TEXT_COLOR)])

    def _build_ui(self) -> None:
        root = tk.Frame(self, bg=APP_BG)
        root.pack(fill=tk.BOTH, expand=True)

        sidebar = tk.Frame(root, bg=SIDEBAR_BG, width=248)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)
        self._build_sidebar(sidebar)

        main = tk.Frame(root, bg=MAIN_BG, padx=18, pady=14)
        main.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        page_head = tk.Frame(
            main,
            bg=SURFACE_BG,
            highlightbackground=LINE_COLOR,
            highlightthickness=1,
            bd=0,
            padx=18,
            pady=14,
        )
        page_head.pack(fill=tk.X, pady=(0, 10))

        title_block = tk.Frame(page_head, bg=SURFACE_BG)
        title_block.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Label(
            title_block,
            text="V-TECH TRANSPORT CONTROL ROOM",
            bg=SURFACE_BG,
            fg=ACCENT_BLUE,
            font=("Segoe UI Semibold", 8),
            anchor="w",
        ).pack(anchor="w")
        tk.Label(
            title_block,
            textvariable=self.page_title_var,
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            font=("Segoe UI Semibold", 25),
            anchor="w",
        ).pack(anchor="w")
        tk.Label(
            title_block,
            textvariable=self.page_subtitle_var,
            bg=SURFACE_BG,
            fg=MUTED_COLOR,
            font=("Segoe UI", 10),
            anchor="w",
        ).pack(anchor="w", pady=(3, 0))
        chip_row = tk.Frame(title_block, bg=SURFACE_BG)
        chip_row.pack(anchor="w", pady=(10, 0))
        for text, color in [
            ("LIVE", SUCCESS_GREEN),
            ("BRT LTL", ACCENT_BLUE),
            ("KN LTL/FTL", "#8b5cf6"),
            ("MARGINI", WARNING_AMBER),
        ]:
            tk.Label(
                chip_row,
                text=text,
                bg="#0b1626",
                fg=color,
                font=("Segoe UI Semibold", 8),
                padx=10,
                pady=3,
                highlightbackground=LINE_COLOR,
                highlightthickness=1,
            ).pack(side=tk.LEFT, padx=(0, 7))

        actions = tk.Frame(page_head, bg=SURFACE_BG)
        actions.pack(side=tk.RIGHT, anchor="e", padx=(16, 0))
        buttons = [
            ("Importa e calcola", ACTION_BLUE, "#ffffff", self.import_data),
            ("Scegli vettore", SURFACE_SOFT, TEXT_COLOR, self.choose_selected_carrier),
            ("Pianificato", WARNING_AMBER, "#111827", self.mark_selected_planned),
            ("Da pianificare", SURFACE_SOFT, TEXT_COLOR, self.mark_selected_unplanned),
            ("Consegnata / XML", SUCCESS_GREEN, "#052e16", self.mark_selected_delivered),
            ("Seleziona tutto", SURFACE_SOFT, TEXT_COLOR, self.select_all_current_page),
            ("Elimina", DANGER_RED, "#ffffff", self.delete_selected_shipments),
            ("Pulisci filtri", SURFACE_SOFT, TEXT_COLOR, self.clear_all_filters),
            ("Aggiorna", SURFACE_SOFT, TEXT_COLOR, self.refresh_tables),
        ]
        for index, (text, bg, fg, command) in enumerate(buttons):
            self._action_button(actions, text, bg, fg, command).grid(
                row=index // 5,
                column=index % 5,
                sticky="ew",
                padx=(0, 7),
                pady=(0, 6),
            )

        files = tk.Frame(main, bg=SURFACE_BG, bd=0)
        files.pack(fill=tk.X, pady=(0, 8))
        tk.Label(
            files,
            text="Sorgenti dati operative",
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            font=("Segoe UI Semibold", 10),
            anchor="w",
        ).pack(fill=tk.X, padx=12, pady=(8, 1))
        files_body = tk.Frame(files, bg=SURFACE_BG)
        files_body.pack(fill=tk.X, padx=12, pady=(0, 8))

        self._file_row(files_body, 0, "Report V-Tech", self.vtech_var, [("Excel", "*.xlsx *.xlsm *.xls")])
        self._file_row(files_body, 1, "Tariffe attive", self.active_var, [("Excel", "*.xlsx *.xlsm *.xls")])
        self._file_row(files_body, 2, "Passiva BRT", self.brt_var, [("PDF", "*.pdf")])

        files_body.columnconfigure(1, weight=1)

        self._build_kpi_cards(main)

        page_area = tk.Frame(main, bg=MAIN_BG)
        page_area.pack(fill=tk.BOTH, expand=True)
        page_area.rowconfigure(0, weight=1)
        page_area.columnconfigure(0, weight=1)

        shipments_page = self._make_page(page_area, "shipments")
        customer_page = self._make_page(page_area, "customer")
        groupage_page = self._make_page(page_area, "planned_groupage")
        delivered_page = self._make_page(page_area, "delivered")

        self.open_groupage_tree = self._make_board(shipments_page, "open_groupage", "Groupage da pianificare - BRT LTL", height=13)
        self.open_direct_tree = self._make_board(shipments_page, "open_direct", "Dirette da pianificare - KN LTL / FTL e altri vettori", height=13)
        self.customer_planned_tree = self._make_board(customer_page, "planned_customer", "Customer - dirette pianificate", height=24)
        self.planned_direct_tree = self.customer_planned_tree
        self._build_groupage_mail_bar(groupage_page)
        self.planned_groupage_tree = self._make_board(groupage_page, "planned_groupage", "Groupage pianificato - BRT LTL", height=24)
        self.delivered_tree = self._make_board(delivered_page, "delivered", "Spedizioni consegnate", height=24)

        bottom = tk.Frame(main, bg=SURFACE_BG, bd=0)
        bottom.pack(fill=tk.X, pady=(8, 0))
        tk.Label(
            bottom,
            text="Dettaglio spedizione",
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            font=("Segoe UI Semibold", 10),
            anchor="w",
        ).pack(fill=tk.X, padx=12, pady=(8, 0))
        self.detail_text = tk.Text(
            bottom,
            height=2,
            wrap="word",
            font=("Segoe UI", 9),
            bg=SURFACE_BG,
            fg="#dbeafe",
            relief="flat",
            padx=10,
            pady=6,
            insertbackground=TEXT_COLOR,
        )
        self.detail_text.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6, pady=(0, 7))
        self.detail_text.configure(state="disabled")

        status = ttk.Label(main, textvariable=self.status_var, anchor="w", style="Muted.TLabel", padding=(2, 8))
        status.pack(fill=tk.X)

        self.show_page("shipments")

    def _action_button(
        self,
        parent: tk.Frame,
        text: str,
        bg: str,
        fg: str,
        command: Any,
    ) -> tk.Button:
        border = LINE_COLOR if bg in {SURFACE_BG, SURFACE_SOFT, TABLE_BG} else bg
        button = tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg=fg,
            activebackground=self._hover_color(bg),
            activeforeground=fg,
            font=("Segoe UI Semibold", 9),
            bd=0,
            relief=tk.FLAT,
            highlightthickness=1,
            highlightbackground=border,
            padx=12,
            pady=7,
            cursor="hand2",
        )
        button.bind("<Enter>", lambda _event, widget=button, color=bg: widget.configure(bg=self._hover_color(color)))
        button.bind("<Leave>", lambda _event, widget=button, color=bg: widget.configure(bg=color))
        return button

    def _hover_color(self, color: str) -> str:
        hover_map = {
            ACTION_BLUE: "#1d4ed8",
            SURFACE_SOFT: "#1d334f",
            SUCCESS_GREEN: "#16a34a",
            WARNING_AMBER: "#d97706",
            DANGER_RED: "#dc2626",
        }
        return hover_map.get(color, color)

    def _draw_round_rect(
        self,
        canvas: tk.Canvas,
        x1: int,
        y1: int,
        x2: int,
        y2: int,
        radius: int,
        fill: str,
        outline: str = "",
        width: int = 1,
    ) -> None:
        points = [
            x1 + radius,
            y1,
            x2 - radius,
            y1,
            x2,
            y1,
            x2,
            y1 + radius,
            x2,
            y2 - radius,
            x2,
            y2,
            x2 - radius,
            y2,
            x1 + radius,
            y2,
            x1,
            y2,
            x1,
            y2 - radius,
            x1,
            y1 + radius,
            x1,
            y1,
        ]
        canvas.create_polygon(
            points,
            smooth=True,
            splinesteps=18,
            fill=fill,
            outline=outline,
            width=width,
        )

    def _draw_metric_card(self, config: dict[str, Any]) -> None:
        canvas: tk.Canvas = config["canvas"]
        width = max(170, canvas.winfo_width())
        height = max(112, canvas.winfo_height())
        title = clean_text(config["title"]).upper()
        value = clean_text(config["variable"].get())
        subtitle = clean_text(config["subtitle"])
        accent = clean_text(config["accent"])
        icon = clean_text(config["icon"])

        canvas.delete("all")
        self._draw_round_rect(canvas, 2, 3, width - 3, height - 4, 16, "#101f34", "#2b4260", 1)
        self._draw_round_rect(canvas, 8, 9, width - 9, 36, 14, "#152a44", "", 0)
        canvas.create_line(18, height - 9, width - 18, height - 9, fill="#1f3b5c", width=2)
        canvas.create_line(18, height - 9, max(42, int(width * 0.42)), height - 9, fill=accent, width=2)
        self._draw_round_rect(canvas, 16, 42, 58, 84, 12, accent, "", 0)
        canvas.create_text(
            37,
            63,
            text=icon,
            fill="#ffffff" if accent != WARNING_AMBER else "#111827",
            font=("Segoe UI Semibold", 11),
        )
        canvas.create_text(
            72,
            20,
            text=title,
            fill=MUTED_COLOR,
            font=("Segoe UI Semibold", 8),
            anchor="nw",
            width=max(80, width - 88),
        )
        value_font = 15 if value.startswith("EUR ") else 18 if len(value) > 9 else 24
        canvas.create_text(
            72,
            47,
            text=value,
            fill=TEXT_COLOR,
            font=("Segoe UI Semibold", value_font),
            anchor="nw",
        )
        canvas.create_text(
            72,
            78,
            text=subtitle,
            fill="#b6c6dc",
            font=("Segoe UI", 8),
            anchor="nw",
            width=max(80, width - 88),
        )

    def _build_kpi_cards(self, parent: ttk.Frame) -> None:
        kpi_row = tk.Frame(parent, bg=MAIN_BG)
        kpi_row.pack(fill=tk.X, pady=(0, 10))
        cards = [
            ("Da pianificare", self.kpi_open_var, ACTION_BLUE, "SP", "nuove spedizioni operative"),
            ("Pianificate", self.kpi_planned_var, "#8b5cf6", "PL", "gia pronte per gestione"),
            ("Customer", self.kpi_customer_var, "#06b6d4", "CU", "dirette KN e altri vettori"),
            ("Consegnate", self.kpi_delivered_var, SUCCESS_GREEN, "OK", "storico consegne"),
            ("Margine stimato", self.kpi_margin_var, "#14b8a6", "EU", "attivo meno passivo"),
            ("Filtri", self.kpi_filters_var, WARNING_AMBER, "FX", "vista colonne attiva"),
        ]
        self.metric_cards = []
        for index, (title, variable, accent, icon, subtitle) in enumerate(cards):
            card = tk.Canvas(kpi_row, width=185, height=112, bg=MAIN_BG, bd=0, highlightthickness=0)
            card.grid(row=0, column=index, sticky="nsew", padx=(0, 10 if index < len(cards) - 1 else 0))
            kpi_row.columnconfigure(index, weight=1)
            config = {
                "canvas": card,
                "title": title,
                "variable": variable,
                "accent": accent,
                "icon": icon,
                "subtitle": subtitle,
            }
            self.metric_cards.append(config)

            def redraw_metric(_event: tk.Event | None = None, item: dict[str, Any] = config) -> None:
                self._draw_metric_card(item)

            card.bind("<Configure>", redraw_metric)
            variable.trace_add("write", lambda *_args, item=config: self._draw_metric_card(item))
            self._draw_metric_card(config)

    def _build_groupage_mail_bar(self, parent: ttk.Frame) -> None:
        bar = tk.Frame(parent, bg=SURFACE_BG, bd=0)
        bar.pack(fill=tk.X, pady=(0, 10))

        copy = tk.Frame(bar, bg=SURFACE_BG)
        copy.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=14, pady=12)
        tk.Label(
            copy,
            text="Mail ritiri giornalieri",
            bg=SURFACE_BG,
            fg=TEXT_COLOR,
            font=("Segoe UI Semibold", 11),
            anchor="w",
        ).pack(anchor="w")
        tk.Label(
            copy,
            text="Crea una bozza Outlook con i groupage pianificati selezionati.",
            bg=SURFACE_BG,
            fg=MUTED_COLOR,
            font=("Segoe UI", 9),
            anchor="w",
        ).pack(anchor="w", pady=(2, 0))

        controls = tk.Frame(bar, bg=SURFACE_BG)
        controls.pack(side=tk.RIGHT, padx=14, pady=12)
        self._action_button(
            controls,
            "Genera mail selezionate",
            ACTION_BLUE,
            "#ffffff",
            self.create_groupage_outlook_mail,
        ).grid(row=0, column=0, sticky="ew")

    def _build_sidebar(self, sidebar: tk.Frame) -> None:
        brand = tk.Frame(sidebar, bg=SIDEBAR_BG)
        brand.pack(fill=tk.X, padx=18, pady=(20, 18))
        tk.Label(
            brand,
            text="CONTROL TOWER",
            bg=SIDEBAR_BG,
            fg=ACCENT_BLUE,
            font=("Segoe UI Semibold", 8),
            anchor="w",
        ).pack(fill=tk.X, pady=(0, 5))
        tk.Label(
            brand,
            text="V-Tech",
            bg=SIDEBAR_BG,
            fg=TEXT_COLOR,
            font=("Segoe UI Semibold", 22),
            anchor="w",
        ).pack(fill=tk.X)
        tk.Label(
            brand,
            text="Trasporti",
            bg=SIDEBAR_BG,
            fg=MUTED_COLOR,
            font=("Segoe UI", 10),
            anchor="w",
        ).pack(fill=tk.X, pady=(2, 0))

        tk.Label(
            sidebar,
            text="NAVIGAZIONE",
            bg=SIDEBAR_BG,
            fg="#64748b",
            font=("Segoe UI Semibold", 8),
            anchor="w",
        ).pack(fill=tk.X, padx=18, pady=(8, 8))

        nav = tk.Frame(sidebar, bg=SIDEBAR_BG)
        nav.pack(fill=tk.X, padx=12)
        self._add_nav_button(nav, "shipments", "Spedizioni")
        self._add_nav_button(nav, "customer", "Customer")
        self._add_nav_button(nav, "planned_groupage", "Groupage pianificato")
        self._add_nav_button(nav, "delivered", "Spedizioni consegnate")

        tk.Frame(sidebar, bg=LINE_COLOR, height=1).pack(fill=tk.X, padx=18, pady=(20, 14))
        tk.Label(
            sidebar,
            text="Seleziona una riga per vedere dettaglio, extra e note. Trascina le intestazioni per riordinare le colonne.",
            bg=SIDEBAR_BG,
            fg=MUTED_COLOR,
            font=("Segoe UI", 9),
            justify=tk.LEFT,
            wraplength=180,
            anchor="w",
        ).pack(fill=tk.X, padx=18)

    def _add_nav_button(self, parent: tk.Frame, page_key: str, text: str) -> None:
        button = tk.Button(
            parent,
            text=text,
            command=lambda: self.show_page(page_key),
            anchor="w",
            bd=0,
            relief=tk.FLAT,
            highlightthickness=0,
            padx=16,
            pady=12,
            bg=SIDEBAR_BG,
            fg="#cbd5e1",
            activebackground=SIDEBAR_HOVER,
            activeforeground=TEXT_COLOR,
            font=("Segoe UI Semibold", 10),
            cursor="hand2",
        )
        button.pack(fill=tk.X, pady=(0, 7))
        self.nav_buttons[page_key] = button

    def _make_page(self, parent: ttk.Frame, key: str) -> ttk.Frame:
        page = tk.Frame(parent, bg=MAIN_BG)
        page.grid(row=0, column=0, sticky="nsew")

        canvas = tk.Canvas(page, background=MAIN_BG, highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(page, orient=tk.VERTICAL, command=canvas.yview)
        content = tk.Frame(canvas, bg=MAIN_BG)
        window_id = canvas.create_window((0, 0), window=content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        page.rowconfigure(0, weight=1)
        page.columnconfigure(0, weight=1)

        def refresh_scroll_region(_event: tk.Event) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def stretch_content(event: tk.Event) -> None:
            canvas.itemconfigure(window_id, width=event.width)

        content.bind("<Configure>", refresh_scroll_region)
        canvas.bind("<Configure>", stretch_content)
        self.pages[key] = page
        self.page_canvases[key] = canvas
        return content

    def show_page(self, page_key: str) -> None:
        titles = {
            "shipments": ("Spedizioni", "Da pianificare, margini e assegnazione vettori"),
            "customer": ("Customer", "Spedizioni pianificate dirette KN LTL/FTL e altri vettori"),
            "planned_groupage": ("Groupage pianificato", "Spedizioni BRT LTL gia pianificate"),
            "delivered": ("Spedizioni consegnate", "Storico consegne con possibilita di riportarle indietro"),
        }
        for key, page in self.pages.items():
            if key == page_key:
                page.grid()
            else:
                page.grid_remove()

        for key, button in self.nav_buttons.items():
            if key == page_key:
                button.configure(bg=ACTION_BLUE, fg="#ffffff", activebackground="#1d4ed8")
            else:
                button.configure(bg=SIDEBAR_BG, fg="#cbd5e1", activebackground=SIDEBAR_HOVER)

        title, subtitle = titles.get(page_key, titles["shipments"])
        self.page_title_var.set(title)
        self.page_subtitle_var.set(subtitle)
        self.current_page = page_key

        canvas = self.page_canvases.get(page_key)
        if canvas is not None:
            canvas.yview_moveto(0)
            self.bind_all("<MouseWheel>", lambda event, target=canvas: target.yview_scroll(int(-1 * (event.delta / 120)), "units"))

    def _make_scrollable_tab(self, notebook: ttk.Notebook, title: str) -> ttk.Frame:
        outer = ttk.Frame(notebook)
        notebook.add(outer, text=title)

        canvas = tk.Canvas(outer, background="#f3f6fb", highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        content = ttk.Frame(canvas, padding=(0, 8, 8, 0))
        window_id = canvas.create_window((0, 0), window=content, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        outer.rowconfigure(0, weight=1)
        outer.columnconfigure(0, weight=1)

        def refresh_scroll_region(_event: tk.Event) -> None:
            canvas.configure(scrollregion=canvas.bbox("all"))

        def stretch_content(event: tk.Event) -> None:
            canvas.itemconfigure(window_id, width=event.width)

        def scroll_page(event: tk.Event) -> None:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        content.bind("<Configure>", refresh_scroll_region)
        canvas.bind("<Configure>", stretch_content)
        canvas.bind("<Enter>", lambda _event: canvas.bind_all("<MouseWheel>", scroll_page))
        canvas.bind("<Leave>", lambda _event: canvas.unbind_all("<MouseWheel>"))
        return content

    def _file_row(
        self,
        parent: tk.Frame,
        row: int,
        label: str,
        variable: tk.StringVar,
        filetypes: list[tuple[str, str]],
    ) -> None:
        tk.Label(
            parent,
            text=label,
            bg=SURFACE_BG,
            fg=MUTED_COLOR,
            font=("Segoe UI Semibold", 9),
            anchor="w",
        ).grid(row=row, column=0, sticky="w", pady=3)
        entry = ttk.Entry(parent, textvariable=variable)
        entry.grid(row=row, column=1, sticky="ew", padx=8, pady=3)
        ttk.Button(
            parent,
            text="Sfoglia",
            style="Browse.TButton",
            command=lambda: self.select_file(variable, filetypes),
        ).grid(row=row, column=2, sticky="ew", pady=3)

    def _make_board(self, parent: ttk.Frame, key: str, title: str, height: int = 12) -> ttk.Treeview:
        section = tk.Frame(parent, bg=MAIN_BG, bd=0, highlightthickness=0)
        section.pack(fill=tk.X, expand=False, pady=(0, 16))
        header = tk.Frame(section, bg=MAIN_BG)
        header.pack(fill=tk.X, pady=(0, 8))
        tk.Frame(header, bg=ACCENT_BLUE, width=4, height=22).pack(side=tk.LEFT, padx=(0, 9))
        tk.Label(
            header,
            text=title,
            bg=MAIN_BG,
            fg=TEXT_COLOR,
            font=("Segoe UI Semibold", 13),
            anchor="w",
        ).pack(side=tk.LEFT)
        tk.Label(
            header,
            text="clic intestazione = filtro  |  trascina = riordina",
            bg=MAIN_BG,
            fg=MUTED_COLOR,
            font=("Segoe UI", 8),
            anchor="e",
        ).pack(side=tk.RIGHT)

        frame = tk.Frame(section, bg=MAIN_BG, bd=0, highlightthickness=0)
        frame.pack(fill=tk.X, expand=False)

        tree = ttk.Treeview(
            frame,
            columns=DISPLAY_COLUMNS,
            show="headings",
            selectmode="extended",
            height=4,
        )
        yscroll = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        xscroll = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)

        for column in DISPLAY_COLUMNS:
            tree.heading(
                column,
                text=f"{COLUMN_TITLES.get(column, column)} v",
            )
            tree.column(column, width=COLUMN_WIDTHS.get(column, 100), minwidth=50, stretch=False)
        tree.configure(displaycolumns=self.column_order)

        tree.tag_configure("row_even", background=TABLE_BG, foreground="#dbeafe")
        tree.tag_configure("row_odd", background="#0f1c2e", foreground="#dbeafe")
        tree.tag_configure("loss", background="#241724", foreground="#fecdd3")
        tree.tag_configure("sla", background="#3a210f", foreground="#fed7aa")
        tree.tag_configure("missing", background="#211f2a", foreground="#fde68a")
        tree.tag_configure("planned", background="#112742", foreground="#dbeafe")
        tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        tree.bind("<ButtonRelease-1>", self.on_tree_click)
        tree.bind("<ButtonPress-1>", self.on_tree_button_press, add="+")
        tree.bind("<B1-Motion>", self.on_tree_drag_motion, add="+")

        tree.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)
        self.trees[key] = tree
        self.tree_max_heights[key] = height
        return tree

    def select_file(self, variable: tk.StringVar, filetypes: list[tuple[str, str]]) -> None:
        selected = filedialog.askopenfilename(filetypes=filetypes + [("Tutti i file", "*.*")])
        if selected:
            variable.set(selected)

    def open_column_filter(self, column: str) -> None:
        title = COLUMN_TITLES.get(column, column)
        source_rows = self._rows_for_column_filter(column)
        values = sorted(
            {
                display_value(row, column) if display_value(row, column) else EMPTY_FILTER_VALUE
                for row in source_rows
            },
            key=lambda value: "(vuoto)" if value == EMPTY_FILTER_VALUE else value.lower(),
        )

        window = tk.Toplevel(self)
        window.title(f"Filtro - {title}")
        window.geometry("420x520")
        window.configure(background="#edf1f5")
        window.transient(self)

        frame = ttk.Frame(window, padding=12)
        frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frame, text=f"Valori presenti nella colonna {title}", font=("Segoe UI Semibold", 11)).pack(anchor="w")

        search_var = tk.StringVar()
        ttk.Entry(frame, textvariable=search_var).pack(fill=tk.X, pady=(10, 8))

        listbox = tk.Listbox(frame, activestyle="dotbox", exportselection=False, height=16)
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=listbox.yview)
        listbox.configure(yscrollcommand=scrollbar.set)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        def label_for(value: str) -> str:
            return "(Vuoto)" if value == EMPTY_FILTER_VALUE else value

        current_values = values[:]

        def populate(filter_text: str = "") -> None:
            nonlocal current_values
            needle = filter_text.lower()
            current_values = [
                value for value in values
                if needle in label_for(value).lower()
            ]
            listbox.delete(0, tk.END)
            for value in current_values:
                listbox.insert(tk.END, label_for(value))

        def apply_selected() -> None:
            selection = listbox.curselection()
            if not selection:
                return
            self.set_column_filter(column, current_values[selection[0]])
            window.destroy()

        def clear_selected_filter() -> None:
            self.clear_column_filter(column)
            window.destroy()

        search_var.trace_add("write", lambda *_: populate(search_var.get()))
        listbox.bind("<Double-1>", lambda _event: apply_selected())
        populate()

        actions = ttk.Frame(window, padding=(12, 0, 12, 12))
        actions.pack(fill=tk.X)
        ttk.Button(actions, text="Applica", style="Accent.TButton", command=apply_selected).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(actions, text="Tutti", style="Secondary.TButton", command=clear_selected_filter).pack(side=tk.RIGHT)

    def set_column_filter(self, column: str, value: str) -> None:
        self.filters[column] = value
        self.refresh_tables()

    def clear_column_filter(self, column: str) -> None:
        self.filters.pop(column, None)
        self.refresh_tables()

    def clear_all_filters(self) -> None:
        self.filters = {}
        self.refresh_tables()

    def open_filters_window(self) -> None:
        window = tk.Toplevel(self)
        window.title("Filtri colonne")
        window.geometry("520x680")
        window.configure(background="#edf1f5")
        window.transient(self)

        container = ttk.Frame(window, padding=12)
        container.pack(fill=tk.BOTH, expand=True)
        ttk.Label(
            container,
            text="Inserisci un testo nelle colonne da filtrare. I filtri lavorano in modalita contiene.",
            style="Muted.TLabel",
            wraplength=480,
        ).pack(fill=tk.X, pady=(0, 10))

        canvas = tk.Canvas(container, background="#edf1f5", highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient=tk.VERTICAL, command=canvas.yview)
        fields_frame = ttk.Frame(canvas)
        fields_frame.bind(
            "<Configure>",
            lambda event: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas.create_window((0, 0), window=fields_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        entries: dict[str, tk.StringVar] = {}
        for row_index, column in enumerate(DISPLAY_COLUMNS):
            ttk.Label(fields_frame, text=COLUMN_TITLES.get(column, column)).grid(row=row_index, column=0, sticky="w", pady=3, padx=(0, 8))
            variable = tk.StringVar(value=self.filters.get(column, ""))
            entries[column] = variable
            ttk.Entry(fields_frame, textvariable=variable, width=34).grid(row=row_index, column=1, sticky="ew", pady=3)
        fields_frame.columnconfigure(1, weight=1)

        actions = ttk.Frame(window, padding=(12, 0, 12, 12))
        actions.pack(fill=tk.X)

        def apply_filters() -> None:
            self.filters = {
                column: clean_text(variable.get())
                for column, variable in entries.items()
                if clean_text(variable.get())
            }
            self.refresh_tables()
            window.destroy()

        def clear_filters() -> None:
            self.filters = {}
            self.refresh_tables()
            window.destroy()

        ttk.Button(actions, text="Applica", style="Accent.TButton", command=apply_filters).pack(side=tk.RIGHT, padx=(8, 0))
        ttk.Button(actions, text="Pulisci filtri", style="Secondary.TButton", command=clear_filters).pack(side=tk.RIGHT)

    def import_data(self) -> None:
        if self.import_running:
            return
        vtech_path = Path(self.vtech_var.get())
        if not vtech_path.exists():
            messagebox.showerror("File mancante", "Seleziona il file Excel V-Tech.")
            return

        self._save_current_settings()
        self.status_var.set("Importazione in corso...")
        self.import_running = True
        self._set_buttons_state(tk.DISABLED)

        thread = threading.Thread(target=self._import_worker, daemon=True)
        thread.start()

    def _import_worker(self) -> None:
        try:
            detail_rows, shipment_rows, summary = run_import(
                Path(self.vtech_var.get()),
                Path(self.active_var.get()) if self.active_var.get() else None,
                Path(self.brt_var.get()) if self.brt_var.get() else None,
                save_db=True,
            )
        except Exception as exc:
            self.after(0, self._import_failed, exc)
            return
        self.after(0, self._import_finished, shipment_rows, summary)

    def _import_finished(self, shipment_rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
        self.shipments = load_shipments_from_db()
        self.refresh_tables()
        db_result = summary.get("db", {})
        self.status_var.set(
            f"Importate {summary['shipment_rows']} spedizioni. "
            f"Nuove: {db_result.get('inserted', 0)}, aggiornate: {db_result.get('updated', 0)}. "
            f"Database locale: {DB_PATH}"
        )
        self.import_running = False
        self._set_buttons_state(tk.NORMAL)

    def _import_failed(self, exc: Exception) -> None:
        self.status_var.set("Importazione non riuscita")
        self.import_running = False
        self._set_buttons_state(tk.NORMAL)
        messagebox.showerror("Errore importazione", str(exc))

    def scan_downloads_for_new_reports(self) -> None:
        try:
            if not self.import_running:
                reports = find_outbound_reports()
                latest_report = reports[0] if reports else None
                new_reports = (
                    [latest_report]
                    if latest_report and not is_imported_file_current(latest_report)
                    else []
                )
                if new_reports:
                    self.import_running = True
                    self.status_var.set(f"Trovati {len(new_reports)} nuovi Outbound Report in Download...")
                    thread = threading.Thread(target=self._auto_import_worker, args=(new_reports,), daemon=True)
                    thread.start()
        finally:
            self.after(DOWNLOAD_SCAN_MS, self.scan_downloads_for_new_reports)

    def _auto_import_worker(self, paths: list[Path]) -> None:
        imported: list[dict[str, Any]] = []
        try:
            sorted_paths = sorted(paths, key=lambda item: item.stat().st_mtime)
            for path in sorted_paths:
                _, shipment_rows, summary = run_import(
                    path,
                    Path(self.active_var.get()) if self.active_var.get() else None,
                    Path(self.brt_var.get()) if self.brt_var.get() else None,
                    save_db=True,
                )
                imported.append(
                    {
                        "file": path.name,
                        "shipments": summary.get("shipment_rows", 0),
                        "inserted": summary.get("db", {}).get("inserted", 0),
                        "updated": summary.get("db", {}).get("updated", 0),
                    }
                )
            self.after(0, self._auto_import_finished, sorted_paths[-1], imported)
        except Exception as exc:
            self.after(0, self._auto_import_failed, exc)

    def _auto_import_finished(self, latest_path: Path, imported: list[dict[str, Any]]) -> None:
        self.vtech_var.set(str(latest_path))
        self._save_current_settings()
        self.shipments = load_shipments_from_db()
        self.refresh_tables()
        inserted = sum(item["inserted"] for item in imported)
        updated = sum(item["updated"] for item in imported)
        self.status_var.set(
            f"Download aggiornato automaticamente. Nuove: {inserted}, aggiornate: {updated}."
        )
        self.import_running = False

    def _auto_import_failed(self, exc: Exception) -> None:
        self.status_var.set(f"Import automatico non riuscito: {exc}")
        self.import_running = False

    def _set_buttons_state(self, state: str) -> None:
        for child in self.winfo_children():
            self._set_state_recursive(child, state)

    def _set_state_recursive(self, widget: tk.Widget, state: str) -> None:
        if isinstance(widget, ttk.Button):
            widget.configure(state=state)
        if isinstance(widget, tk.Button):
            widget.configure(state=state)
        for child in widget.winfo_children():
            self._set_state_recursive(child, state)

    def _is_brt_groupage(self, row: dict[str, Any]) -> bool:
        return clean_text(row.get("Tipo Servizio")) == "Groupage - BRT LTL"

    def _parse_groupage_mail_date(self) -> date:
        raw_text = clean_text(self.groupage_mail_date_var.get())
        if not raw_text:
            raise ValueError("Inserisci una data, ad esempio 18/5/26.")

        normalized = raw_text.replace(".", "/").replace("-", "/").replace(" ", "")
        parts = normalized.split("/")
        if len(parts) == 2:
            normalized = f"{parts[0]}/{parts[1]}/{date.today().year}"

        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
            try:
                return datetime.strptime(normalized, fmt).date()
            except ValueError:
                continue

        raise ValueError("Formato data non valido. Usa un formato tipo 18/5/26.")

    def open_groupage_mail_calendar(self) -> None:
        try:
            selected = self._parse_groupage_mail_date()
        except ValueError:
            selected = date.today()

        window = tk.Toplevel(self)
        window.title("Seleziona data Wave")
        window.configure(background="#ffffff")
        window.transient(self)
        window.grab_set()
        window.resizable(False, False)

        state = {"year": selected.year, "month": selected.month}
        header = tk.Frame(window, bg="#ffffff")
        header.pack(fill=tk.X, padx=12, pady=(12, 8))
        title_var = tk.StringVar()

        def shift_month(delta: int) -> None:
            month = state["month"] + delta
            year = state["year"]
            if month < 1:
                month = 12
                year -= 1
            elif month > 12:
                month = 1
                year += 1
            state["month"] = month
            state["year"] = year
            render_days()

        self._action_button(header, "<", "#ffffff", "#0f172a", lambda: shift_month(-1)).pack(side=tk.LEFT)
        tk.Label(
            header,
            textvariable=title_var,
            bg="#ffffff",
            fg="#0f172a",
            font=("Segoe UI Semibold", 11),
            width=18,
        ).pack(side=tk.LEFT, padx=8)
        self._action_button(header, ">", "#ffffff", "#0f172a", lambda: shift_month(1)).pack(side=tk.LEFT)

        grid = tk.Frame(window, bg="#ffffff")
        grid.pack(fill=tk.BOTH, padx=12, pady=(0, 12))

        def choose_day(day: int) -> None:
            chosen = date(state["year"], state["month"], day)
            self.groupage_mail_date_var.set(chosen.strftime("%d/%m/%y"))
            window.destroy()

        def render_days() -> None:
            for child in grid.winfo_children():
                child.destroy()

            month_names = [
                "",
                "Gennaio",
                "Febbraio",
                "Marzo",
                "Aprile",
                "Maggio",
                "Giugno",
                "Luglio",
                "Agosto",
                "Settembre",
                "Ottobre",
                "Novembre",
                "Dicembre",
            ]
            title_var.set(f"{month_names[state['month']]} {state['year']}")
            for col, label in enumerate(["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]):
                tk.Label(
                    grid,
                    text=label,
                    bg="#ffffff",
                    fg="#64748b",
                    font=("Segoe UI Semibold", 8),
                    width=5,
                ).grid(row=0, column=col, padx=2, pady=2)

            weeks = calendar.Calendar(firstweekday=0).monthdayscalendar(state["year"], state["month"])
            for row_index, week in enumerate(weeks, start=1):
                for col, day in enumerate(week):
                    if day == 0:
                        tk.Label(grid, text="", bg="#ffffff", width=5).grid(row=row_index, column=col, padx=2, pady=2)
                        continue
                    is_selected = (
                        day == selected.day
                        and state["month"] == selected.month
                        and state["year"] == selected.year
                    )
                    bg = "#2563eb" if is_selected else "#f8fafc"
                    fg = "#ffffff" if is_selected else "#0f172a"
                    tk.Button(
                        grid,
                        text=str(day),
                        command=lambda selected_day=day: choose_day(selected_day),
                        bg=bg,
                        fg=fg,
                        activebackground="#dbeafe",
                        activeforeground="#0f172a",
                        font=("Segoe UI Semibold", 9),
                        bd=0,
                        relief=tk.FLAT,
                        width=5,
                        pady=5,
                        cursor="hand2",
                    ).grid(row=row_index, column=col, padx=2, pady=2)

        render_days()

    def _wave_tokens_for_date(self, target_date: date) -> set[str]:
        year_short = str(target_date.year)[-2:]
        return {
            f"{target_date.day:02d}/{target_date.month:02d}",
            f"{target_date.day}/{target_date.month}",
            f"{target_date.day:02d}/{target_date.month}",
            f"{target_date.day}/{target_date.month:02d}",
            f"{target_date.day:02d}/{target_date.month:02d}/{target_date.year}",
            f"{target_date.day}/{target_date.month}/{target_date.year}",
            f"{target_date.day:02d}/{target_date.month:02d}/{year_short}",
            f"{target_date.day}/{target_date.month}/{year_short}",
            target_date.strftime("%Y-%m-%d"),
        }

    def _wave_contains_date(self, wave: Any, target_date: date) -> bool:
        wave_text = clean_text(wave).replace(" ", "")
        if not wave_text:
            return False
        return any(token in wave_text for token in self._wave_tokens_for_date(target_date))

    def _groupage_mail_rows_for_shipments(self, shipments: list[str]) -> list[dict[str, Any]]:
        selected_shipments = [clean_text(shipment) for shipment in shipments if clean_text(shipment)]
        if not selected_shipments:
            raise ValueError("Seleziona una o piu spedizioni groupage pianificate.")

        row_by_shipment = {
            clean_text(row.get("Shipment")): row
            for row in self.shipments
            if clean_text(row.get("Shipment"))
        }
        rows: list[dict[str, Any]] = []
        invalid_shipments: list[str] = []
        for shipment in selected_shipments:
            row = row_by_shipment.get(shipment)
            if not row or clean_text(row.get("Stato")) != STATUS_PLANNED or not self._is_brt_groupage(row):
                invalid_shipments.append(shipment)
                continue
            rows.append(row)

        if invalid_shipments:
            raise ValueError(
                "La mail Groupage si puo generare solo dalle spedizioni groupage pianificate selezionate. "
                f"Controlla: {', '.join(invalid_shipments)}"
            )
        return rows

    def _groupage_mail_columns(self) -> list[str]:
        columns = [
            column for column in self.column_order
            if column in DISPLAY_COLUMNS and column not in GROUPAGE_MAIL_EXCLUDED_COLUMNS
        ]
        for mail_column in GROUPAGE_MAIL_ONLY_COLUMNS:
            if mail_column in columns:
                continue
            if mail_column == "Grand Total Shipment Ftp Vol m3" and "Grand Total Shipment Ftp Wgt Kg" in columns:
                columns.insert(columns.index("Grand Total Shipment Ftp Wgt Kg") + 1, mail_column)
            else:
                columns.append(mail_column)
        return columns

    def _build_groupage_mail_html(self, rows: list[dict[str, Any]], target_date: date) -> str:
        columns = self._groupage_mail_columns()
        total_pallets = sum(ceil_pallets(row.get("Theoretical Pallets")) for row in rows)
        total_pallets_text = str(total_pallets)
        date_text = target_date.strftime("%d/%m/%Y")

        header_cells = "".join(
            f"<th>{html_lib.escape(COLUMN_TITLES.get(column, column))}</th>"
            for column in columns
        )
        body_rows = []
        for row in rows:
            cells = "".join(
                f"<td>{html_lib.escape(display_value(row, column))}</td>"
                for column in columns
            )
            body_rows.append(f"<tr>{cells}</tr>")

        return f"""
<html>
  <head>
    <meta charset="utf-8">
    <style>
      body {{
        font-family: Segoe UI, Arial, sans-serif;
        color: #0f172a;
        font-size: 11pt;
      }}
      .summary {{
        margin: 16px 0 12px 0;
        padding: 12px 14px;
        background: #eef6ff;
        border-left: 4px solid #2563eb;
      }}
      table {{
        border-collapse: collapse;
        width: 100%;
      }}
      th {{
        background: #0f172a;
        color: #ffffff;
        text-align: left;
        padding: 8px;
        border: 1px solid #dbe4ef;
        font-weight: 600;
      }}
      td {{
        padding: 7px 8px;
        border: 1px solid #dbe4ef;
        vertical-align: top;
      }}
      tr:nth-child(even) td {{
        background: #f8fafc;
      }}
    </style>
  </head>
  <body>
    <p>Buongiorno,<br>
    di seguito il dettaglio delle spedizioni che caricherete con il ritiro n&deg;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; per il cliente Vtech:</p>
    <div class="summary">
      <strong>Totale bancali da ritirare:</strong> {html_lib.escape(total_pallets_text)}
    </div>
    <table>
      <thead><tr>{header_cells}</tr></thead>
      <tbody>{''.join(body_rows)}</tbody>
    </table>
  </body>
</html>
""".strip()

    def _display_outlook_mail(self, subject: str, html_body: str) -> None:
        try:
            import win32com.client  # type: ignore[import-not-found]

            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.Display()
            try:
                mail.GetInspector.Activate()
            except Exception:
                pass
            return
        except Exception:
            pass

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        html_path = OUTPUT_DIR / f"groupage_mail_{timestamp}.html"
        script_path = OUTPUT_DIR / "create_outlook_mail.vbs"
        log_path = OUTPUT_DIR / "outlook_mail_error.log"
        html_path.write_text(html_body, encoding="utf-8")
        script_path.write_text(
            """
On Error Resume Next
Dim htmlPath, subject, logPath, stream, htmlBody, outlook, mail, inspector, fso, logFile
htmlPath = WScript.Arguments.Item(0)
subject = WScript.Arguments.Item(1)
logPath = WScript.Arguments.Item(2)

Set stream = CreateObject("ADODB.Stream")
stream.Type = 2
stream.Charset = "utf-8"
stream.Open
stream.LoadFromFile htmlPath
htmlBody = stream.ReadText
stream.Close

Set outlook = CreateObject("Outlook.Application")
Set mail = outlook.CreateItem(0)
mail.Subject = subject
mail.HTMLBody = htmlBody
mail.Display
Set inspector = mail.GetInspector
inspector.Activate

If Err.Number <> 0 Then
    Set fso = CreateObject("Scripting.FileSystemObject")
    Set logFile = fso.OpenTextFile(logPath, 8, True)
    logFile.WriteLine Now & " - " & Err.Number & " - " & Err.Description
    logFile.Close
End If
""".strip(),
            encoding="utf-8",
        )
        subprocess.Popen(
            [
                "wscript.exe",
                str(script_path),
                str(html_path),
                subject,
                str(log_path),
            ],
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )

    def create_groupage_outlook_mail(self) -> None:
        selected_shipments = self._selected_shipments_or_warn()
        if not selected_shipments:
            return

        try:
            rows = self._groupage_mail_rows_for_shipments(selected_shipments)
        except ValueError as exc:
            messagebox.showerror("Selezione non valida", str(exc))
            return

        mail_date = date.today()
        subject = f"Ritiri groupage V-Tech - {mail_date.strftime('%d/%m/%Y')}"
        html_body = self._build_groupage_mail_html(rows, mail_date)
        try:
            self._display_outlook_mail(subject, html_body)
        except Exception as exc:
            messagebox.showerror("Outlook non disponibile", str(exc))
            return

        total_pallets = sum(ceil_pallets(row.get("Theoretical Pallets")) for row in rows)
        self.status_var.set(
            f"Bozza Outlook creata: {len(rows)} spedizioni, {total_pallets} bancali."
        )

    def refresh_tables(self) -> None:
        for tree in self.trees.values():
            self._clear_tree(tree)
            self._update_tree_headings(tree)

        filtered_shipments = [row for row in self.shipments if self._row_matches_filters(row)]
        open_rows = [
            row for row in filtered_shipments
            if clean_text(row.get("Stato")) not in {STATUS_PLANNED, STATUS_DEPARTED, STATUS_DELIVERED}
        ]
        planned_rows = [
            row for row in filtered_shipments
            if clean_text(row.get("Stato")) == STATUS_PLANNED
        ]

        open_groupage = [row for row in open_rows if self._is_brt_groupage(row)]
        open_direct = [row for row in open_rows if not self._is_brt_groupage(row)]
        planned_groupage = [row for row in planned_rows if self._is_brt_groupage(row)]
        planned_customer = [row for row in planned_rows if not self._is_brt_groupage(row)]
        delivered_rows = [row for row in filtered_shipments if clean_text(row.get("Stato")) == STATUS_DELIVERED]

        for row in open_groupage:
            self._insert_row(self.open_groupage_tree, row)
        for row in open_direct:
            self._insert_row(self.open_direct_tree, row)
        for row in planned_groupage:
            self._insert_row(self.planned_groupage_tree, row)
        for row in planned_customer:
            self._insert_row(self.planned_direct_tree, row)
        for row in delivered_rows:
            self._insert_row(self.delivered_tree, row)

        self._fit_tree_height(self.open_groupage_tree, "open_groupage")
        self._fit_tree_height(self.open_direct_tree, "open_direct")
        self._fit_tree_height(self.planned_groupage_tree, "planned_groupage")
        self._fit_tree_height(self.planned_direct_tree, "planned_customer")
        self._fit_tree_height(self.delivered_tree, "delivered")

        visible_rows = open_rows + planned_rows
        total_margin = sum((to_float(row.get("Margine")) or 0) for row in visible_rows)
        filter_label = f"  |  Filtri attivi: {len(self.filters)}" if self.filters else ""
        self.kpi_open_var.set(str(len(open_rows)))
        self.kpi_planned_var.set(str(len(planned_rows)))
        self.kpi_customer_var.set(str(len(planned_customer)))
        self.kpi_delivered_var.set(str(len(delivered_rows)))
        self.kpi_margin_var.set(f"EUR {format_number(total_margin)}")
        self.kpi_filters_var.set(f"{len(self.filters)} attivi" if self.filters else "Nessun filtro")
        self.kpi_var.set(
            f"Da pianificare: {len(open_rows)}  |  "
            f"Pianificate: {len(planned_rows)}  |  "
            f"Consegnate: {len(delivered_rows)}  |  "
            f"BRT LTL: {len(open_groupage) + len(planned_groupage)}  |  "
            f"Customer pianificate: {len(planned_customer)}  |  "
            f"Margine stimato: EUR {format_number(total_margin)}"
            f"{filter_label}"
        )

    def _fit_tree_height(self, tree: ttk.Treeview, key: str) -> None:
        max_height = self.tree_max_heights.get(key, 12)
        row_count = len(tree.get_children())
        height = max(3, min(max_height, row_count + 1))
        tree.configure(height=height)

    def _row_matches_filters(self, row: dict[str, Any]) -> bool:
        for column, expected in self.filters.items():
            actual = display_value(row, column)
            if expected == EMPTY_FILTER_VALUE:
                if actual:
                    return False
                continue
            if actual != expected:
                return False
        return True

    def _rows_for_column_filter(self, target_column: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        for row in self.shipments:
            matches_other_filters = True
            for column, expected in self.filters.items():
                if column == target_column:
                    continue
                actual = display_value(row, column)
                if expected == EMPTY_FILTER_VALUE:
                    if actual:
                        matches_other_filters = False
                        break
                    continue
                if actual != expected:
                    matches_other_filters = False
                    break
            if matches_other_filters:
                rows.append(row)
        return rows

    def _update_tree_headings(self, tree: ttk.Treeview) -> None:
        for column in DISPLAY_COLUMNS:
            marker = " *" if column in self.filters else ""
            tree.heading(
                column,
                text=f"{COLUMN_TITLES.get(column, column)}{marker} v",
            )
        tree.configure(displaycolumns=self.column_order)

    def _clear_tree(self, tree: ttk.Treeview) -> None:
        for item in tree.get_children():
            tree.delete(item)

    def _insert_row(self, tree: ttk.Treeview, row: dict[str, Any]) -> None:
        margin = to_float(row.get("Margine"))
        base_tag = "row_odd" if len(tree.get_children()) % 2 else "row_even"
        if "non rispettato" in clean_text(row.get("SLA Contratto")).lower():
            tag = "sla"
        elif clean_text(row.get("Stato")) == STATUS_PLANNED:
            tag = "planned"
        elif margin is None:
            tag = "missing"
        elif margin < 0:
            tag = "loss"
        else:
            tag = base_tag

        tree.insert(
            "",
            tk.END,
            iid=f"{tree}_{clean_text(row.get('Shipment'))}",
            values=[display_value(row, column) for column in DISPLAY_COLUMNS],
            tags=(tag,),
        )

    def _visible_column_from_x(self, tree: ttk.Treeview, x_position: int) -> str | None:
        column_id = tree.identify_column(x_position)
        if not column_id:
            return None
        try:
            column_index = int(column_id.replace("#", "")) - 1
        except ValueError:
            return None
        if column_index < 0 or column_index >= len(self.column_order):
            return None
        return self.column_order[column_index]

    def on_tree_button_press(self, event: tk.Event) -> None:
        tree = event.widget
        if tree.identify_region(event.x, event.y) != "heading":
            self.column_drag_source = None
            self.column_drag_moved = False
            return

        self.column_drag_source = self._visible_column_from_x(tree, event.x)
        self.column_drag_start_x = event.x
        self.column_drag_moved = False

    def on_tree_drag_motion(self, event: tk.Event) -> None:
        if self.column_drag_source and abs(event.x - self.column_drag_start_x) > 8:
            self.column_drag_moved = True

    def _finish_column_header_action(self, event: tk.Event) -> bool:
        tree = event.widget
        if tree.identify_region(event.x, event.y) != "heading" and not self.column_drag_source:
            return False

        source = self.column_drag_source
        target = self._visible_column_from_x(tree, event.x)
        moved = self.column_drag_moved
        self.column_drag_source = None
        self.column_drag_moved = False

        if not source:
            return True

        if moved and target and target != source:
            self._move_column(source, target)
        elif not moved:
            self.open_column_filter(source)
        return True

    def _move_column(self, source: str, target: str) -> None:
        if source not in self.column_order or target not in self.column_order:
            return

        new_order = self.column_order[:]
        source_index = new_order.index(source)
        target_index = new_order.index(target)
        moved_column = new_order.pop(source_index)
        new_order.insert(target_index, moved_column)
        self.column_order = new_order
        self._apply_column_order()
        self._save_column_order()
        self.status_var.set(
            f"Colonna '{COLUMN_TITLES.get(source, source)}' spostata. Ordine colonne salvato."
        )

    def _trees_for_current_page(self) -> list[ttk.Treeview]:
        if self.current_page == "shipments":
            return [self.open_groupage_tree, self.open_direct_tree]
        if self.current_page == "customer":
            return [self.customer_planned_tree]
        if self.current_page == "planned_groupage":
            return [self.planned_groupage_tree]
        if self.current_page == "delivered":
            return [self.delivered_tree]
        return []

    def _shipment_from_tree_item(self, tree: ttk.Treeview, item_id: str) -> str:
        values = tree.item(item_id, "values")
        return clean_text(values[0]) if values else ""

    def _sync_selected_shipments(self) -> None:
        selected: list[str] = []
        for tree in self._trees_for_current_page():
            for item_id in tree.selection():
                shipment = self._shipment_from_tree_item(tree, item_id)
                if shipment and shipment not in selected:
                    selected.append(shipment)
        self.selected_shipments = selected
        self.selected_shipment = selected[0] if selected else ""

    def _selected_shipments_or_warn(self) -> list[str]:
        self._sync_selected_shipments()
        if not self.selected_shipments:
            messagebox.showinfo("Nessuna selezione", "Seleziona una o piu spedizioni dalla tabella.")
            return []
        return self.selected_shipments[:]

    def select_all_current_page(self) -> None:
        selected: list[str] = []
        for tree in self._trees_for_current_page():
            item_ids = tree.get_children()
            if item_ids:
                tree.selection_set(item_ids)
                tree.focus(item_ids[0])
                tree.see(item_ids[0])
            for item_id in item_ids:
                shipment = self._shipment_from_tree_item(tree, item_id)
                if shipment and shipment not in selected:
                    selected.append(shipment)

        self.selected_shipments = selected
        self.selected_shipment = selected[0] if selected else ""
        if len(selected) == 1:
            row = next((item for item in self.shipments if clean_text(item.get("Shipment")) == selected[0]), None)
            if row:
                self._show_detail(row)
        elif selected:
            self._show_selection_summary(selected)
        self.status_var.set(f"Selezionate {len(selected)} spedizioni nella pagina corrente.")

    def _show_selection_summary(self, shipments: list[str]) -> None:
        total_pallets = sum(
            (to_float(row.get("Theoretical Pallets")) or 0)
            for row in self.shipments
            if clean_text(row.get("Shipment")) in shipments
        )
        lines = [
            f"Spedizioni selezionate: {len(shipments)}",
            f"Pallet teorici selezionati: {format_number(total_pallets)}",
            f"Shipment: {', '.join(shipments[:12])}{' ...' if len(shipments) > 12 else ''}",
        ]
        self.detail_text.configure(state="normal")
        self.detail_text.delete("1.0", tk.END)
        self.detail_text.insert(tk.END, "\n".join(lines))
        self.detail_text.configure(state="disabled")

    def on_tree_select(self, event: tk.Event) -> None:
        tree = event.widget
        selection = tree.selection()
        if not selection:
            self._sync_selected_shipments()
            return

        self._sync_selected_shipments()
        if len(self.selected_shipments) > 1:
            self._show_selection_summary(self.selected_shipments)
            return

        shipment = self.selected_shipments[0] if self.selected_shipments else ""
        row = next((item for item in self.shipments if clean_text(item.get("Shipment")) == shipment), None)
        if row:
            self._show_detail(row)

    def on_tree_click(self, event: tk.Event) -> None:
        if self._finish_column_header_action(event):
            return
        self._show_extra_brt_if_clicked(event)

    def _show_extra_brt_if_clicked(self, event: tk.Event) -> None:
        tree = event.widget
        row_id = tree.identify_row(event.y)
        if tree.identify_region(event.x, event.y) != "cell":
            return
        if not row_id:
            return
        column = self._visible_column_from_x(tree, event.x)
        if not column:
            return
        values = tree.item(row_id, "values")
        shipment = values[0] if values else ""
        row = next((item for item in self.shipments if clean_text(item.get("Shipment")) == shipment), None)
        if not row:
            return
        if column == "Extra BRT Totale":
            details = clean_text(row.get("Extra BRT Applicati")) or "Nessun extra BRT calcolato."
            messagebox.showinfo(f"Extra BRT - {shipment}", details)

    def _show_detail(self, row: dict[str, Any]) -> None:
        lines = [
            f"Shipment: {clean_text(row.get('Shipment'))} | Ordine: {clean_text(row.get('Orders'))}",
            f"Tipo: {clean_text(row.get('Tipo Servizio'))} | Vettore scelto: {clean_text(row.get('Carrier Scelto'))}",
            f"Freight Code: {clean_text(row.get('Freight Code'))} | Prenotazione scarico: {clean_text(row.get('Prenotazione Scarico'))}",
            f"Late ship: {display_value(row, 'Late Ship Date')} | Early delivery: {display_value(row, 'Early Delivery Date')}",
            f"SLA: {clean_text(row.get('SLA Contratto'))} | Min. ship: {display_value(row, 'Data Ship Minima SLA')} | Min. consegna: {display_value(row, 'Prima Consegna SLA')}",
            f"Dettaglio SLA: {clean_text(row.get('Dettaglio SLA'))}",
            f"Cliente: {clean_text(row.get('Route to Customer'))} | Indirizzo: {clean_text(row.get('Route To Address'))}",
            f"Attiva: {clean_text(row.get('Tariffa Attiva Applicata'))}",
            f"Passiva: {clean_text(row.get('Tariffa Passiva Applicata'))}",
            f"Extra BRT applicati: {clean_text(row.get('Extra BRT Applicati'))}",
            f"Note: {clean_text(row.get('Note Text'))}",
        ]
        if clean_text(row.get("XML Consegna")):
            lines.append(f"XML consegna: {clean_text(row.get('XML Consegna'))}")
        self.detail_text.configure(state="normal")
        self.detail_text.delete("1.0", tk.END)
        self.detail_text.insert(tk.END, "\n".join(lines))
        self.detail_text.configure(state="disabled")

    def choose_selected_carrier(self) -> None:
        selected_shipments = self._selected_shipments_or_warn()
        if not selected_shipments:
            return

        row = next((item for item in self.shipments if clean_text(item.get("Shipment")) == selected_shipments[0]), None)
        current_carrier = clean_text(row.get("Carrier Scelto")) if row else ""

        window = tk.Toplevel(self)
        window.title("Scegli vettore")
        window.geometry("360x170")
        window.configure(background="#edf1f5")
        window.transient(self)
        window.grab_set()

        frame = ttk.Frame(window, padding=14)
        frame.pack(fill=tk.BOTH, expand=True)
        title = (
            f"Shipment {selected_shipments[0]}"
            if len(selected_shipments) == 1
            else f"{len(selected_shipments)} spedizioni selezionate"
        )
        ttk.Label(frame, text=title, font=("Segoe UI Semibold", 11)).pack(anchor="w", pady=(0, 10))
        carrier_var = tk.StringVar(value=current_carrier if current_carrier in AVAILABLE_CARRIERS else (current_carrier or "BRT"))
        combo = ttk.Combobox(frame, textvariable=carrier_var, values=AVAILABLE_CARRIERS, state="readonly")
        combo.pack(fill=tk.X)

        def apply_choice() -> None:
            selected = clean_text(carrier_var.get()).upper()
            if selected == "ALTRO":
                custom = simpledialog.askstring("Vettore", "Nome vettore:", parent=window)
                selected = clean_text(custom).upper()
                if not selected:
                    return
            try:
                for shipment in selected_shipments:
                    set_manual_carrier(
                        shipment,
                        selected,
                        Path(self.active_var.get()) if self.active_var.get() else None,
                        Path(self.brt_var.get()) if self.brt_var.get() else None,
                    )
            except Exception as exc:
                messagebox.showerror("Errore vettore", str(exc))
                return
            self.shipments = load_shipments_from_db()
            self.refresh_tables()
            self.status_var.set(f"Vettore {selected} assegnato a {len(selected_shipments)} spedizioni. Calcoli aggiornati.")
            window.destroy()

        actions = ttk.Frame(frame)
        actions.pack(fill=tk.X, pady=(14, 0))
        ttk.Button(actions, text="Applica", style="Accent.TButton", command=apply_choice).pack(side=tk.RIGHT)
        ttk.Button(actions, text="Annulla", style="Secondary.TButton", command=window.destroy).pack(side=tk.RIGHT, padx=(0, 8))

    def mark_selected_planned(self) -> None:
        selected_shipments = self._selected_shipments_or_warn()
        if not selected_shipments:
            return

        for shipment in selected_shipments:
            mark_planned(shipment)
        self.shipments = load_shipments_from_db()
        self.refresh_tables()
        self.status_var.set(f"{len(selected_shipments)} spedizioni spostate tra le pianificate.")

    def mark_selected_unplanned(self) -> None:
        selected_shipments = self._selected_shipments_or_warn()
        if not selected_shipments:
            return

        for shipment in selected_shipments:
            mark_unplanned(shipment)
        self.shipments = load_shipments_from_db()
        self.refresh_tables()
        self.status_var.set(f"{len(selected_shipments)} spedizioni rimesse tra le da pianificare.")

    def mark_selected_delivered(self) -> None:
        selected_shipments = self._selected_shipments_or_warn()
        if not selected_shipments:
            return
        xml_paths: list[Path] = []
        try:
            delivered_at = ""
            for shipment in selected_shipments:
                delivered_at, xml_path = mark_delivered(shipment)
                xml_paths.append(xml_path)
        except Exception as exc:
            messagebox.showerror("XML non creato", str(exc))
            return
        self.shipments = load_shipments_from_db()
        self.refresh_tables()
        self.status_var.set(
            f"{len(selected_shipments)} spedizioni consegnate il {delivered_at}. XML creati: {len(xml_paths)}."
        )

    def delete_selected_shipments(self) -> None:
        selected_shipments = self._selected_shipments_or_warn()
        if not selected_shipments:
            return

        count = len(selected_shipments)
        label = "questa spedizione" if count == 1 else f"queste {count} spedizioni"
        confirmed = messagebox.askyesno(
            "Elimina definitivamente",
            f"Vuoi eliminare definitivamente {label}?\n\n"
            "Non saranno piu visibili nel gestionale e non verranno reimportate automaticamente dagli stessi file.",
        )
        if not confirmed:
            return

        deleted = delete_shipments_permanently(selected_shipments)
        self.shipments = load_shipments_from_db()
        self.selected_shipments = []
        self.selected_shipment = ""
        self.refresh_tables()
        self.status_var.set(f"{deleted} spedizioni eliminate definitivamente.")

    def _save_current_settings(self) -> None:
        settings = {
            "vtech_path": self.vtech_var.get(),
            "active_rates_path": self.active_var.get(),
            "brt_passive_path": self.brt_var.get(),
            "column_order": self.column_order[:],
        }
        self.settings = settings
        save_settings(settings)


def smoke_test(args: argparse.Namespace) -> None:
    detail_rows, shipment_rows, summary = run_import(
        Path(args.vtech),
        Path(args.active_rates) if args.active_rates else None,
        Path(args.brt_passive) if args.brt_passive else None,
        save_db=False,
    )
    groupage = [row for row in shipment_rows if clean_text(row.get("Tipo Servizio")) == "Groupage - BRT LTL"]
    direct = [row for row in shipment_rows if clean_text(row.get("Tipo Servizio")) == "Diretta - KN LTL/FTL"]
    print(
        json.dumps(
            {
                "detail_rows": len(detail_rows),
                "shipment_rows": len(shipment_rows),
                "groupage_brt_ltl": len(groupage),
                "direct_kn_ltl_ftl": len(direct),
                "summary": summary,
            },
            ensure_ascii=False,
            default=serialize,
        )
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Programma locale V-Tech Trasporti.")
    parser.add_argument("--smoke-test", action="store_true", help="Esegue import e calcoli senza aprire la finestra.")
    parser.add_argument("--vtech", default=str(DEFAULT_VTECH_PATH), help="File Excel V-Tech per smoke test.")
    parser.add_argument("--active-rates", default=str(DEFAULT_ACTIVE_PATH), help="File Excel tariffe attive.")
    parser.add_argument("--brt-passive", default=str(DEFAULT_BRT_PATH), help="PDF passiva BRT.")
    args = parser.parse_args()

    if args.smoke_test:
        smoke_test(args)
        return

    app = VTechApp()
    app.mainloop()


if __name__ == "__main__":
    main()
